"""
Zero-collection extension tables.

Appends aligned zero-collection summary tables on the right side (col AA+) of
a generated hourly report. Each designation section in the base report
(Region / Division / Area / Branch) gets a matching ext table at the SAME
row as the base section's title.

Table schema: NAME col + 6 bucket cols
  Regular Demand | 1-30 DPD | 31-60 DPD | PNPA (61-90) | 1-90 DPD | NPA

Cell value = unit name if that unit has zero collection in that bucket; blank else.
Rows sorted DESC by bucket-hit count (units hitting more buckets float to top).

Color palette matches base tables:
  Title   FCE4D6 bold 14
  Header  F4B084 bold 11
  Name    FCE4D6
  Filled  E2EFDA (light green)
  Blank   FCE4D6
"""

from pathlib import Path
import pandas as pd

from services.eod_processor import get_fy_label
from services.column_matcher import find_column


# Bucket → (label, demand_col, collection_col, demand_amt_col, coll_amt_col, is_derived)
BUCKET_COLS = [
    ('Regular Demand',   'reg_demand',   'hourly_reg_collection',    'reg_demand_amt',   'reg_collection_amt',    False),
    ('1-30 DPD',         'dem_130',      'hourly_col_130',           'dem_130_amt',      'col_130_amt',           False),
    ('31-60 DPD',        'dem_3160',     'hourly_col_3160',          'dem_3160_amt',     'col_3160_amt',          False),
    ('PNPA (61-90)',     'pnpa_demand',  'hourly_pnpa_collection',   'pnpa_demand_amt',  'pnpa_collection_amt',   False),
    ('1-90 DPD',         '__1_90_dem',   '__1_90_col',               '__1_90_dem_amt',   '__1_90_col_amt',        True),
    ('NPA',              'npa_cases',    'npa_hourly_acc',           None,               'npa_hourly_amt',        False),
]

# Designation → (filter_type, title keywords, label)
DESIG_SECTIONS = [
    ('All_Region',   ['REGION - WISE', 'REGION WISE'],     'REGION'),
    ('All_Division', ['DIVISION - WISE', 'DIVISION WISE'], 'DIVISION'),
    ('All_Area',     ['AREA - WISE', 'AREA WISE'],         'AREA'),
    ('All_Branch',   ['BRANCH - WISE', 'BRANCH WISE', 'BRANCH + OFFICER', 'BRANCH / OFFICER'], 'BRANCH'),
]

PRODUCT = 'ALL'
START_COL = 27   # AA
TABLE_WIDTH = 7  # 1 label + 6 buckets


def build_branch_region_map(df):
    """Return {branch_name: region} dict from a dataframe with Region + BranchName cols."""
    reg_col = find_column(df, 'Region')
    br_col = find_column(df, 'Branch Name', 'BranchName')
    if not reg_col or not br_col:
        return {}
    hmap = df[[br_col, reg_col]].dropna().drop_duplicates()
    return dict(zip(
        hmap[br_col].astype(str).str.strip(),
        hmap[reg_col].astype(str).str.strip(),
    ))


def append_zero_collection_tables(xlsx_path, pc, target_date,
                                  branch_region=None,
                                  selected_date_str=None,
                                  selected_time_str=None):
    """Reopen the hourly xlsx and append aligned zero-collection tables per sheet."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    xlsx_path = str(xlsx_path)
    branch_region = branch_region or {}
    fy_label = get_fy_label(target_date)

    # Derive 1-90 aggregates
    pc = pc.copy()
    pc['__1_90_dem'] = pc['dem_130']        + pc['dem_3160']       + pc['pnpa_demand']
    pc['__1_90_col'] = pc['hourly_col_130'] + pc['hourly_col_3160'] + pc['hourly_pnpa_collection']

    wb = load_workbook(xlsx_path)

    # Styles (match base report palette)
    thin = Side(border_style='thin', color='808080')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(bold=True, size=14, color='000000')
    title_fill = PatternFill('solid', fgColor='FCE4D6')
    header_font = Font(bold=True, size=11, color='000000')
    header_fill = PatternFill('solid', fgColor='F4B084')
    name_font = Font(size=11, color='000000')
    name_fill = PatternFill('solid', fgColor='FCE4D6')
    filled_font = Font(size=11, color='000000')
    filled_fill = PatternFill('solid', fgColor='E2EFDA')
    blank_font = Font(size=11, color='BFBFBF')
    blank_fill = PatternFill('solid', fgColor='FCE4D6')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    BUCKET_LABELS = [b[0] for b in BUCKET_COLS]
    sheet_scopes = [('OverAll', 'OA'), (fy_label, 'FY')]

    dt_suffix = ''
    if selected_date_str:
        dt_suffix = f" — {selected_date_str}"
        if selected_time_str:
            dt_suffix += f" @ {selected_time_str}"

    def zero_hits_for(filter_type, scope):
        q = pc[
            (pc['filter_type'] == filter_type) &
            (pc['scope'] == scope) &
            (pc['product'] == PRODUCT) &
            (pc['group_value'] != 'Grand Total')
        ]
        rows = []
        for bucket_label, d_col, c_col, _da, _ca, _dv in BUCKET_COLS:
            if len(q) == 0:
                continue
            col_num = pd.to_numeric(q[c_col], errors='coerce').fillna(0)
            dem_num = pd.to_numeric(q[d_col], errors='coerce').fillna(0)
            hits = q[(col_num == 0) & (dem_num > 0)]
            for _, r in hits.iterrows():
                rows.append({
                    'name': str(r['group_value']).strip(),
                    'bucket': bucket_label,
                })
        return pd.DataFrame(rows)

    def find_section_row(ws, keywords):
        for r in range(1, 400):
            for col in (1, 2):
                v = ws.cell(row=r, column=col).value
                if v and isinstance(v, str):
                    up = v.upper()
                    for kw in keywords:
                        if kw in up:
                            return r
        return None

    def write_title(ws, row, text):
        ws.cell(row=row, column=START_COL + 1, value=text)
        ws.merge_cells(start_row=row, start_column=START_COL + 1,
                       end_row=row, end_column=START_COL + TABLE_WIDTH)
        c = ws.cell(row=row, column=START_COL + 1)
        c.font = title_font
        c.fill = title_fill
        c.alignment = center
        c.border = border

    def write_col_headers(ws, row, label_text):
        headers = [label_text] + BUCKET_LABELS
        for i, h in enumerate(headers):
            c = ws.cell(row=row, column=START_COL + 1 + i, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = border

    for sheet_name, scope in sheet_scopes:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        anchors = {}
        for ft, kws, lbl in DESIG_SECTIONS:
            anchors[ft] = find_section_row(ws, kws)

        for ft, kws, lbl in DESIG_SECTIONS:
            anchor = anchors.get(ft)
            if anchor is None:
                continue

            write_title(ws, anchor, f"ZERO COLLECTION — {lbl}-WISE{dt_suffix}")
            write_col_headers(ws, anchor + 1, f"{lbl} NAME")

            hits = zero_hits_for(ft, scope)
            data_start = anchor + 2

            if len(hits) == 0:
                c = ws.cell(row=data_start, column=START_COL + 1,
                            value=f"(no {lbl.lower()}s with zero collection)")
                ws.merge_cells(start_row=data_start, start_column=START_COL + 1,
                               end_row=data_start, end_column=START_COL + TABLE_WIDTH)
                c.font = blank_font
                c.fill = blank_fill
                c.alignment = center
                c.border = border
                continue

            counts = hits.groupby('name').size().to_dict()
            units_sorted = sorted(counts.keys(), key=lambda n: (-counts[n], n.lower()))

            r = data_start
            for unit in units_sorted:
                buckets_hit = set(hits[hits['name'] == unit]['bucket'].tolist())
                cnt = len(buckets_hit)
                lc = ws.cell(row=r, column=START_COL + 1, value=f"{unit}  ({cnt})")
                lc.font = name_font
                lc.fill = name_fill
                lc.alignment = left
                lc.border = border
                for bi, blabel in enumerate(BUCKET_LABELS):
                    is_hit = blabel in buckets_hit
                    val = unit if is_hit else ''
                    cc = ws.cell(row=r, column=START_COL + 2 + bi, value=val)
                    cc.font = filled_font if is_hit else blank_font
                    cc.fill = filled_fill if is_hit else blank_fill
                    cc.alignment = center
                    cc.border = border
                r += 1

        ws.column_dimensions[get_column_letter(START_COL + 1)].width = 32
        for bi in range(len(BUCKET_LABELS)):
            ws.column_dimensions[get_column_letter(START_COL + 2 + bi)].width = 22

    wb.save(xlsx_path)
