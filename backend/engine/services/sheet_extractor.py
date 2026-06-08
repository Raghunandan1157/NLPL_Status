"""
sheet_extractor.py - Extract every sheet from a multi-sheet workbook into individual .xlsx files.
=================================================================================================
After report_builder.py generates EOD_Report_Latest.xlsx (~166 sheets, ~10MB),
this module splits it into one .xlsx per sheet.  Email assembly then merges 2-3
small files instead of copying+stripping the giant workbook (3-10s -> <0.5s).

Provides:
  - extract_all_sheets(report_path, output_dir)  -- bulk extraction + manifest
  - get_manifest(sheets_dir)                      -- load & validate cached manifest
"""

import json
import logging
import re
import time
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Filename sanitisation
# ---------------------------------------------------------------------------
_UNSAFE_RE = re.compile(r'[<>:"/\\|?*\x00-\x1f]')


def _sanitize_filename(name: str) -> str:
    """Turn a sheet name into a safe filesystem filename (without extension).

    Replaces characters forbidden on Windows/macOS/Linux with underscores
    and strips leading/trailing whitespace and dots.
    """
    safe = _UNSAFE_RE.sub('_', name)
    safe = safe.strip(' ._')
    # Collapse multiple underscores
    safe = re.sub(r'_+', '_', safe)
    safe = safe.strip('_')
    return safe or 'sheet'


# ---------------------------------------------------------------------------
# Cell-level copy (values + styles + number formats + merged ranges)
# ---------------------------------------------------------------------------

def _copy_cell_style(src_cell, dst_cell):
    """Copy visual formatting from *src_cell* to *dst_cell*."""
    if src_cell.has_style:
        dst_cell.font = src_cell.font.copy()
        dst_cell.border = src_cell.border.copy()
        dst_cell.fill = src_cell.fill.copy()
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = src_cell.protection.copy()
        dst_cell.alignment = src_cell.alignment.copy()


def _copy_sheet(src_ws, dst_ws):
    """Deep-copy cell values, styles, merges, column widths and row heights."""

    # -- cells --
    for row in src_ws.iter_rows():
        for src_cell in row:
            if isinstance(src_cell, MergedCell):
                continue  # merged slave cells carry no data; skip
            dst_cell = dst_ws.cell(
                row=src_cell.row,
                column=src_cell.column,
                value=src_cell.value,
            )
            _copy_cell_style(src_cell, dst_cell)

    # -- merged ranges --
    for rng in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(rng))

    # -- column widths --
    for col_letter, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = dim.width
        if dim.hidden:
            dst_ws.column_dimensions[col_letter].hidden = True

    # -- row heights --
    for row_idx, dim in src_ws.row_dimensions.items():
        if dim.height is not None:
            dst_ws.row_dimensions[row_idx].height = dim.height
        if dim.hidden:
            dst_ws.row_dimensions[row_idx].hidden = True

    # -- sheet-level properties --
    dst_ws.sheet_format = src_ws.sheet_format
    if src_ws.freeze_panes:
        dst_ws.freeze_panes = src_ws.freeze_panes
    if src_ws.auto_filter.ref:
        dst_ws.auto_filter.ref = src_ws.auto_filter.ref
    dst_ws.sheet_properties = src_ws.sheet_properties


# ---------------------------------------------------------------------------
# Main extraction
# ---------------------------------------------------------------------------

def extract_all_sheets(report_path, output_dir):
    """Extract every sheet from *report_path* into individual .xlsx files.

    Parameters
    ----------
    report_path : str | Path
        Path to EOD_Report_Latest.xlsx (or any multi-sheet workbook).
    output_dir : str | Path
        Directory to write individual sheet files (created if absent).

    Returns
    -------
    dict
        Manifest with ``sheet_name -> {path, size_bytes}`` plus metadata.
    """
    report_path = Path(report_path)
    output_dir = Path(output_dir)

    if not report_path.exists():
        raise FileNotFoundError(f"Source workbook not found: {report_path}")

    t0 = time.perf_counter()

    # Clear previous extraction
    if output_dir.exists():
        for old in output_dir.glob('*.xlsx'):
            try:
                old.unlink()
            except OSError:
                pass
    output_dir.mkdir(parents=True, exist_ok=True)

    # Load source workbook (read_only=False so we can read styles; data_only=True for cached values)
    logger.info(f"Loading source workbook: {report_path.name}")
    src_wb = load_workbook(report_path, read_only=False, data_only=True)
    sheet_names = src_wb.sheetnames
    total = len(sheet_names)
    logger.info(f"Source workbook has {total} sheets")

    manifest_sheets = {}
    errors = []

    for idx, sheet_name in enumerate(sheet_names, 1):
        safe_name = _sanitize_filename(sheet_name)
        out_file = output_dir / f"{safe_name}.xlsx"

        # Handle duplicate sanitized names by appending a suffix
        if out_file.exists():
            counter = 2
            while out_file.exists():
                out_file = output_dir / f"{safe_name}_{counter}.xlsx"
                counter += 1

        try:
            dst_wb = Workbook()
            dst_ws = dst_wb.active
            dst_ws.title = sheet_name

            src_ws = src_wb[sheet_name]
            _copy_sheet(src_ws, dst_ws)

            dst_wb.save(out_file)
            dst_wb.close()

            size_bytes = out_file.stat().st_size
            size_kb = size_bytes / 1024
            manifest_sheets[sheet_name] = {
                'path': out_file.name,
                'size_bytes': size_bytes,
            }
            logger.info(f"Extracted sheet {idx}/{total}: {sheet_name} ({size_kb:.0f}KB)")

        except Exception as err:
            logger.warning(f"Failed to extract sheet {idx}/{total} '{sheet_name}': {err}")
            errors.append({'sheet': sheet_name, 'error': str(err)})
            continue

    src_wb.close()

    elapsed = time.perf_counter() - t0
    source_mtime = report_path.stat().st_mtime

    manifest = {
        'source': report_path.name,
        'source_mtime': source_mtime,
        'extracted_at': datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%S'),
        'sheet_count': len(manifest_sheets),
        'extraction_seconds': round(elapsed, 2),
        'sheets': manifest_sheets,
    }
    if errors:
        manifest['errors'] = errors

    manifest_path = output_dir / 'manifest.json'
    manifest_path.write_text(json.dumps(manifest, indent=2))

    logger.info(
        f"Sheet extraction complete: {len(manifest_sheets)}/{total} sheets "
        f"in {elapsed:.1f}s -> {output_dir}"
    )
    return manifest


# ---------------------------------------------------------------------------
# Manifest helpers
# ---------------------------------------------------------------------------

def get_manifest(sheets_dir):
    """Load and validate the sheet manifest.  Returns None if stale or missing.

    Parameters
    ----------
    sheets_dir : str | Path
        Directory containing manifest.json and individual .xlsx files
        (typically ``data/backend/sheets/``).

    Returns
    -------
    dict | None
        The manifest dict, or None if the manifest is missing, unreadable,
        or the source workbook has been updated since extraction.
    """
    sheets_dir = Path(sheets_dir)
    manifest_path = sheets_dir / 'manifest.json'
    if not manifest_path.exists():
        return None

    try:
        manifest = json.loads(manifest_path.read_text())
    except (json.JSONDecodeError, OSError) as err:
        logger.warning(f"Cannot read sheet manifest: {err}")
        return None

    # Verify source file mtime still matches
    source_path = sheets_dir.parent / manifest.get('source', '')
    if source_path.exists():
        if source_path.stat().st_mtime != manifest.get('source_mtime'):
            logger.info("Sheet manifest is stale (source mtime changed)")
            return None
    else:
        logger.info("Sheet manifest source file missing")
        return None

    return manifest
