"""
WhatsApp Web automation via Playwright.

Uses a persistent Chromium profile so the QR code scan is only needed once.
The browser stays open after sending so the user can verify delivery.

All Playwright operations run on a single dedicated thread (greenlet requirement).
Flask request threads submit tasks via a queue and wait for results.
"""
import os
import re
import sys
import time
import logging
import subprocess
import tempfile
import threading
from queue import Queue
from pathlib import Path

log = logging.getLogger(__name__)

# Persistent browser profile directory (survives restarts)
WHATSAPP_PROFILE_DIR = Path(__file__).parent.parent / 'data' / 'whatsapp-profile'
WHATSAPP_PROFILE_DIR.mkdir(parents=True, exist_ok=True)

# Temp dir for generated images
_TEMP_DIR = Path(__file__).parent.parent / 'data' / 'temp'
_TEMP_DIR.mkdir(parents=True, exist_ok=True)


# ── Indian number formatting ──────────────────────────────────────────

def _indian_format(n: int) -> str:
    """Format an integer using the Indian numbering system.
    e.g. 1618542 → '16,18,542', 16475 → '16,475'."""
    s = str(abs(n))
    if len(s) <= 3:
        return ('-' if n < 0 else '') + s
    # Last 3 digits, then groups of 2
    last3 = s[-3:]
    rest = s[:-3]
    groups = []
    while rest:
        groups.append(rest[-2:])
        rest = rest[:-2]
    groups.reverse()
    return ('-' if n < 0 else '') + ','.join(groups) + ',' + last3


def _is_pct_format(fmt: str) -> bool:
    """Check if an Excel number_format string is a percentage format."""
    if not fmt:
        return False
    return '%' in fmt


# ── Excel range → image extraction ────────────────────────────────────

def _extract_range_as_image(xlsx_path: str, cell_range: str = 'B2:Y11',
                            sheet_name: str = 'OverAll') -> str:
    """Read an Excel range from a specific sheet, render as a styled HTML
    table, screenshot it using Playwright, and return the path to the PNG.

    This runs INSIDE the Playwright worker thread so it can use the browser.
    """
    import openpyxl
    from openpyxl.utils import range_boundaries

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
        log.warning(f'Sheet "{sheet_name}" not found, using active sheet: {ws.title}')

    min_col, min_row, max_col, max_row = range_boundaries(cell_range)

    # Collect merged cell info for the range
    merged = {}  # (row, col) -> (rowspan, colspan)
    skip = set()  # cells absorbed by a merge
    for mr in ws.merged_cells.ranges:
        if (mr.min_row >= min_row and mr.max_row <= max_row and
                mr.min_col >= min_col and mr.max_col <= max_col):
            rspan = mr.max_row - mr.min_row + 1
            cspan = mr.max_col - mr.min_col + 1
            merged[(mr.min_row, mr.min_col)] = (rspan, cspan)
            for r in range(mr.min_row, mr.max_row + 1):
                for c in range(mr.min_col, mr.max_col + 1):
                    if (r, c) != (mr.min_row, mr.min_col):
                        skip.add((r, c))

    # ── Pre-scan: collect percentage values per column for conditional formatting ──
    # VBA applies icon sets + top/bottom highlight to Collection % columns.
    # Identify percentage columns and gather data-row values for ranking.
    pct_cols = {}  # col -> list of (row, value)
    grand_total_row = max_row  # last data row is typically Grand Total
    for c in range(min_col, max_col + 1):
        for r in range(min_row, max_row + 1):
            cell = ws.cell(row=r, column=c)
            nfmt = cell.number_format or 'General'
            if isinstance(cell.value, (int, float)) and _is_pct_format(nfmt):
                pct_cols.setdefault(c, []).append((r, cell.value))

    # For each percentage column, compute the conditional formatting:
    # - Grand Total value = last row's value in that column
    # - Top 3 by value → green highlight; Bottom 3 → red highlight
    # - Arrow icon: ▲ (green) if >= grand total, ▼ (red) if below
    cond_fmt = {}  # (row, col) -> {'bg': color, 'fg': color, 'arrow': str}
    for c, entries in pct_cols.items():
        if len(entries) < 2:
            continue
        # Last entry is Grand Total row — exclude from ranking
        gt_val = entries[-1][1]
        data_entries = entries[:-1]
        if not data_entries:
            continue

        # Sort by value descending for top/bottom ranking
        ranked = sorted(data_entries, key=lambda x: x[1], reverse=True)
        top3 = {row for row, _ in ranked[:3]}
        bottom3 = {row for row, _ in ranked[-3:]}

        for row, val in data_entries:
            fmt = {}
            # Top/Bottom highlight (VBA: AddTopBottomHighlight)
            if row in top3:
                fmt['bg'] = '#92D050'   # RGB(146,208,80)
                fmt['fg'] = '#006100'   # RGB(0,97,0)
                fmt['bold'] = True
            elif row in bottom3:
                fmt['bg'] = '#FFC7CE'   # RGB(255,199,206)
                fmt['fg'] = '#9C0006'   # RGB(156,0,6)
                fmt['bold'] = True

            # Icon arrow based on Grand Total comparison (VBA: AddIconSet)
            if val >= gt_val:
                fmt['arrow'] = '\u25B2'  # ▲ green arrow
                fmt['arrow_color'] = '#006100'
            else:
                fmt['arrow'] = '\u25BC'  # ▼ red arrow
                fmt['arrow_color'] = '#9C0006'

            cond_fmt[(row, c)] = fmt

    # Build HTML rows
    rows_html = []
    for r in range(min_row, max_row + 1):
        cells_html = []
        for c in range(min_col, max_col + 1):
            if (r, c) in skip:
                continue

            cell = ws.cell(row=r, column=c)
            val = cell.value
            nfmt = cell.number_format or 'General'
            is_pct = isinstance(val, (int, float)) and _is_pct_format(nfmt)
            if val is None:
                val = ''
            elif is_pct:
                # Excel stores 95.49% as 0.9549 — convert to display %
                pct_val = val * 100
                # Determine decimal places from format (e.g. '0.00%' → 2)
                m = re.search(r'\.(\d+)', nfmt)
                decimals = len(m.group(1)) if m else 2
                pct_str = f'{pct_val:.{decimals}f}%'
                # Prepend arrow icon if conditional formatting applies
                cf = cond_fmt.get((r, c))
                if cf and 'arrow' in cf:
                    val = f'{cf["arrow"]}{pct_str}'
                else:
                    val = pct_str
            elif isinstance(val, float):
                if val == int(val):
                    val = _indian_format(int(val))
                else:
                    formatted = f'{val:,.2f}'
                    val = formatted
            elif isinstance(val, int):
                val = _indian_format(val)
            else:
                val = str(val)

            # Text-based conditional formatting for Performance column
            # VBA applies: "Above Average"→green, "Below Average"→red, "N/A"→gray
            text_cf = None
            if isinstance(val, str):
                if 'Above Average' in val:
                    text_cf = {'bg': '#C6EFCE', 'fg': '#006100'}
                elif 'Below Average' in val:
                    text_cf = {'bg': '#FFC7CE', 'fg': '#9C0006'}
                elif val.strip() in ('N/A', '\u25CF N/A'):
                    text_cf = {'bg': '#D9D9D9', 'fg': '#808080'}

            # Extract styles
            styles = []

            # Apply conditional formatting colors (overrides cell fill)
            cf = cond_fmt.get((r, c))
            has_cond = False
            if text_cf:
                # Text-based conditional (Performance column)
                styles.append(f'background:{text_cf["bg"]}')
                styles.append(f'color:{text_cf["fg"]}')
                has_cond = True
            elif cf and 'bg' in cf:
                # Numeric conditional (Collection % columns)
                styles.append(f'background:{cf["bg"]}')
                if cf.get('fg'):
                    styles.append(f'color:{cf["fg"]}')
                if cf.get('bold'):
                    styles.append('font-weight:bold')
                has_cond = True
            else:
                fill = cell.fill
                if fill and fill.start_color and fill.start_color.rgb and fill.start_color.rgb not in ('00000000', None):
                    rgb = fill.start_color.rgb
                    if isinstance(rgb, str) and len(rgb) == 8:
                        rgb = rgb[2:]  # strip alpha
                    styles.append(f'background:#{rgb}')

            font = cell.font
            if font:
                if font.bold and 'font-weight:bold' not in styles:
                    styles.append('font-weight:bold')
                if font.size:
                    styles.append(f'font-size:{font.size}pt')
                if font.color and font.color.rgb and font.color.rgb not in ('00000000', None):
                    if not has_cond:  # don't override conditional color
                        frgb = font.color.rgb
                        if isinstance(frgb, str) and len(frgb) == 8:
                            frgb = frgb[2:]
                        styles.append(f'color:#{frgb}')

            align = cell.alignment
            if align and align.horizontal:
                styles.append(f'text-align:{align.horizontal}')

            style_attr = f' style="{";".join(styles)}"' if styles else ''

            # Merge attrs
            merge_attr = ''
            if (r, c) in merged:
                rspan, cspan = merged[(r, c)]
                if rspan > 1:
                    merge_attr += f' rowspan="{rspan}"'
                if cspan > 1:
                    merge_attr += f' colspan="{cspan}"'

            cells_html.append(f'<td{merge_attr}{style_attr}>{val}</td>')

        rows_html.append('<tr>' + ''.join(cells_html) + '</tr>')

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><style>
body {{ margin:0; padding:0; background:#fff; }}
table {{ border-collapse:collapse; font-family:'Calibri','Segoe UI',sans-serif; font-size:11pt; }}
td {{ border:1px solid #bbb; padding:5px 10px; white-space:nowrap; }}
</style></head><body>
<table>{''.join(rows_html)}</table>
</body></html>"""

    # Write HTML to temp file
    html_path = _TEMP_DIR / '_wa_range_preview.html'
    html_path.write_text(html, encoding='utf-8')

    wb.close()
    return str(html_path)


_screenshot_counter = 0

def _screenshot_html(page_browser, html_path: str) -> str:
    """Open the HTML in a new tab, screenshot the table, close the tab.
    Returns path to the PNG. Runs inside the Playwright worker thread."""
    global _screenshot_counter
    _screenshot_counter += 1
    png_path = str(_TEMP_DIR / f'wa_range_preview_{_screenshot_counter}.png')

    tab = page_browser.new_page()
    try:
        tab.goto(f'file://{html_path}', wait_until='load')
        time.sleep(0.5)
        table = tab.locator('table')
        table.screenshot(path=png_path)
        log.info(f'Screenshot saved: {png_path}')
    finally:
        tab.close()

    return png_path


# ── Clipboard helpers ──────────────────────────────────────────────────

def _copy_file_to_clipboard(file_path: str):
    """Copy a file to the OS clipboard so it can be pasted (Cmd/Ctrl+V).

    macOS: uses osascript.
    Windows: uses PowerShell to place the file on the clipboard.
    """
    if sys.platform == 'darwin':
        script = f'set the clipboard to (POSIX file "{file_path}")'
        subprocess.run(['osascript', '-e', script], check=True)
    elif sys.platform == 'win32':
        ps_cmd = (
            f'$f = Get-Item -LiteralPath "{file_path}";'
            f'Add-Type -AssemblyName System.Windows.Forms;'
            f'$col = New-Object System.Collections.Specialized.StringCollection;'
            f'$col.Add($f.FullName);'
            f'[System.Windows.Forms.Clipboard]::SetFileDropList($col)'
        )
        subprocess.run(['powershell', '-NoProfile', '-Command', ps_cmd], check=True)
    else:
        raise RuntimeError(f'Clipboard file copy not supported on {sys.platform}')


def _copy_image_to_clipboard(image_path: str):
    """Copy a PNG image to the OS clipboard for pasting.

    macOS: uses osascript to read image as «class PNGf».
    Windows: uses PowerShell to set clipboard image.
    """
    if sys.platform == 'darwin':
        script = (
            f'set the clipboard to '
            f'(read (POSIX file "{image_path}") as «class PNGf»)'
        )
        subprocess.run(['osascript', '-e', script], check=True)
    elif sys.platform == 'win32':
        ps_cmd = (
            f'Add-Type -AssemblyName System.Windows.Forms;'
            f'Add-Type -AssemblyName System.Drawing;'
            f'$img = [System.Drawing.Image]::FromFile("{image_path}");'
            f'[System.Windows.Forms.Clipboard]::SetImage($img);'
            f'$img.Dispose()'
        )
        subprocess.run(['powershell', '-NoProfile', '-Command', ps_cmd], check=True)
    else:
        raise RuntimeError(f'Clipboard image copy not supported on {sys.platform}')


# ── Paste + send helper (used for both image and file) ─────────────────

def _paste_and_send(page, label: str, caption: str = None, wait_after: int = 10):
    """Focus message input, Cmd/Ctrl+V, optionally type a caption, click send, wait.

    When an image is pasted, WhatsApp shows a preview dialog with its own
    caption input field. The caption is typed there (not in the main message box).
    """
    msg_selector = 'div[contenteditable="true"][data-tab="10"]'
    msg_box = page.locator(msg_selector)
    msg_box.wait_for(timeout=10_000)
    msg_box.click()
    time.sleep(1)

    paste_key = 'Meta+v' if sys.platform == 'darwin' else 'Control+v'
    page.keyboard.press(paste_key)
    log.info(f'Pasted {label} into chat.')

    # Wait for preview to appear; caption field is already focused
    log.info('Waiting 2s for preview...')
    time.sleep(2)

    # If a caption is provided, just type it — the field is auto-selected
    if caption:
        page.keyboard.type(caption, delay=20)
        log.info(f'Typed caption: {caption}')
        time.sleep(1)

    send_btn = page.locator('div[aria-label="Send"]').first
    if send_btn.count() == 0:
        send_btn = page.locator('span[data-icon="send"]').first
    send_btn.click()
    log.info(f'Clicked send ({label}).')

    log.info(f'Waiting {wait_after}s for delivery...')
    time.sleep(wait_after)


# ── Playwright worker thread ──────────────────────────────────────────

_task_queue = Queue()
_worker_thread = None
_worker_lock = threading.Lock()


def _ensure_worker():
    """Start the Playwright worker thread if not already running."""
    global _worker_thread
    with _worker_lock:
        if _worker_thread is not None and _worker_thread.is_alive():
            return
        _worker_thread = threading.Thread(target=_playwright_worker, daemon=True,
                                          name='whatsapp-playwright')
        _worker_thread.start()


def _playwright_worker():
    """
    Long-running thread that owns the Playwright browser.
    Receives (action, args, result_queue) tuples from _task_queue.
    """
    from playwright.sync_api import sync_playwright

    pw = sync_playwright().start()
    browser = pw.chromium.launch_persistent_context(
        user_data_dir=str(WHATSAPP_PROFILE_DIR),
        headless=False,
        args=['--start-maximized'],
        accept_downloads=True,
    )
    page = browser.pages[0] if browser.pages else browser.new_page()
    log.info('Playwright worker started, browser launched.')

    while True:
        action, args, result_q = _task_queue.get()

        if action == 'shutdown':
            try:
                browser.close()
                pw.stop()
            except Exception:
                pass
            result_q.put({'success': True, 'message': 'Shutdown complete.'})
            break

        try:
            if action == 'open':
                result = _do_open(page)
            elif action == 'send':
                result = _do_send(page, browser, **args)
            else:
                result = {'success': False, 'error': f'Unknown action: {action}'}
        except Exception as e:
            log.exception(f'Playwright worker error during {action}')
            result = {'success': False, 'error': str(e)}

        result_q.put(result)


def _do_open(page) -> dict:
    """Navigate to WhatsApp Web and wait for it to load."""
    log.info('Opening WhatsApp Web...')
    page.goto('https://web.whatsapp.com', wait_until='domcontentloaded')

    log.info('Waiting for WhatsApp to load (scan QR if first time)...')
    search_selector = 'div[contenteditable="true"][data-tab="3"]'
    page.wait_for_selector(search_selector, timeout=120_000)
    log.info('WhatsApp loaded and ready.')

    return {'success': True, 'message': 'WhatsApp Web is open and ready.'}


def _send_to_contact(page, contact_name, image1_path, image1_caption,
                     image2_path, image2_caption, file_path, wait_after_send):
    """Send images + file to a single contact. Assumes WhatsApp is already open."""
    search_selector = 'div[contenteditable="true"][data-tab="3"]'
    page.wait_for_selector(search_selector, timeout=30_000)
    time.sleep(1)

    # Search and open contact
    search_box = page.locator(search_selector)
    search_box.click()
    search_box.fill(contact_name)
    log.info(f'Searching for contact: {contact_name}')

    time.sleep(2)
    contact_selector = f'span[title="{contact_name}"]'
    contact = page.locator(contact_selector).first
    contact.wait_for(timeout=10_000)
    contact.click()
    log.info(f'Opened chat with {contact_name}.')
    time.sleep(1)

    # Send image 1 — Region-wise (B2:Y11)
    if image1_path:
        _copy_image_to_clipboard(image1_path)
        log.info(f'[{contact_name}] Image 1 (region) copied to clipboard.')
        _paste_and_send(page, 'image-region', caption=image1_caption, wait_after=2)

    # Send image 2 — Area-wise (B15:Y49)
    if image2_path:
        _copy_image_to_clipboard(image2_path)
        log.info(f'[{contact_name}] Image 2 (area) copied to clipboard.')
        _paste_and_send(page, 'image-area', caption=image2_caption, wait_after=2)

    # Send the Excel file
    abs_path = str(file_path.resolve())
    _copy_file_to_clipboard(abs_path)
    log.info(f'[{contact_name}] Excel file copied to clipboard.')
    _paste_and_send(page, 'file', wait_after=wait_after_send)

    log.info(f'[{contact_name}] Done.')


# Contacts CSV path
CONTACTS_CSV = Path(__file__).parent.parent / 'data' / 'whatsapp_contacts.csv'


def _load_contacts() -> list:
    """Read contact names from the CSV file."""
    if not CONTACTS_CSV.exists():
        return []
    names = []
    with open(CONTACTS_CSV, 'r', encoding='utf-8') as f:
        import csv
        reader = csv.DictReader(f)
        for row in reader:
            name = row.get('name', '').strip()
            if name:
                names.append(name)
    return names


def _do_send(page, browser, file_path: str, contact_name: str = 'Raghunandan',
             wait_after_send: int = 10) -> dict:
    """
    1. Extract B2:Y11 and B15:Y49 from OverAll sheet as PNG screenshots.
    2. For each contact in SEND_CONTACTS: search, open chat, send images + file.
    Browser stays open.
    """
    file_path = Path(file_path)
    if not file_path.exists():
        return {'success': False, 'error': f'File not found: {file_path}'}

    # ── Step 1: Extract range images + captions from OverAll ────────────
    image1_path = None
    image1_caption = None
    image2_path = None
    image2_caption = None
    is_xlsx = file_path.suffix.lower() in ('.xlsx', '.xls')
    if is_xlsx:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(file_path), data_only=True)
            sheet_name = 'OverAll'
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

            b2_val = ws['B2'].value
            if b2_val:
                image1_caption = str(b2_val).strip()

            b15_val = ws['B15'].value
            if b15_val:
                image2_caption = str(b15_val).strip()

            wb.close()
            log.info(f'Captions: B2={image1_caption}, B15={image2_caption}')

            # Image 1: Region-wise (B2:Y11)
            html1 = _extract_range_as_image(str(file_path), 'B2:Y11', 'OverAll')
            image1_path = _screenshot_html(browser, html1)
            log.info(f'Extracted image 1 (region): {image1_path}')

            # Image 2: Area-wise (B15:Y49)
            html2 = _extract_range_as_image(str(file_path), 'B15:Y49', 'OverAll')
            image2_path = _screenshot_html(browser, html2)
            log.info(f'Extracted image 2 (area): {image2_path}')

        except Exception as e:
            log.warning(f'Could not extract range images: {e}. Sending Excel only.')

    # ── Step 2: Make sure we're on WhatsApp Web ────────────────────────
    if 'web.whatsapp.com' not in (page.url or ''):
        result = _do_open(page)
        if not result['success']:
            return result

    # ── Step 3: Send to each contact ───────────────────────────────────
    sent_to = []
    contacts = _load_contacts()
    if not contacts:
        return {'success': False, 'error': 'No contacts in whatsapp_contacts.csv'}

    for name in contacts:
        try:
            _send_to_contact(page, name, image1_path, image1_caption,
                             image2_path, image2_caption, file_path,
                             wait_after_send)
            sent_to.append(name)
        except Exception as e:
            log.exception(f'Failed to send to {name}')

    log.info(f'Done. Sent to: {", ".join(sent_to)}. Browser stays open.')
    return {'success': True,
            'message': f'Sent to {", ".join(sent_to)}'}


# ── Queue helpers ─────────────────────────────────────────────────────

def _submit(action: str, args: dict = None, timeout: float = 180) -> dict:
    """Submit a task to the Playwright worker and wait for the result."""
    _ensure_worker()
    result_q = Queue()
    _task_queue.put((action, args or {}, result_q))
    try:
        return result_q.get(timeout=timeout)
    except Exception:
        return {'success': False, 'error': 'Timed out waiting for Playwright worker.'}


# ── Public API (called from Flask routes) ──────────────────────────────

def open_whatsapp() -> dict:
    """Open WhatsApp Web in Chromium. Browser stays open."""
    return _submit('open')


def send_file_to_contact(file_path: str, contact_name: str = 'Raghunandan',
                         wait_after_send: int = 10) -> dict:
    """Send image preview + Excel file to contact. Browser stays open."""
    return _submit('send', {
        'file_path': file_path,
        'contact_name': contact_name,
        'wait_after_send': wait_after_send,
    })
