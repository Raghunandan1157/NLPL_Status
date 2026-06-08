"""
Supabase Sync Blueprint
=======================
Mirrors the EOD daily_performance push (AWS Coll_Db) into the Supabase
project NLPL_Version_2, schema `Grow_With_Me`, table `_stage_daily_performance`.

The custom `Grow_With_Me` schema is not exposed over PostgREST, so writes go
through two SECURITY DEFINER wrappers in the `public` schema:
  - public.eod_check_stage_date(date)        -> int   (rows already on date)
  - public.eod_sync_stage_daily(date, jsonb) -> int   (delete + insert, count)

Auth: service_role JWT in config.SUPABASE_SERVICE_KEY (.env). If the key is
blank the endpoints return success=False with a clear message so `z eod`
can warn and continue instead of failing the pipeline.

Parsing of the EOD Employee Report mirrors Coll_Db/server/index.js
`/api/upload-daily` exactly (sheet skip rules, header detection, 25 metrics)
so AWS and Supabase receive identical numbers.
"""

import logging

from flask import Blueprint, jsonify, request
import requests as http_requests
from openpyxl import load_workbook

import config

logger = logging.getLogger(__name__)
supabase_bp = Blueprint('supabase_sync', __name__)

# Sheet (product) name -> Grow_With_Me.product_type.product_type_id.
# VVY is renamed to IL, same as the AWS Node upload.
_PRODUCT_TYPE_ID = {'IGL': 1, 'FIG': 2, 'IL': 3, 'VVY': 3}

# (db_column, normalised header text) — order defines positional fallback.
_METRIC_HEADERS = [
    ('regular_demand',         'regular demand'),
    ('regular_collection',     'regular collection'),
    ('demand_1_30',            '1-30 demand'),
    ('collection_1_30',        '1-30 collection'),
    ('demand_31_60',           '31-60 demand'),
    ('collection_31_60',       '31-60 collection'),
    ('pnpa_demand',            'pnpa demand'),
    ('pnpa_collection',        'pnpa collection'),
    ('npa_cases',              'npa cases'),
    ('npa_act_acc',            'npa act acc'),
    ('npa_act_amt',            'npa act amt'),
    ('npa_clo_acc',            'npa clo acc'),
    ('npa_clo_amt',            'npa clo amt'),
    ('on_date_demand',         'on-date demand'),
    ('on_date_collection',     'on-date collection'),
    ('regular_demand_amt',     'regular demand amt'),
    ('regular_collection_amt', 'regular collection amt'),
    ('demand_1_30_amt',        '1-30 demand amt'),
    ('collection_1_30_amt',    '1-30 collection amt'),
    ('demand_31_60_amt',       '31-60 demand amt'),
    ('collection_31_60_amt',   '31-60 collection amt'),
    ('pnpa_demand_amt',        'pnpa demand amt'),
    ('pnpa_collection_amt',    'pnpa collection amt'),
    ('on_date_demand_amt',     'on-date demand amt'),
    ('on_date_collection_amt', 'on-date collection amt'),
]
_METRIC_DB_COLS = [c for c, _ in _METRIC_HEADERS]


def _norm(h):
    return ' '.join(str(h or '').strip().lower().split())


def _num(v):
    if v is None or v == '':
        return 0
    try:
        n = float(v)
    except (TypeError, ValueError):
        return 0
    return n if n == n and n not in (float('inf'), float('-inf')) else 0


def _employee_report_path():
    """Locate the latest EOD Employee Report — same lookup as /eod/sync-daily."""
    for name in ('Employee_Report_Latest.xlsx', 'EOD_Report_Latest.xlsx'):
        p = config.BACKEND_DATA_DIR / name
        if p.exists():
            return p
    return None


def _parse_report(path):
    """Return list of row dicts: {emp_id, product_type_id, <25 metrics>}.

    Sheets ending in `_FY` and the `POS` / `EMP_POS` sheets are skipped, so
    only the per-product EOD sheets (IGL / FIG / VVY) are pushed.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    out = []
    try:
        for sheet_name in wb.sheetnames:
            if sheet_name.endswith('_FY') or sheet_name in ('POS', 'EMP_POS'):
                continue
            pt_id = _PRODUCT_TYPE_ID.get(sheet_name.strip().upper())
            if pt_id is None:
                logger.info(f"Supabase sync: skipping unknown sheet '{sheet_name}'")
                continue

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            header = rows[0]
            # emp id column
            emp_idx = 3
            norm_hdr = [_norm(h) for h in header]
            for i, h in enumerate(norm_hdr):
                if h in ('emp id', 'empid', 'emp_id'):
                    emp_idx = i
                    break
            metrics_start = emp_idx + 2

            # header-name -> column index (first match wins)
            hdr_idx = {}
            for i, h in enumerate(norm_hdr):
                if h and h not in hdr_idx:
                    hdr_idx[h] = i
            col_map = {db: hdr_idx[txt] for db, txt in _METRIC_HEADERS if txt in hdr_idx}
            use_header_map = len(col_map) >= 20

            for row in rows[1:]:
                if not row or len(row) <= emp_idx:
                    continue
                if not row[0] or not row[emp_idx]:
                    continue
                emp_id = str(row[emp_idx]).strip()
                if not emp_id:
                    continue
                rec = {'emp_id': emp_id, 'product_type_id': pt_id}
                for i, db_col in enumerate(_METRIC_DB_COLS):
                    if use_header_map and db_col in col_map:
                        ci = col_map[db_col]
                    else:
                        ci = metrics_start + i
                    rec[db_col] = _num(row[ci]) if ci < len(row) else 0
                out.append(rec)
    finally:
        wb.close()
    return out


def _rpc(fn, payload):
    """Call a Supabase public RPC. Returns (ok, result_or_errmsg)."""
    key = config.SUPABASE_SERVICE_KEY
    if not key:
        return False, 'SUPABASE_SERVICE_KEY not configured (.env)'
    url = f"{config.SUPABASE_URL.rstrip('/')}/rest/v1/rpc/{fn}"
    try:
        resp = http_requests.post(
            url,
            json=payload,
            headers={
                'apikey': key,
                'Authorization': f'Bearer {key}',
                'Content-Type': 'application/json',
            },
            timeout=120,
        )
    except http_requests.exceptions.RequestException as e:
        return False, f'Supabase not reachable: {e}'
    if resp.status_code not in (200, 204):
        return False, f'Supabase RPC {fn} failed ({resp.status_code}): {(resp.text or "")[:300]}'
    try:
        return True, resp.json()
    except ValueError:
        return True, None


@supabase_bp.route('/check-date', methods=['GET'])
def check_date():
    """How many rows already exist in _stage_daily_performance for a date.

    Query param: date=YYYY-MM-DD
    Response: {success, date, exists: bool, count: int}
    """
    date = (request.args.get('date') or '').strip()
    if not date:
        return jsonify({'success': False, 'message': 'date required (YYYY-MM-DD)'}), 400
    ok, result = _rpc('eod_check_stage_date', {'p_date': date})
    if not ok:
        return jsonify({'success': False, 'message': result}), 502
    count = int(result or 0)
    return jsonify({'success': True, 'date': date, 'exists': count > 0, 'count': count})


@supabase_bp.route('/sync-daily', methods=['POST'])
def sync_daily():
    """Push the latest EOD Employee Report into Grow_With_Me._stage_daily_performance.

    The RPC deletes existing rows for the date, then inserts — so calling this
    is always an override. `z eod` decides per-DB whether to call it at all.

    JSON body: {"date": "YYYY-MM-DD"}
    """
    data = request.get_json(silent=True) or {}
    date = (data.get('date') or '').strip()
    if not date:
        return jsonify({'success': False, 'message': 'date is required (YYYY-MM-DD)'}), 400

    if not config.SUPABASE_SERVICE_KEY:
        return jsonify({
            'success': False,
            'message': 'Supabase service key not configured — set SUPABASE_SERVICE_KEY in .env',
        }), 503

    path = _employee_report_path()
    if not path:
        return jsonify({
            'success': False,
            'message': 'No EOD Employee Report found. Run EOD processing first.',
        }), 404

    try:
        rows = _parse_report(path)
    except Exception as e:
        logger.warning(f"Supabase sync: report parse failed: {e}")
        return jsonify({'success': False, 'message': f'Report parse failed: {e}'}), 500

    if not rows:
        return jsonify({'success': False, 'message': 'No employee rows found in report.'}), 400

    ok, result = _rpc('eod_sync_stage_daily', {'p_date': date, 'p_rows': rows})
    if not ok:
        logger.warning(f"Supabase sync failed for {date}: {result}")
        return jsonify({'success': False, 'message': result}), 502

    inserted = int(result or 0)
    logger.info(f"Supabase sync: {inserted} rows pushed to _stage_daily_performance for {date}")

    # Auto-explode the staging rows into collection_period and child facts
    explode_ok, explode_result = _rpc('gw_sync_explode_stage', {'p_dates': [date]})
    if not explode_ok:
        logger.warning(f"Supabase explode failed for {date}: {explode_result}")
        return jsonify({
            'success': False,
            'message': f"Staged {inserted} rows, but explode failed: {explode_result}"
        }), 502

    # Parse exploded counts
    res_list = explode_result or [{}]
    res_map = res_list[0] if isinstance(res_list, list) and res_list else {}
    p_rows = res_map.get('period_rows', 0)
    dpd_rows = res_map.get('dpd_rows', 0)
    npa_rows = res_map.get('npa_rows', 0)
    skipped_emp = res_map.get('skipped_emp', 0)

    logger.info(
        f"Supabase sync: exploded {date} (period={p_rows} dpd={dpd_rows} npa={npa_rows} skipped_emp={skipped_emp})"
    )

    return jsonify({
        'success': True,
        'date': date,
        'inserted': inserted,
        'message': (
            f"{inserted} rows synced to Supabase for {date} (exploded: "
            f"period={p_rows} dpd={dpd_rows} npa={npa_rows} skipped_emp={skipped_emp})"
        ),
    })


# ── Hourly sync ───────────────────────────────────────────────────────
# Mirrors the EOD daily push, but for the Quick Hourly Report. There is no
# date dimension: _stage_hourly_performance is a single live snapshot, so
# every sync deletes ALL rows and re-inserts — same as AWS /api/upload-hourly
# which DROPs and recreates hourly_performance on each call.

# Quick Report carries no product split — all rows are combined.
_QUICK_HOURLY_PT_ID = 0


def _safe_num(v):
    if v is None or v == '' or v == '-':
        return 0
    try:
        n = float(v)
    except (TypeError, ValueError):
        return 0
    return n if n == n and n not in (float('inf'), float('-inf')) else 0


def _quick_report_path():
    """Locate the latest Quick Hourly Report — same file /quick writes."""
    p = config.BACKEND_DATA_DIR / 'Quick_Report_Latest.xlsx'
    return p if p.exists() else None


def _parse_quick_report(path):
    """Parse the Quick Report into hourly row dicts.

    Mirrors Coll_Db/server/index.js /api/upload-hourly Quick Report parser:
      - OverAll sheet Branch+Officer section (starts 4 rows after 'EMP ID')
      - OverAll_On-Date sheet supplies on_date_demand / on_date_collection
      - column layout (0-indexed): 0=EMP ID 2=Reg Demand 3=Reg Collection
        6=1-30 Demand 7=1-30 Collection 10=31-60 Demand 11=31-60 Collection
        14=PNPA Demand 15=PNPA Collection 18-21=1-90 DPD (skipped — derived)
        22=NPA Cases 23=NPA Act Account 24=NPA Act Amount.
        All rows carry product_type_id = 0 (combined).
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if 'OverAll' not in wb.sheetnames:
            raise ValueError("OverAll sheet not found in Quick Report")
        over_rows = list(wb['OverAll'].iter_rows(values_only=True))

        officer_start = -1
        for r, row in enumerate(over_rows):
            if row and row[0] is not None and str(row[0]).strip() == 'EMP ID':
                officer_start = r
                break
        if officer_start < 0:
            raise ValueError("Could not find Branch+Officer section (no EMP ID header)")
        data_start = officer_start + 4

        # On-Date sheet → {emp_id: {demand, collection}}
        on_date = {}
        if 'OverAll_On-Date' in wb.sheetnames:
            od_rows = list(wb['OverAll_On-Date'].iter_rows(values_only=True))
            od_start = -1
            for r, row in enumerate(od_rows):
                if row and row[0] is not None and str(row[0]).strip() == 'EMP ID':
                    od_start = r
                    break
            if od_start >= 0:
                for row in od_rows[od_start + 4:]:
                    if not row or not row[0]:
                        continue
                    emp = str(row[0]).strip()
                    if not emp or emp == 'EMP ID':
                        continue
                    on_date[emp] = {
                        'demand': _safe_num(row[2] if len(row) > 2 else 0),
                        'collection': _safe_num(row[3] if len(row) > 3 else 0),
                    }

        out = []
        for row in over_rows[data_start:]:
            if not row:
                continue
            col0 = str(row[0]).strip() if row[0] is not None else ''
            # Skip branch rows (blank col0), Grand Total, nested headers.
            if not col0 or col0 in ('Grand Total', 'EMP ID'):
                continue
            emp = col0
            od = on_date.get(emp, {'demand': 0, 'collection': 0})

            def g(i):
                return _safe_num(row[i]) if len(row) > i else 0

            out.append({
                'emp_id': emp,
                'product_type_id': _QUICK_HOURLY_PT_ID,
                'regular_demand': g(2),
                'regular_collection': g(3),
                'demand_1_30': g(6),
                'collection_1_30': g(7),
                'demand_31_60': g(10),
                'collection_31_60': g(11),
                'pnpa_demand': g(14),
                'pnpa_collection': g(15),
                'npa_cases': g(22),
                'npa_act_acc': g(23),
                'npa_act_amt': g(24),
                'npa_clo_acc': 0,
                'npa_clo_amt': 0,
                'on_date_demand': od['demand'],
                'on_date_collection': od['collection'],
                'regular_demand_amt': 0, 'regular_collection_amt': 0,
                'demand_1_30_amt': 0, 'collection_1_30_amt': 0,
                'demand_31_60_amt': 0, 'collection_31_60_amt': 0,
                'pnpa_demand_amt': 0, 'pnpa_collection_amt': 0,
                'on_date_demand_amt': 0, 'on_date_collection_amt': 0,
            })
        return out
    finally:
        wb.close()


@supabase_bp.route('/check-hourly', methods=['GET'])
def check_hourly():
    """How many rows currently sit in _stage_hourly_performance (no date arg)."""
    ok, result = _rpc('hourly_check_stage', {})
    if not ok:
        return jsonify({'success': False, 'message': result}), 502
    count = int(result or 0)
    return jsonify({'success': True, 'exists': count > 0, 'count': count})


@supabase_bp.route('/sync-hourly', methods=['POST'])
def sync_hourly():
    """Push the latest Quick Report into Grow_With_Me._stage_hourly_performance.

    The RPC deletes ALL existing rows then inserts — every call is a full
    override, mirroring AWS /api/upload-hourly. No date is involved.
    """
    if not config.SUPABASE_SERVICE_KEY:
        return jsonify({
            'success': False,
            'message': 'Supabase service key not configured — set SUPABASE_SERVICE_KEY in .env',
        }), 503

    path = _quick_report_path()
    if not path:
        return jsonify({
            'success': False,
            'message': 'No Quick Report found. Run Quick Hourly processing first.',
        }), 404

    try:
        rows = _parse_quick_report(path)
    except Exception as e:
        logger.warning(f"Supabase hourly sync: report parse failed: {e}")
        return jsonify({'success': False, 'message': f'Quick Report parse failed: {e}'}), 500

    if not rows:
        return jsonify({'success': False, 'message': 'No employee rows found in Quick Report.'}), 400

    ok, result = _rpc('hourly_sync_stage', {'p_rows': rows})
    if not ok:
        logger.warning(f"Supabase hourly sync failed: {result}")
        return jsonify({'success': False, 'message': result}), 502

    inserted = int(result or 0)
    logger.info(f"Supabase hourly sync: {inserted} rows pushed to _stage_hourly_performance")
    return jsonify({
        'success': True,
        'inserted': inserted,
        'message': f"{inserted} rows synced to Supabase (_stage_hourly_performance)",
    })


# ── Disbursement sync ─────────────────────────────────────────────────
# Mirrors the EOD daily push, for the Disbursement file. The target,
# Grow_With_Me.disbursement_daily, is date-keyed: the RPC deletes the given
# dates then inserts. The monthly disbursement file is cumulative, so the CLI
# decides override-all vs add-new-only and passes only the dates to push.
# Writes route through public wrappers (PostgREST exposes public only):
#   public.disb_check_dates(date[])        -> TABLE(disb_date, row_count)
#   public.disb_sync_daily(jsonb, date[])  -> jsonb {inserted, skipped, dates}

# Disbursement product name -> Grow_With_Me.product_type.product_type_id.
_DISB_PRODUCT_TYPE_ID = {'IGL': 1, 'FIG': 2, 'IL': 3}


def _disbursement_rows(path, filename, keep_dates):
    """Parse a disbursement file into (rows, dates) for disb_sync_daily.

    Each row is a positional JSON array matching gw_sync_disbursement_daily:
      [disb_date, branch_name, emp_code, product_type_id,
       officer_name, disb_count, disb_amount]
    """
    from blueprints.disbursement import _parse_file, _aggregate
    parsed = _parse_file(path, filename)
    agg, dates = _aggregate(parsed, keep_dates=keep_dates)
    rows = []
    for (iso, branch, emp_id, prod), v in agg.items():
        rows.append([
            iso,
            branch,
            emp_id or None,
            _DISB_PRODUCT_TYPE_ID.get(prod, 1),
            v['officer_name'] or None,
            v['cnt'],
            round(v['amt'], 2),
        ])
    return rows, sorted(dates)


@supabase_bp.route('/check-disbursement', methods=['GET'])
def check_disbursement():
    """Per-date row counts in Grow_With_Me.disbursement_daily.

    Query param: dates=YYYY-MM-DD,YYYY-MM-DD,...
    Response: {success, counts: {date: count, ...}}
    """
    raw = (request.args.get('dates') or '').strip()
    dates = [d.strip() for d in raw.split(',') if d.strip()]
    if not dates:
        return jsonify({
            'success': False,
            'message': 'dates required (comma-separated YYYY-MM-DD)',
        }), 400
    ok, result = _rpc('disb_check_dates', {'p_dates': dates})
    if not ok:
        return jsonify({'success': False, 'message': result}), 502
    counts = {}
    for r in (result or []):
        counts[str(r.get('disb_date'))] = int(r.get('row_count') or 0)
    return jsonify({'success': True, 'counts': counts})


@supabase_bp.route('/sync-disbursement', methods=['POST'])
def sync_disbursement():
    """Push disbursement file rows into Grow_With_Me.disbursement_daily.

    Form fields:
      - file:  disbursement CSV/XLSX
      - dates: comma-separated YYYY-MM-DD to push. Empty/missing -> all dates.
    The RPC deletes those dates then inserts (override for that date set).
    """
    if not config.SUPABASE_SERVICE_KEY:
        return jsonify({
            'success': False,
            'message': 'Supabase service key not configured — set SUPABASE_SERVICE_KEY in .env',
        }), 503

    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'success': False, 'message': 'No file uploaded'}), 400
    if not f.filename.lower().endswith(('.csv', '.xlsx', '.xls')):
        return jsonify({'success': False, 'message': 'Expected .csv / .xlsx file'}), 400

    raw_dates = (request.form.get('dates') or '').strip()
    keep_dates = None
    if raw_dates:
        keep_dates = {d.strip() for d in raw_dates.split(',') if d.strip()}

    from blueprints.disbursement import _save_upload
    tmp = _save_upload(f, f.filename)
    try:
        rows, dates = _disbursement_rows(tmp, f.filename, keep_dates)
        if not rows:
            return jsonify({
                'success': False,
                'message': (
                    'No Active disbursement rows match the selected dates.'
                    if keep_dates else
                    'No Active disbursement rows found in file.'
                ),
            }), 400

        ok, result = _rpc('disb_sync_daily', {'p_rows': rows, 'p_dates': dates})
        if not ok:
            logger.warning(f"Supabase disbursement sync failed: {result}")
            return jsonify({'success': False, 'message': result}), 502

        res = result or {}
        inserted = int(res.get('inserted') or 0)
        skipped = int(res.get('skipped_branch_unmatched') or 0)
        logger.info(
            f"Supabase disbursement sync: {inserted} rows across {len(dates)} "
            f"date(s), {skipped} skipped (branch unmatched)"
        )
        msg = (
            f"{inserted} rows synced to Supabase (disbursement_daily) "
            f"across {len(dates)} date(s)"
        )
        if skipped:
            msg += f" · {skipped} skipped (branch unmatched)"
        return jsonify({
            'success': True,
            'inserted': inserted,
            'skipped_branch_unmatched': skipped,
            'dates': len(dates),
            'message': msg,
        })
    finally:
        try:
            tmp.unlink()
        except Exception:
            pass
