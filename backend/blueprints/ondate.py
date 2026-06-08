"""
ON_DATE Blueprint - Monthly On-Date Report extraction and management.
Migrated from ON_DATE/server.py into the Unified app.
"""

from flask import Blueprint, send_from_directory, jsonify, request, send_file
import tempfile
import json
import gc
import logging
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from copy import copy
import calendar

from config import REPORTS_DIR, STATIC_DIR
from services.memory_manager import try_acquire_processing, release_processing, gc_checkpoint
from services.error_handler import user_error

logger = logging.getLogger(__name__)

ondate_bp = Blueprint('ondate', __name__)

# Static directory for ON_DATE frontend files
ONDATE_STATIC_DIR = STATIC_DIR / 'ondate'

# FY pattern for matching any FY_XX-YY sheet name dynamically
import re
_FY_SHEET_RE = re.compile(r'^fy_\d{2}-\d{2}$', re.IGNORECASE)
_FY_ONDATE_SHEET_RE = re.compile(r'^fy_\d{2}-\d{2}[_ ]on-?date$', re.IGNORECASE)


def _build_sheet_configs(fy_label=None):
    """Build sheet configs with dynamic FY label.

    If fy_label is None, derives it from the current date.
    The FY source_names use regex matching (handled in the sheet-matching logic)
    so any FY_XX-YY pattern will be matched automatically.
    """
    if fy_label is None:
        from services.eod_processor import get_fy_label
        fy_label = get_fy_label(None)

    return [
        {
            'source_names': ['overall'],
            'canonical_source': 'OverAll',
            'target_sheet': 'OverAll',
            'source_col_start': 2,   # B
            'source_col_end': 25,    # Y
            'num_cols': 24,
            'label_cols': 1,
            'skip_header_row': True,
        },
        {
            'source_names': ['overall_on-date', 'on-date', 'ondate', 'on date'],
            'canonical_source': 'OverAll_On-Date',
            'target_sheet': 'OverAll_On-Date',
            'source_col_start': 2,   # B
            'source_col_end': 5,     # E
            'num_cols': 4,
            'label_cols': 1,
            'skip_header_row': True,
        },
        {
            'source_names': [fy_label.lower()],
            'source_names_regex': _FY_SHEET_RE,
            'canonical_source': fy_label,
            'target_sheet': fy_label,
            'source_col_start': 2,   # B
            'source_col_end': 25,    # Y
            'num_cols': 24,
            'label_cols': 1,
            'skip_header_row': True,
        },
        {
            'source_names': [f'{fy_label.lower()}_on-date', f'{fy_label.lower()}_ondate'],
            'source_names_regex': _FY_ONDATE_SHEET_RE,
            'canonical_source': f'{fy_label}_On-Date',
            'target_sheet': f'{fy_label}_On-Date',
            'source_col_start': 2,   # B
            'source_col_end': 5,     # E
            'num_cols': 4,
            'label_cols': 1,
            'skip_header_row': True,
        },
    ]


def _sheet_matches_config(sheet_name, cfg):
    """Check if a sheet name matches a config entry (exact list or regex)."""
    if sheet_name.lower() in cfg['source_names']:
        return True
    regex = cfg.get('source_names_regex')
    if regex and regex.match(sheet_name):
        return True
    return False


def _detect_fy_label_from_sheetnames(sheetnames):
    """Auto-detect the FY label from Excel sheet names like 'FY_26-27'."""
    for name in sheetnames:
        if _FY_SHEET_RE.match(name):
            return name.upper()
    return None


def _build_sheet_configs_from_sheetnames(sheetnames):
    """Build SHEET_CONFIGS using the FY label found in the uploaded file's sheets."""
    detected = _detect_fy_label_from_sheetnames(sheetnames)
    if detected:
        return _build_sheet_configs(detected)
    return _build_sheet_configs()


# Default configs (will be rebuilt dynamically when target_date is known)
SHEET_CONFIGS = _build_sheet_configs()


# ------------------------------------------------------------------
# Static file serving
# ------------------------------------------------------------------

@ondate_bp.route('/')
def index():
    """Serve the ON_DATE index page."""
    return send_from_directory(str(ONDATE_STATIC_DIR), 'index.html')


@ondate_bp.route('/<path:filename>')
def serve_static(filename):
    """Serve ON_DATE static files (CSS, JS, etc.)."""
    return send_from_directory(str(ONDATE_STATIC_DIR), filename)


# ------------------------------------------------------------------
# Health check
# ------------------------------------------------------------------

@ondate_bp.route('/api/health')
def health():
    return jsonify({'status': 'ok', 'message': 'ON_DATE module is running'})


# ------------------------------------------------------------------
# Helper functions
# ------------------------------------------------------------------

def _populate_date_headers(ws, month_name, year, num_cols=4, label_cols=1):
    """
    Populate a worksheet with date headers for the entire month.
    Reserves `label_cols` fixed columns on the left for Region/Area/Branch names.
    Each date gets `data_cols` data columns + 1 gap column for visual separation.
    """
    data_cols = num_cols - label_cols  # Actual data columns per date block

    month_num = datetime.strptime(month_name, "%B").month
    num_days = calendar.monthrange(int(year), month_num)[1]

    # Define styles for headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    white_font = Font(bold=True, size=11, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Day-of-week row style (slightly lighter blue)
    day_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    day_font = Font(bold=True, size=10, color="FFFFFF")

    # Gap column style (light gray)
    gap_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

    # Label column style (gold)
    label_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    label_font = Font(bold=True, size=11, color="000000")

    # Day name lookup
    DAY_NAMES = ['MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN']

    # --- Reserve label column(s) on the left ---
    for lc in range(1, label_cols + 1):
        ws.column_dimensions[get_column_letter(lc)].width = 25
        # Row 1 header for label column
        lbl_cell_r1 = ws.cell(row=1, column=lc, value="Region / Branch")
        lbl_cell_r1.font = label_font
        lbl_cell_r1.fill = label_fill
        lbl_cell_r1.alignment = center_align
        lbl_cell_r1.border = thin_border
        # Row 2 (day-of-week row) — leave blank but styled
        lbl_cell_r2 = ws.cell(row=2, column=lc, value="")
        lbl_cell_r2.fill = label_fill
        lbl_cell_r2.border = thin_border

    # Add gap column between labels and date blocks
    gap_after_label = label_cols + 1
    ws.column_dimensions[get_column_letter(gap_after_label)].width = 3
    for gr in [1, 2]:
        gc = ws.cell(row=gr, column=gap_after_label, value="")
        gc.fill = gap_fill

    # Freeze panes: label column(s) + gap column stay visible when scrolling
    ws.freeze_panes = ws.cell(row=3, column=label_cols + 2)

    # Adjust column width based on how many data columns per date
    col_width = 15 if data_cols <= 4 else 12

    # Create date headers — start after label columns + gap
    col = label_cols + 2
    weekday_count = {}  # Track occurrence of each weekday (1st MON, 2nd MON, etc.)
    for day in range(1, num_days + 1):
        date_obj = datetime(int(year), month_num, day)
        date_header = date_obj.strftime("%d-%m-%Y")
        wd = date_obj.weekday()  # MON=0 .. SUN=6
        weekday_count[wd] = weekday_count.get(wd, 0) + 1
        day_name = f"{weekday_count[wd]} {DAY_NAMES[wd]}"  # e.g. "1 SAT", "2 SAT"

        # Row 1: Date header
        cell = ws.cell(row=1, column=col, value=date_header)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

        # Merge cells for the date header (spans data_cols columns)
        end_merge_col = col + data_cols - 1
        if end_merge_col <= 16384:  # Excel max columns
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=end_merge_col)

        # Row 2: Day-of-week header (e.g. MON, TUE, WED...)
        day_cell = ws.cell(row=2, column=col, value=day_name)
        day_cell.font = day_font
        day_cell.fill = day_fill
        day_cell.alignment = center_align
        day_cell.border = thin_border

        # Merge day-of-week across same columns
        if end_merge_col <= 16384:
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=end_merge_col)

        # Set column widths for data columns
        for i in range(data_cols):
            ws.column_dimensions[get_column_letter(col + i)].width = col_width

        # Add gap column after data columns (narrow, light gray)
        gap_col = col + data_cols
        ws.column_dimensions[get_column_letter(gap_col)].width = 3  # Narrow gap
        gap_cell_r1 = ws.cell(row=1, column=gap_col, value="")
        gap_cell_r1.fill = gap_fill
        gap_cell_r2 = ws.cell(row=2, column=gap_col, value="")
        gap_cell_r2.fill = gap_fill

        col += data_cols + 1  # Move to next date (data_cols + 1 gap column)

    return num_days


def create_master_report_template(report_path, month_name, year, configs=None):
    """
    Auto-create a master report template with date headers for the entire month.
    Creates separate sheets for each configured source (On-Date Report, FY On-Date Report).
    Returns the Workbook object (caller is responsible for saving).
    """
    logger.info("[ON_DATE] AUTO-CREATE: Creating new master report template...")
    if configs is None:
        configs = SHEET_CONFIGS

    wb = Workbook()

    for idx, config in enumerate(configs):
        if idx == 0:
            ws = wb.active
            ws.title = config['target_sheet']
        else:
            ws = wb.create_sheet(title=config['target_sheet'])

        num_cols = config.get('num_cols', 4)
        label_cols = config.get('label_cols', 1)
        num_days = _populate_date_headers(ws, month_name, year, num_cols=num_cols, label_cols=label_cols)
        logger.info(f"[ON_DATE] AUTO-CREATE: Sheet '{config['target_sheet']}' created with {num_days} date headers ({num_cols} cols/date, {label_cols} label cols)")

    return wb


def copy_cell_complete(source_cell, target_cell, style_cache):
    """
    Copy ALL properties from source cell to target cell.
    Preserves: value, font, fill, border, alignment, number format.
    Uses aggressive caching for performance.
    """
    # Copy value
    target_cell.value = source_cell.value

    try:
        # Create cache key based on style attributes
        cache_key = (
            str(source_cell.font.color.rgb if source_cell.font and source_cell.font.color and hasattr(source_cell.font.color, 'rgb') else None),
            str(source_cell.fill.fgColor.rgb if source_cell.fill and source_cell.fill.fgColor and hasattr(source_cell.fill.fgColor, 'rgb') else None),
            source_cell.font.bold if source_cell.font else None,
            source_cell.font.size if source_cell.font else None,
            source_cell.number_format,
        )

        if cache_key not in style_cache:
            # Cache the copied styles
            style_cache[cache_key] = {
                'font': copy(source_cell.font) if source_cell.font else None,
                'fill': copy(source_cell.fill) if source_cell.fill else None,
                'border': copy(source_cell.border) if source_cell.border else None,
                'alignment': copy(source_cell.alignment) if source_cell.alignment else None,
                'number_format': source_cell.number_format
            }

        # Apply cached styles
        cached = style_cache[cache_key]
        if cached['font']:
            target_cell.font = cached['font']
        if cached['fill']:
            target_cell.fill = cached['fill']
        if cached['border']:
            target_cell.border = cached['border']
        if cached['alignment']:
            target_cell.alignment = cached['alignment']
        if cached['number_format']:
            target_cell.number_format = cached['number_format']

    except Exception as e:
        logger.debug(f"Style copy skipped: {e}")


def _apply_performance_coloring(ws, col_idx, start_row, end_row):
    """
    Apply direct cell fills to the PERFORMANCE column based on text content.
    Replicates conditional formatting that can't be copied cell-by-cell.
    """
    # Conditional formatting colors from source
    CF_RULES = [
        {
            'text': 'Above Average',
            'fill': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            'font': Font(color="006100"),
        },
        {
            'text': 'Below Average',
            'fill': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            'font': Font(color="9C0006"),
        },
        {
            'text': 'N/A',
            'fill': PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            'font': Font(color="808080"),
        },
    ]

    applied = 0
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        val = str(cell.value) if cell.value is not None else ''
        for rule in CF_RULES:
            if rule['text'] in val:
                cell.fill = rule['fill']
                cell.font = rule['font']
                applied += 1
                break
    return applied


def _auto_fit_columns(ws, label_cols=1):
    """
    Auto-fit column widths based on actual cell content.
    Skips label columns, gap columns, and merged-cell titles.
    """
    # Identify top-left cells of horizontally-merged ranges — their content
    # spans multiple columns, so they shouldn't inflate a single column's width.
    wide_merge_tops = set()
    for mr in ws.merged_cells.ranges:
        if mr.max_col > mr.min_col:
            wide_merge_tops.add((mr.min_row, mr.min_col))

    sample_rows = min(ws.max_row or 1, 250)
    adjusted = 0

    # Start after label col(s) + gap col
    for col_idx in range(label_cols + 2, (ws.max_column or 1) + 1):
        col_letter = get_column_letter(col_idx)
        current_width = ws.column_dimensions[col_letter].width
        if current_width is not None and current_width <= 4:
            continue  # Skip gap columns

        max_len = 0
        for row_idx in range(1, sample_rows + 1):
            if (row_idx, col_idx) in wide_merge_tops:
                continue
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))

        if max_len > 0:
            new_width = min(25, max(10, max_len * 1.2 + 2))
            if abs(new_width - (current_width or 8)) > 0.5:
                ws.column_dimensions[col_letter].width = new_width
                adjusted += 1

    return adjusted


def get_merged_ranges_for_columns(worksheet, min_col, max_col):
    """
    Get merged cell ranges that intersect with specified columns.
    """
    merged = []
    try:
        for merged_range in worksheet.merged_cells.ranges:
            # Check if merged range intersects with our columns (B-E = 2-5)
            if merged_range.min_col <= max_col and merged_range.max_col >= min_col:
                merged.append({
                    'min_row': merged_range.min_row,
                    'max_row': merged_range.max_row,
                    'min_col': merged_range.min_col,
                    'max_col': merged_range.max_col
                })
    except Exception as e:
        logger.warning(f"[ON_DATE] Could not read merged cells: {e}")
    return merged


# ------------------------------------------------------------------
# Endpoints
# ------------------------------------------------------------------

@ondate_bp.route('/check-step2-report', methods=['POST'])
def check_step2_report():
    """Check if a report for the given date already exists."""
    try:
        data = request.get_json()
        selected_date = data.get('date')  # Single date in YYYY-MM-DD format

        if not selected_date:
            return jsonify({'exists': False})

        # Parse the selected date to get month/year
        try:
            date_parts = selected_date.split('-')
            year = int(date_parts[0])
            month = int(date_parts[1])
            month_names = ["January", "February", "March", "April", "May", "June",
                          "July", "August", "September", "October", "November", "December"]
            month_name = month_names[month - 1]
        except Exception:
            # Fallback to current month/year
            now = datetime.now()
            month_name = now.strftime("%B")
            year = now.year

        # Check if the REPORTS folder for this month exists
        month_folder = REPORTS_DIR / f"{month_name}_{year}"
        if not month_folder.exists():
            return jsonify({'exists': False})

        # Check for existing report file
        report_filename = f"{month_name}_{year}_Report.xlsx"
        report_path = month_folder / report_filename

        if report_path.exists():
            return jsonify({
                'exists': True,
                'filename': report_filename,
                'path': str(report_path.relative_to(REPORTS_DIR)),
                'date': selected_date
            })

        return jsonify({'exists': False})

    except Exception as e:
        err = user_error(e, context='ondate-check-step2')
        return jsonify({
            'error': err['user_message'],
            'suggestion': err['suggestion'],
        }), 500


@ondate_bp.route('/extract-ondate-report', methods=['POST'])
def extract_ondate_report():
    """
    Extract On-Date sheet with FULL STRUCTURAL INTEGRITY.
    Preserves: data, fonts, colors, borders, merged cells, number formats, row heights.
    """
    if not try_acquire_processing():
        return jsonify({
            'error': 'Server is busy processing another request. Please try again in a moment.'
        }), 503
    try:
        logger.info(f"\n{'='*60}")
        logger.info("[ON_DATE] EXTRACT - FULL STRUCTURAL INTEGRITY MODE")
        logger.info(f"{'='*60}")

        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400

        file = request.files['file']
        month = request.form.get('month', datetime.now().strftime("%B"))
        year = request.form.get('year', str(datetime.now().year))

        # Validate month and year
        valid_months = ["January", "February", "March", "April", "May", "June",
                        "July", "August", "September", "October", "November", "December"]
        if month not in valid_months:
            return jsonify({'error': f'Invalid month: {month}'}), 400
        if not year.isdigit() or len(year) != 4:
            return jsonify({'error': f'Invalid year: {year}'}), 400

        target_date_str = request.form.get('date')

        logger.info(f"[ON_DATE] File: {file.filename}")
        logger.info(f"[ON_DATE] Target: {month} {year}, Date: {target_date_str}")

        # Save uploaded file temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            file.save(tmp.name)
            tmp_path = tmp.name

        try:
            file_size_mb = Path(tmp_path).stat().st_size / (1024 * 1024)
            logger.info(f"[ON_DATE] File size: {file_size_mb:.1f} MB")

            gc_checkpoint("ondate-pre-load")

            # ============================================
            # STEP 1: Load source workbook & auto-cleanup
            # ============================================
            logger.info("[ON_DATE] 1. Loading source workbook (full mode for formatting)...")
            if file_size_mb > 50:
                logger.info("[ON_DATE]    Large file - this may take a few minutes...")

            source_wb = load_workbook(tmp_path, data_only=True)

            # Rebuild SHEET_CONFIGS dynamically based on the uploaded file's FY sheets
            active_configs = _build_sheet_configs_from_sheetnames(source_wb.sheetnames)

            # Auto-cleanup: keep only recognized On-Date sheets
            sheets_to_keep = set()
            for s in source_wb.sheetnames:
                for cfg in active_configs:
                    if _sheet_matches_config(s, cfg):
                        sheets_to_keep.add(s)
            if not sheets_to_keep:
                source_wb.close()
                return jsonify({'error': 'No recognizable On-Date sheets found'}), 400
            sheets_to_delete = [s for s in source_wb.sheetnames if s not in sheets_to_keep]
            for sheet_name in sheets_to_delete:
                del source_wb[sheet_name]
            logger.info(f"[ON_DATE] Auto-cleanup: kept {sheets_to_keep}, removed {len(sheets_to_delete)} sheet(s)")
            logger.info(f"[ON_DATE]    Loaded! Sheets: {source_wb.sheetnames}")

            # ============================================
            # STEP 2: Prepare Master Report
            # ============================================
            logger.info("[ON_DATE] 2. Preparing Master Report...")

            month_folder = REPORTS_DIR / f"{month}_{year}"
            month_folder.mkdir(parents=True, exist_ok=True)
            report_filename = f"{month}_{year}_Report.xlsx"
            report_path = month_folder / report_filename

            if report_path.exists():
                master_wb = load_workbook(report_path)
                # Check if existing report has all required sheets; regenerate if stale
                expected_sheets = {cfg['target_sheet'] for cfg in active_configs}
                existing_sheets = set(master_wb.sheetnames)
                missing = expected_sheets - existing_sheets
                if missing:
                    logger.info(f"[ON_DATE]    Existing report missing sheets {missing}, adding them...")
                    for cfg in active_configs:
                        if cfg['target_sheet'] not in existing_sheets:
                            ws = master_wb.create_sheet(title=cfg['target_sheet'])
                            _populate_date_headers(ws, month, year, num_cols=cfg.get('num_cols', 4), label_cols=cfg.get('label_cols', 1))
                            logger.info(f"[ON_DATE]    Added missing sheet '{cfg['target_sheet']}'")

                # -- Backward compatibility: detect stale layouts and regenerate --
                first_sheet_cfg = active_configs[0]
                first_ws = master_wb[first_sheet_cfg['target_sheet']] if first_sheet_cfg['target_sheet'] in master_wb.sheetnames else None
                needs_regen = False
                label_cols_cfg = first_sheet_cfg.get('label_cols', 1)

                if first_ws:
                    def _is_date_value(val):
                        """Check if a cell value looks like a date header."""
                        if isinstance(val, datetime):
                            return True
                        if isinstance(val, str):
                            try:
                                datetime.strptime(val, "%d-%m-%Y")
                                return True
                            except ValueError:
                                pass
                        return False

                    # Check 1: Very old layout — dates start at col 1 (no label column)
                    if _is_date_value(first_ws.cell(row=1, column=1).value):
                        needs_regen = True
                        logger.warning("[ON_DATE]    Old-layout detected (dates at col 1). Regenerating...")

                    # Check 2: Missing gap column — dates start at col 2 instead of col 3
                    if not needs_regen:
                        gap_col = label_cols_cfg + 1  # Should be the gap column (col B)
                        if _is_date_value(first_ws.cell(row=1, column=gap_col).value):
                            needs_regen = True
                            logger.warning("[ON_DATE]    No-gap layout detected (dates at col 2). Regenerating...")

                    # Check 3: Label contamination — labels embedded inside date blocks
                    if not needs_regen:
                        first_date_col = label_cols_cfg + 2  # After label col + gap col
                        for r in [3, 4, 5, 6, 7]:
                            date_block_val = first_ws.cell(row=r, column=first_date_col).value
                            label_val = first_ws.cell(row=r, column=1).value
                            if date_block_val and label_val and str(date_block_val) == str(label_val):
                                needs_regen = True
                                logger.warning("[ON_DATE]    Label contamination detected. Regenerating...")
                                break

                    if needs_regen:
                        master_wb.close()
                        master_wb = create_master_report_template(report_path, month, year, configs=active_configs)

                logger.info("[ON_DATE]    Loaded existing report")
            else:
                logger.info("[ON_DATE]    Auto-creating template...")
                master_wb = create_master_report_template(report_path, month, year, configs=active_configs)

            # ============================================
            # STEP 3: Validate date parameter
            # ============================================
            if not target_date_str:
                source_wb.close()
                master_wb.close()
                return jsonify({'error': 'Date is required'}), 400

            target_date_obj = datetime.strptime(target_date_str, "%Y-%m-%d")
            target_date_header = target_date_obj.strftime("%d-%m-%Y")

            # ============================================
            # STEP 4-7: Process each sheet from active_configs
            # ============================================
            processed_sheets = []

            for config in active_configs:
                # Find source sheet by matching sheetname against config
                source_sheet_name = None
                for name in source_wb.sheetnames:
                    if _sheet_matches_config(name, config):
                        source_sheet_name = name
                        break

                if not source_sheet_name:
                    logger.info(f"[ON_DATE]    Source sheet for '{config['target_sheet']}' not found, skipping")
                    continue

                source_ws = source_wb[source_sheet_name]
                total_rows = source_ws.max_row or 0
                logger.info(f"[ON_DATE]    Processing '{source_sheet_name}' -> '{config['target_sheet']}' ({total_rows:,} rows)")

                # -- Read column range from config --
                src_col_start = config.get('source_col_start', 2)
                src_col_end = config.get('source_col_end', 5)
                num_cols = config.get('num_cols', 4)
                label_cols = config.get('label_cols', 1)
                data_cols = num_cols - label_cols  # Data columns per date block

                # -- Get merged cells from source --
                source_merged = get_merged_ranges_for_columns(source_ws, min_col=src_col_start, max_col=src_col_end)
                logger.info(f"[ON_DATE]    Found {len(source_merged)} merged regions (cols {src_col_start}-{src_col_end})")

                # -- Find or create target sheet in master workbook --
                target_sheet_name = config['target_sheet']
                if target_sheet_name in master_wb.sheetnames:
                    master_ws = master_wb[target_sheet_name]
                else:
                    logger.info(f"[ON_DATE]    Creating new sheet '{target_sheet_name}' in master")
                    master_ws = master_wb.create_sheet(title=target_sheet_name)
                    _populate_date_headers(master_ws, month, year, num_cols=num_cols, label_cols=label_cols)

                # -- Find target column in row 1 --
                logger.info(f"[ON_DATE]    Finding column for: '{target_date_header}' in '{target_sheet_name}'")
                target_col_start = None
                for cell in master_ws[1]:
                    cell_val = cell.value
                    if cell_val == target_date_header:
                        target_col_start = cell.column
                        break
                    elif isinstance(cell_val, datetime):
                        if cell_val.strftime("%d-%m-%Y") == target_date_header:
                            target_col_start = cell.column
                            break

                if not target_col_start:
                    logger.warning(f"[ON_DATE]    Date {target_date_header} not found in '{target_sheet_name}', skipping")
                    continue

                logger.info(f"[ON_DATE]    Target column: {target_col_start}")

                # -- Copy data WITH ALL FORMATTING --
                # Part A: Labels (column B) go to fixed master column(s) 1..label_cols — written ONCE
                # Part B: Data (columns C+) go to the date-specific block at target_col_start
                logger.info(f"[ON_DATE]    Copying data: {label_cols} label col(s) + {data_cols} data cols -> labels=col 1, data=col {target_col_start}")
                logger.info(f"[ON_DATE]    Column mapping: source B(label)->master A, source C-{get_column_letter(src_col_end)}(data)->master col {target_col_start}")

                style_cache = {}
                row_count = 0

                # -- Detect title rows: source rows where a merge spans from label col into data cols --
                # These are section titles (e.g. "REGION - WISE COLLECTION REPORT - as on ...")
                # stored in merged cells like B2:Y2. Col B has the value, cols C+ are None.
                # Title rows go ONLY into the date block (merged), NOT into column A.
                title_rows = set()
                data_src_min_col = src_col_start + label_cols  # First data column in source
                for merge in source_merged:
                    if merge['min_col'] <= src_col_start and merge['max_col'] >= data_src_min_col:
                        for r in range(merge['min_row'], merge['max_row'] + 1):
                            title_rows.add(r)
                if title_rows:
                    logger.info(f"[ON_DATE]    Detected {len(title_rows)} title row(s) in source: {sorted(title_rows)}")

                # Check if labels already exist in master (from a previously processed date)
                # Check multiple rows to avoid false negatives from blank separator rows
                labels_already_written = any(
                    master_ws.cell(row=r, column=1).value is not None
                    for r in range(3, min(20, (master_ws.max_row or 2) + 1))
                )
                logger.info(f"[ON_DATE]    Labels already written: {labels_already_written}")

                start_row = 2 if config.get('skip_header_row', True) else 1
                for row_idx in range(start_row, total_rows + 1):
                    row_count += 1
                    target_row = row_idx + 2  # Row 1=date, Row 2=day name, data starts Row 3

                    if row_idx in title_rows:
                        # Title row: write to date block only (merged), NOT to column A
                        title_value = source_ws.cell(row=row_idx, column=src_col_start).value
                        if title_value is not None:
                            title_cell = master_ws.cell(row=target_row, column=target_col_start)
                            copy_cell_complete(
                                source_ws.cell(row=row_idx, column=src_col_start),
                                title_cell, style_cache
                            )
                            if data_cols > 1:
                                master_ws.merge_cells(
                                    start_row=target_row,
                                    start_column=target_col_start,
                                    end_row=target_row,
                                    end_column=target_col_start + data_cols - 1
                                )
                    else:
                        # Part A — Labels: copy source col B → master col 1..label_cols (only once)
                        if not labels_already_written:
                            for lc in range(label_cols):
                                source_col = src_col_start + lc        # Column B (=2) + offset
                                target_label_col = 1 + lc              # Master column 1+
                                source_cell = source_ws.cell(row=row_idx, column=source_col)
                                target_cell = master_ws.cell(row=target_row, column=target_label_col)
                                if target_cell.value is None:  # Extra safety: don't overwrite existing labels
                                    copy_cell_complete(source_cell, target_cell, style_cache)

                        # Part B — Data: copy source cols (B+label_cols)..end → date block
                        for col_offset in range(data_cols):
                            source_col = src_col_start + label_cols + col_offset  # Skip label col(s) in source
                            target_col = target_col_start + col_offset

                            source_cell = source_ws.cell(row=row_idx, column=source_col)
                            target_cell = master_ws.cell(row=target_row, column=target_col)

                            copy_cell_complete(source_cell, target_cell, style_cache)

                    if row_count % 5000 == 0:
                        pct = int(100 * row_count / total_rows) if total_rows > 0 else 0
                        logger.info(f"[ON_DATE]    ... {row_count:,} / {total_rows:,} rows ({pct}%)")

                    if row_count % 15000 == 0:
                        gc_checkpoint(f"ondate-rows-{row_count}")

                logger.info(f"[ON_DATE]    Copied {row_count:,} rows with styles (labels_written={'skipped' if labels_already_written else 'yes'})")

                # -- Post-process: Clear section titles from column A --
                # Section titles (e.g. "REGION - WISE COLLECTION REPORT - as on ...")
                # belong only in date block columns, not in the fixed label column.
                cleared_titles = 0
                for row in range(3, (master_ws.max_row or 2) + 1):
                    cell = master_ws.cell(row=row, column=1)
                    if cell.value and isinstance(cell.value, str):
                        val_upper = cell.value.strip().upper()
                        if 'REPORT' in val_upper or 'DEMAND VS COLLECTION' in val_upper or 'DEMAND' == val_upper or 'ON DATE' in val_upper:
                            cell.value = None
                            cleared_titles += 1
                if cleared_titles:
                    logger.info(f"[ON_DATE]    Cleared {cleared_titles} section title(s) from column A")

                # -- Apply merged cells to target --
                # Split into label merges (fixed col 1) and data merges (date-specific block)
                merged_count = 0
                if source_merged:
                    logger.info(f"[ON_DATE]    Applying {len(source_merged)} merged cell regions...")

                    # Boundary: source columns that are labels vs data
                    label_src_max_col = src_col_start + label_cols - 1  # e.g. col 2 (B) when label_cols=1
                    data_src_min_col = src_col_start + label_cols       # e.g. col 3 (C)

                    for merge in source_merged:
                        try:
                            # Skip title merges (full-width merges already handled during copy)
                            if merge['min_col'] <= src_col_start and merge['max_col'] >= data_src_min_col:
                                continue

                            target_min_row = merge['min_row'] + 2
                            target_max_row = merge['max_row'] + 2

                            # Case 1: Merge falls entirely within label columns
                            if merge['max_col'] <= label_src_max_col:
                                if not labels_already_written:
                                    lbl_min_col = 1 + (merge['min_col'] - src_col_start)
                                    lbl_max_col = 1 + (merge['max_col'] - src_col_start)
                                    master_ws.merge_cells(
                                        start_row=target_min_row,
                                        start_column=lbl_min_col,
                                        end_row=target_max_row,
                                        end_column=lbl_max_col
                                    )
                                    merged_count += 1

                            # Case 2: Merge falls entirely within data columns
                            elif merge['min_col'] >= data_src_min_col:
                                col_offset = merge['min_col'] - data_src_min_col
                                target_min_col = target_col_start + col_offset
                                target_max_col = target_col_start + (merge['max_col'] - data_src_min_col)

                                if target_max_col <= target_col_start + data_cols - 1:
                                    master_ws.merge_cells(
                                        start_row=target_min_row,
                                        start_column=target_min_col,
                                        end_row=target_max_row,
                                        end_column=target_max_col
                                    )
                                    merged_count += 1

                            # Case 3: Merge spans both label and data columns — split it
                            else:
                                # Label portion
                                if not labels_already_written:
                                    lbl_min_col = 1 + (merge['min_col'] - src_col_start)
                                    lbl_max_col = 1 + (label_src_max_col - src_col_start)
                                    if lbl_min_col != lbl_max_col or target_min_row != target_max_row:
                                        master_ws.merge_cells(
                                            start_row=target_min_row,
                                            start_column=lbl_min_col,
                                            end_row=target_max_row,
                                            end_column=lbl_max_col
                                        )
                                        merged_count += 1
                                # Data portion
                                target_min_col = target_col_start
                                target_max_col = target_col_start + (merge['max_col'] - data_src_min_col)
                                if target_max_col <= target_col_start + data_cols - 1:
                                    master_ws.merge_cells(
                                        start_row=target_min_row,
                                        start_column=target_min_col,
                                        end_row=target_max_row,
                                        end_column=target_max_col
                                    )
                                    merged_count += 1

                        except Exception as e:
                            logger.warning(f"[ON_DATE]    Merge failed at rows {merge['min_row']}-{merge['max_row']}: {e}")
                    logger.info(f"[ON_DATE]    Applied {merged_count} merged regions")

                # -- Copy row heights --
                for row_idx in range(1, total_rows + 1):
                    try:
                        src_height = source_ws.row_dimensions[row_idx].height
                        if src_height:
                            master_ws.row_dimensions[row_idx + 2].height = src_height
                    except Exception as e:
                        logger.warning(f"[ON_DATE]    Row height copy failed for row {row_idx}: {e}")

                # -- Apply conditional formatting as direct fills (PERFORMANCE column) --
                # Column Y (25) in source = last column of 24-col sheets
                # With label_cols split out, the perf column is at data_cols offset from target_col_start
                if num_cols == 24:
                    perf_target_col = target_col_start + data_cols - 1  # Last data column in block
                    data_start_row = 3  # Row 1=date, Row 2=day, Row 3+=data
                    data_end_row = total_rows + 2
                    cf_applied = _apply_performance_coloring(master_ws, perf_target_col, data_start_row, data_end_row)
                    logger.info(f"[ON_DATE]    Applied PERFORMANCE coloring to {cf_applied} cells")

                # -- Auto-fit column widths based on content --
                fit_count = _auto_fit_columns(master_ws, label_cols=label_cols)
                logger.info(f"[ON_DATE]    Auto-fitted {fit_count} column(s)")

                processed_sheets.append({
                    'sheet_name': config['target_sheet'],
                    'source_sheet': source_sheet_name,
                    'rows': row_count,
                    'merged_cells': merged_count
                })
                logger.info(f"[ON_DATE]    Done with '{config['target_sheet']}'")

            # ============================================
            # STEP 8: Save master workbook ONCE after all sheets
            # ============================================
            source_wb.close()
            gc_checkpoint("ondate-source-closed")

            if not processed_sheets:
                master_wb.close()
                return jsonify({'error': 'No matching source sheets found in uploaded file'}), 400

            logger.info("[ON_DATE] 8. Saving report...")
            master_wb.save(report_path)
            master_wb.close()
            gc_checkpoint("ondate-report-saved")

            total_rows_all = sum(s['rows'] for s in processed_sheets)
            total_merged_all = sum(s['merged_cells'] for s in processed_sheets)

            logger.info("[ON_DATE] SUCCESS!")
            logger.info(f"[ON_DATE]   - Sheets processed: {len(processed_sheets)}")
            logger.info(f"[ON_DATE]   - Total rows: {total_rows_all:,}")
            logger.info(f"[ON_DATE]   - Total merged cells: {total_merged_all}")
            logger.info(f"[ON_DATE]   - Saved to: {report_path}")

            return jsonify({
                'success': True,
                'message': f'Report updated for {target_date_header}',
                'path': str(report_path),
                'sheets_processed': len(processed_sheets),
                'details': processed_sheets,
                'rows': total_rows_all,
                'merged_cells': total_merged_all
            })

        finally:
            try:
                tmp_file = Path(tmp_path)
                if tmp_file.exists():
                    tmp_file.unlink()
            except PermissionError:
                logger.warning(f"[ON_DATE] Could not delete temp file {tmp_path} (still in use)")
            except OSError as e:
                logger.warning(f"[ON_DATE] Could not delete temp file {tmp_path}: {e}")

    except Exception as e:
        err = user_error(e, context='ondate-process')
        return jsonify({
            'error': err['user_message'],
            'suggestion': err['suggestion'],
        }), 500
    finally:
        gc_checkpoint("ondate-request-complete")
        release_processing()


@ondate_bp.route('/save-step2-report', methods=['POST'])
def save_step2_report():
    """Save the Step 2 report to REPORTS/[Month]_[Year]/ folder."""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400

        file = request.files['file']
        selected_date = request.form.get('date', '')  # Single date in YYYY-MM-DD format
        month = request.form.get('month', datetime.now().strftime("%B"))
        year = request.form.get('year', str(datetime.now().year))

        # Create folder structure: REPORTS/[Month]_[Year]/
        month_folder = REPORTS_DIR / f"{month}_{year}"
        month_folder.mkdir(parents=True, exist_ok=True)

        # Save the file
        filename = f"{month}_{year}_Report.xlsx"
        file_path = month_folder / filename

        file.save(file_path)
        logger.info(f"[ON_DATE] Step 2 Report saved: {file_path}")

        # Save metadata about processed date
        metadata_path = month_folder / "processed_dates.json"
        try:
            existing_metadata = {}
            if metadata_path.exists():
                with open(metadata_path, 'r') as f:
                    existing_metadata = json.load(f)

            existing_metadata['last_updated'] = datetime.now().isoformat()
            existing_metadata['date'] = selected_date
            existing_metadata['filename'] = filename

            with open(metadata_path, 'w') as f:
                json.dump(existing_metadata, f, indent=2)
        except Exception as e:
            logger.warning(f"Could not save metadata: {e}")

        return jsonify({
            'success': True,
            'filename': filename,
            'path': str(file_path.relative_to(REPORTS_DIR)),
            'folder': str(month_folder.relative_to(REPORTS_DIR))
        })

    except Exception as e:
        err = user_error(e, context='ondate-save-step2')
        return jsonify({
            'error': err['user_message'],
            'suggestion': err['suggestion'],
        }), 500


@ondate_bp.route('/get-step2-reports', methods=['GET'])
def get_step2_reports():
    """Get list of existing Step 2 reports."""
    try:
        reports = []

        if not REPORTS_DIR.exists():
            return jsonify({'reports': []})

        # Scan all month folders
        for month_folder in REPORTS_DIR.iterdir():
            if month_folder.is_dir():
                for report_file in month_folder.glob("*.xlsx"):
                    reports.append({
                        'folder': month_folder.name,
                        'filename': report_file.name,
                        'path': str(report_file.relative_to(REPORTS_DIR)),
                        'created': datetime.fromtimestamp(report_file.stat().st_mtime).isoformat()
                    })

        # Sort by creation date (newest first)
        reports.sort(key=lambda x: x['created'], reverse=True)

        return jsonify({'reports': reports})

    except Exception as e:
        err = user_error(e, context='ondate-list-step2')
        return jsonify({
            'error': err['user_message'],
            'suggestion': err['suggestion'],
        }), 500


@ondate_bp.route('/download-step2-report/<path:filepath>', methods=['GET'])
def download_step2_report(filepath):
    """Download a Step 2 report file."""
    try:
        file_path = REPORTS_DIR / filepath
        # Security: prevent path traversal
        if not file_path.resolve().is_relative_to(REPORTS_DIR.resolve()):
            return jsonify({'error': 'Invalid file path'}), 403
        if not file_path.exists():
            return jsonify({'error': 'File not found'}), 404

        return send_file(
            file_path,
            as_attachment=True,
            download_name=file_path.name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        err = user_error(e, context='ondate-download-step2')
        return jsonify({
            'error': err['user_message'],
            'suggestion': err['suggestion'],
        }), 500
