"""
EOD Blueprint - Regular Demand vs Collection
=============================================
Migrated from EOD/server.py into a Flask Blueprint.
All endpoints preserved, paths use config module.
"""

from flask import Blueprint, send_from_directory, jsonify, request, send_file, Response
from pathlib import Path
import tempfile
import queue
import logging
import json
import time
import shutil
import platform
import io
import hashlib
import requests as http_requests
from datetime import datetime
import pandas as pd
import duckdb
from openpyxl import load_workbook
import config
from services.db_manager import get_db_manager
from services import file_manager
from services import eod_processor as processor
from services.excel_reader import compute_file_hash, save_upload_to_temp, smart_read_excel
from services.memory_manager import try_acquire_processing, release_processing, gc_checkpoint, get_rss_mb
from services.hardware_profile import MEMORY_BUDGET_MB
from services.error_handler import user_error
from functools import wraps
import threading

try:
    from services.employee_processor import invalidate_merged_df_cache
except ImportError:
    logging.warning("employee_processor not available — merged DF cache invalidation disabled")
    def invalidate_merged_df_cache(date_str=None):
        pass


# ---------------------------------------------------------------------------
# API key authentication
# ---------------------------------------------------------------------------
def require_api_key(f):
    """Simple API key authentication for sensitive endpoints."""
    @wraps(f)
    def decorated(*args, **kwargs):
        api_key = config.EOD_API_KEY
        if not api_key:
            # No key configured = auth disabled (dev mode)
            return f(*args, **kwargs)
        # Check header or query param
        provided = request.headers.get('X-API-Key') or request.args.get('api_key')
        if provided != api_key:
            return jsonify({'success': False, 'message': 'Unauthorized. Provide valid X-API-Key header.'}), 401
        return f(*args, **kwargs)
    return decorated


# ---------------------------------------------------------------------------
# Blueprint
# ---------------------------------------------------------------------------
eod_bp = Blueprint('eod', __name__)

# ---------------------------------------------------------------------------
# Paths from config
# ---------------------------------------------------------------------------
STATIC_EOD_DIR = config.STATIC_DIR / 'eod'
BACKEND_DATA_DIR = config.BACKEND_DATA_DIR
ARCHIVE_DIR = config.ARCHIVE_DIR
DB_CACHE_DIR = config.DB_CACHE_DIR
TEMP_DIR = config.TEMP_DIR

# ---------------------------------------------------------------------------
# DB Manager - shared singleton (single DuckDB connection for all blueprints)
# ---------------------------------------------------------------------------
db_manager = get_db_manager()

# ---------------------------------------------------------------------------
# SSE Logging Setup
# ---------------------------------------------------------------------------
log_queue = queue.Queue()


class QueueHandler(logging.Handler):
    def emit(self, record):
        try:
            msg = self.format(record)
            log_queue.put(msg)
        except Exception:
            self.handleError(record)


# Attach handler to root logger
_queue_handler = QueueHandler()
_formatter = logging.Formatter(
    '%(asctime)s | %(levelname)s | %(message)s', datefmt='%Y-%m-%d %H:%M:%S'
)
_queue_handler.setFormatter(_formatter)
logging.getLogger().addHandler(_queue_handler)

# ---------------------------------------------------------------------------
# CSV persistence helpers
# ---------------------------------------------------------------------------
import csv

CACHE_HISTORY_CSV = BACKEND_DATA_DIR / 'cache_history.csv'
EMAIL_CONFIG_CSV = BACKEND_DATA_DIR / 'email_config.csv'


def _append_cache_history(file_type, original_name, cached_path):
    """Write/overwrite the single row for this file_type in cache_history.csv."""
    rows = {}
    if CACHE_HISTORY_CSV.exists():
        with open(CACHE_HISTORY_CSV, newline='') as f:
            for row in csv.DictReader(f):
                rows[row['type']] = row
    rows[file_type] = {
        'type': file_type,
        'name': original_name,
        'path': str(cached_path),
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }
    with open(CACHE_HISTORY_CSV, 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=['type', 'name', 'path', 'timestamp'])
        w.writeheader()
        for r in rows.values():
            w.writerow(r)


# ---------------------------------------------------------------------------
# Archive / Session helpers
# ---------------------------------------------------------------------------
_session_lock = threading.Lock()
_current_session_dir = None


def get_session_dir():
    """
    Get or create the session directory for the current processing run.
    Structure: ARCHIVE/Year/Month/Day/Session_N/
    """
    global _current_session_dir

    with _session_lock:
        if _current_session_dir is not None:
            return _current_session_dir

        now = datetime.now()
        year = str(now.year)
        month = now.strftime("%B")
        day = now.strftime("%d")

        day_dir = ARCHIVE_DIR / year / month / day
        day_dir.mkdir(parents=True, exist_ok=True)

        existing_sessions = list(day_dir.glob("Session_*"))
        session_num = len(existing_sessions) + 1

        session_dir = day_dir / f"Session_{session_num}"
        (session_dir / "Input_Files").mkdir(parents=True, exist_ok=True)
        (session_dir / "Output_Files").mkdir(parents=True, exist_ok=True)
        (session_dir / "Logs").mkdir(parents=True, exist_ok=True)

        _current_session_dir = session_dir
        logging.info(f"Created session: {session_dir.relative_to(ARCHIVE_DIR)}")
        return session_dir


def reset_session():
    """Reset session for new processing run."""
    global _current_session_dir
    _current_session_dir = None


def archive_file(file_path: Path, file_type: str, is_output: bool = False, is_monthly: bool = False):
    """
    Archive a file to the current session folder.
    Structure: ARCHIVE/Year/Month/Day/Session_N/Input_Files or Output_Files
    """
    session_dir = get_session_dir()
    now = datetime.now()
    year = str(now.year)
    month = now.strftime("%B")

    subfolder = "Output_Files" if is_output else "Input_Files"
    target_dir = session_dir / subfolder

    if is_monthly:
        month_dir = ARCHIVE_DIR / year / month
        existing = list(month_dir.rglob(f"{file_type}_*"))
        if existing:
            logging.info(f"Skipped: {file_type} already archived for {month} {year}")
            return existing[0]

        extension = file_path.suffix
        new_filename = f"{file_type}_{month}_{year}{extension}"
    else:
        timestamp = now.strftime("%H-%M-%S")
        extension = file_path.suffix
        new_filename = f"{file_type}_{timestamp}{extension}"

    target_path = target_dir / new_filename
    shutil.copy2(file_path, target_path)
    logging.info(f"Archived: {new_filename}")
    return target_path


# ---------------------------------------------------------------------------
# Cache cleanup helpers
# ---------------------------------------------------------------------------
def cleanup_old_daily_caches():
    """
    Clean up OLD daily cache files - keep only the NEWEST cache for each type.
    """
    deleted_count = 0
    total_size = 0

    if not DB_CACHE_DIR.exists():
        return {'deleted_count': 0, 'total_size': 0}

    try:
        par_caches = list(DB_CACHE_DIR.glob("daily_par_cache_*.parquet"))
        if len(par_caches) > 1:
            par_caches.sort(key=lambda f: f.stat().st_mtime, reverse=True)
            for cache_file in par_caches[1:]:
                try:
                    total_size += cache_file.stat().st_size
                    cache_file.unlink()
                    deleted_count += 1
                    logging.info(f"Deleted old PAR cache: {cache_file.name}")
                except Exception as e:
                    logging.warning(f"Could not delete {cache_file.name}: {e}")

        coll_caches = list(DB_CACHE_DIR.glob("daily_collection_cache_*.parquet"))
        if len(coll_caches) > 1:
            coll_caches.sort(key=lambda f: f.stat().st_mtime, reverse=True)
            for cache_file in coll_caches[1:]:
                try:
                    total_size += cache_file.stat().st_size
                    cache_file.unlink()
                    deleted_count += 1
                    logging.info(f"Deleted old Collection cache: {cache_file.name}")
                except Exception as e:
                    logging.warning(f"Could not delete {cache_file.name}: {e}")
    except Exception as e:
        logging.warning(f"Cache cleanup error: {e}")

    return {'deleted_count': deleted_count, 'total_size': total_size}


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@eod_bp.route('/last-cache', methods=['GET'])
def last_cache_endpoint():
    """Return info about the last cached PAR and Collection Excel copies."""
    result = {'par': None, 'collection': None, 'history': []}

    if not DB_CACHE_DIR.exists():
        return jsonify(result)

    for file_type in ['par', 'collection']:
        f = DB_CACHE_DIR / f"daily_{file_type}_last.xlsx"
        if f.exists():
            stat = f.stat()
            original_name = f.name
            meta_path = DB_CACHE_DIR / f"daily_{file_type}_last.meta.json"
            if meta_path.exists():
                try:
                    meta = json.loads(meta_path.read_text())
                    original_name = meta.get('originalName', f.name)
                except (json.JSONDecodeError, OSError) as e:
                    logging.debug(f"Could not read cache meta: {e}")
            result[file_type] = {
                'name': original_name,
                'size': stat.st_size,
                'modified': stat.st_mtime,
            }

    # Append history from CSV if available
    if CACHE_HISTORY_CSV.exists():
        try:
            with open(CACHE_HISTORY_CSV, newline='') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    result['history'].append(row)
        except (OSError, csv.Error) as e:
            logging.debug(f"Could not read cache history: {e}")

    return jsonify(result)


@eod_bp.route('/')
def index():
    """Serve EOD index.html."""
    return send_from_directory(str(STATIC_EOD_DIR), 'index.html')


@eod_bp.route('/<path:filename>')
def serve_static(filename):
    """Serve EOD static files (CSS, JS)."""
    return send_from_directory(str(STATIC_EOD_DIR), filename)


@eod_bp.route('/events')
def events():
    """SSE endpoint for live log streaming."""
    def generate():
        _keepalive_count = 0
        while True:
            try:
                message = log_queue.get(timeout=0.5)

                step = None
                done = False
                if "STEP " in message:
                    try:
                        parts = message.split("STEP ")
                        if len(parts) > 1:
                            val = parts[1].split(":")[0].strip()
                            step = int(val)
                    except (IndexError, ValueError):
                        pass
                if "ALL STEPS COMPLETED" in message:
                    done = True

                rss = get_rss_mb()
                mem = {'rss_mb': round(rss), 'budget_mb': MEMORY_BUDGET_MB} if rss > 0 else {}
                data = {'log': message, 'step': step, **mem}
                if done:
                    data['done'] = True
                yield f"data: {json.dumps(data)}\n\n"
            except queue.Empty:
                yield ": keep-alive\n\n"
                _keepalive_count += 1
                if _keepalive_count > 3600:
                    break
            except GeneratorExit:
                break

    return Response(generate(), mimetype='text/event-stream')


@eod_bp.route('/fix-sheets', methods=['POST'])
def fix_sheets_endpoint():
    """
    Fixes sheet names in uploaded files by renaming the first sheet to 'Sheet1'.
    Returns a report of what was changed.
    """
    try:
        from openpyxl import load_workbook
        changes = []

        if 'par' in request.files:
            par = request.files['par']
            par_tmp = save_upload_to_temp(par, prefix="fix_par_")
            try:
                wb = load_workbook(str(par_tmp))
                first_sheet = wb.sheetnames[0]

                if first_sheet != 'Sheet1':
                    old_name = first_sheet
                    wb[first_sheet].title = 'Sheet1'
                    temp_par = TEMP_DIR / 'fixed_par.xlsx'
                    wb.save(temp_par)
                    wb.close()
                    changes.append({
                        'file': 'PAR',
                        'oldName': old_name,
                        'newName': 'Sheet1',
                        'path': str(temp_par)
                    })
                    logging.info(f"Fixed PAR: '{old_name}' -> 'Sheet1'")
                else:
                    wb.close()
                    logging.info("PAR already has 'Sheet1'")
            finally:
                par_tmp.unlink(missing_ok=True)

        if 'collection' in request.files:
            collection = request.files['collection']
            coll_tmp = save_upload_to_temp(collection, prefix="fix_coll_")
            try:
                wb = load_workbook(str(coll_tmp))
                first_sheet = wb.sheetnames[0]

                if first_sheet != 'Sheet1':
                    old_name = first_sheet
                    wb[first_sheet].title = 'Sheet1'
                    temp_coll = TEMP_DIR / 'fixed_collection.xlsx'
                    wb.save(temp_coll)
                    wb.close()
                    changes.append({
                        'file': 'Collection',
                        'oldName': old_name,
                        'newName': 'Sheet1',
                        'path': str(temp_coll)
                    })
                    logging.info(f"Fixed Collection: '{old_name}' -> 'Sheet1'")
                else:
                    wb.close()
                    logging.info("Collection already has 'Sheet1'")
            finally:
                coll_tmp.unlink(missing_ok=True)

        if changes:
            return jsonify({
                'success': True,
                'fixed': True,
                'changes': changes,
                'message': f"Fixed {len(changes)} file(s). Sheets renamed to 'Sheet1'."
            })
        else:
            return jsonify({
                'success': True,
                'fixed': False,
                'changes': [],
                'message': "No fixes needed. All sheets are already named 'Sheet1'."
            })

    except Exception as e:
        err = user_error(e, context='eod-fix-sheets')
        return jsonify({
            'error': err['user_message'],
            'suggestion': err['suggestion'],
        }), 500


# ---------------------------------------------------------------------------
# Google Drive endpoints (no API key — public folders only)
# ---------------------------------------------------------------------------

@eod_bp.route('/gdrive-config', methods=['GET'])
def gdrive_config_get():
    """Return the saved GDrive folder URL (persisted across restarts)."""
    from services.gdrive import load_gdrive_config
    cfg = load_gdrive_config(config.EOD_GDRIVE_CONFIG_PATH)
    return jsonify({'success': True, 'folder_url': cfg.get('folder_url', '')})


@eod_bp.route('/gdrive-config', methods=['POST'])
def gdrive_config_save():
    """Save the GDrive folder URL to disk."""
    from services.gdrive import save_gdrive_config
    body = request.get_json(silent=True) or {}
    folder_url = body.get('folder_url', '').strip()
    save_gdrive_config(config.EOD_GDRIVE_CONFIG_PATH, {'folder_url': folder_url})
    return jsonify({'success': True})


@eod_bp.route('/gdrive-scan', methods=['POST'])
def gdrive_scan_endpoint():
    """Step 1: Scan a public Google Drive folder for PAR and Collection files."""
    try:
        from services.gdrive import (
            parse_folder_id, list_folder_files_public,
            find_required_files, save_gdrive_config,
        )

        body = request.get_json(silent=True) or {}
        folder_url = body.get('folder_url', '').strip()
        if not folder_url:
            return jsonify({'success': False, 'message': 'No folder URL provided.'}), 400

        folder_id = parse_folder_id(folder_url)
        if not folder_id:
            return jsonify({'success': False, 'message': 'Could not parse folder ID from URL.'}), 400

        # Persist the folder URL so it survives restarts
        save_gdrive_config(config.EOD_GDRIVE_CONFIG_PATH, {'folder_url': folder_url})

        logging.info(f"GDrive scan: folder_id={folder_id}")
        all_files = list_folder_files_public(folder_id)
        matched = find_required_files(all_files)

        missing = [k for k in ('par', 'collection') if not matched[k]]
        if missing:
            available_names = [f['name'] for f in all_files]
            return jsonify({
                'success': False,
                'message': f"Missing required files: {', '.join(missing)}. Folder must contain files starting with 'par' and 'collection'.",
                'missing': missing,
                'available_files': available_names,
            })

        return jsonify({
            'success': True,
            'par': matched['par'],           # list — may have >1 match
            'collection': matched['collection'],  # list — may have >1 match
            'total_files': len(all_files),
        })

    except Exception as e:
        err = user_error(e, context='eod-gdrive-scan')
        return jsonify({'success': False, 'message': err['user_message'], 'suggestion': err['suggestion']}), 500


@eod_bp.route('/gdrive-download', methods=['POST'])
def gdrive_download_endpoint():
    """Step 2: Download chosen files from Google Drive, cache them, and stage for processing."""
    try:
        from services.gdrive import download_file as gdrive_download

        body = request.get_json(silent=True) or {}
        par_id = body.get('par_id')
        par_name = body.get('par_name', 'par.xlsx')
        collection_id = body.get('collection_id')
        collection_name = body.get('collection_name', 'collection.xlsx')

        if not par_id or not collection_id:
            return jsonify({'success': False, 'message': 'Missing par_id or collection_id.'}), 400

        gdrive_dir = config.GDRIVE_DOWNLOAD_DIR
        gdrive_dir.mkdir(parents=True, exist_ok=True)

        # Clear log queue before download for clean SSE stream
        with log_queue.mutex:
            log_queue.queue.clear()

        def _make_progress_fn(label):
            """Return a callback that pushes structured download progress into the SSE log queue."""
            def _progress(downloaded, total):
                dl_mb = downloaded / (1024 * 1024)
                if total:
                    total_mb = total / (1024 * 1024)
                    pct = min(int(downloaded * 100 / total), 100)
                    msg = json.dumps({
                        'gdrive_progress': True,
                        'label': label,
                        'downloaded': round(dl_mb, 1),
                        'total': round(total_mb, 1),
                        'pct': pct,
                    })
                else:
                    msg = json.dumps({
                        'gdrive_progress': True,
                        'label': label,
                        'downloaded': round(dl_mb, 1),
                        'total': None,
                        'pct': None,
                    })
                log_queue.put(msg)
            return _progress

        # Download both files with progress reporting
        logging.info(f"GDrive: downloading PAR — {par_name}")
        par_path = gdrive_download(par_id, gdrive_dir / par_name,
                                   progress_fn=_make_progress_fn('par'))
        logging.info(f"GDrive: PAR downloaded ({par_path.stat().st_size / (1024*1024):.1f} MB)")

        logging.info(f"GDrive: downloading Collection — {collection_name}")
        collection_path = gdrive_download(collection_id, gdrive_dir / collection_name,
                                          progress_fn=_make_progress_fn('collection'))
        logging.info(f"GDrive: Collection downloaded ({collection_path.stat().st_size / (1024*1024):.1f} MB)")

        # Cache both files (same logic as /cache-file endpoint)
        logging.info("GDrive: caching files to Parquet...")
        par_cached = _gdrive_cache_file(par_path, 'par', par_name)
        coll_cached = _gdrive_cache_file(collection_path, 'collection', collection_name)
        logging.info("GDrive: caching complete")

        # Signal download done
        log_queue.put(json.dumps({'gdrive_done': True}))

        return jsonify({
            'success': True,
            'par_path': str(par_path),
            'collection_path': str(collection_path),
            'par_cached': par_cached,
            'collection_cached': coll_cached,
        })

    except Exception as e:
        err = user_error(e, context='eod-gdrive-download')
        return jsonify({'success': False, 'message': err['user_message'], 'suggestion': err['suggestion']}), 500


def _gdrive_cache_file(file_path: Path, file_type: str, original_name: str) -> bool:
    """Cache a downloaded GDrive file to parquet + save as last-cache Excel copy."""
    try:
        file_hash = compute_file_hash(file_path)
        DB_CACHE_DIR.mkdir(parents=True, exist_ok=True)

        if file_type == 'par':
            cache_pattern = "daily_par_cache_*.parquet"
            cache_path = DB_CACHE_DIR / f"daily_par_cache_{file_hash}.parquet"
        else:
            cache_pattern = "daily_collection_cache_*.parquet"
            cache_path = DB_CACHE_DIR / f"daily_collection_cache_{file_hash}.parquet"

        # Delete old caches for this type
        for old in DB_CACHE_DIR.glob(cache_pattern):
            try:
                old.unlink()
            except OSError:
                pass

        # Read Excel -> Parquet
        df = smart_read_excel(file_path)
        if file_type == 'par':
            cols = ['AccountID', 'Days Group', 'Days group', 'DaysGroup', 'DPD Group', 'DPD Days', 'DPDDays']
        else:
            cols = ['AccountID', 'CollectionTotal', 'Trxdate', 'ReverseTotal']
        available = [c for c in cols if c in df.columns]
        if available:
            df = df[available]

        # Fix mixed-type columns
        numeric_cols = {'CollectionTotal', 'ReverseTotal'}
        for col in df.columns:
            if col in numeric_cols and df[col].dtype == 'object':
                df[col] = pd.to_numeric(df[col], errors='coerce')

        df.to_parquet(cache_path, index=False)

        # Save full Excel copy for "Last Cache"
        full_copy = DB_CACHE_DIR / f"daily_{file_type}_last.xlsx"
        shutil.copy(file_path, full_copy)
        meta_path = DB_CACHE_DIR / f"daily_{file_type}_last.meta.json"
        meta_path.write_text(json.dumps({'originalName': original_name}))
        _append_cache_history(file_type, original_name, full_copy)

        logging.info(f"GDrive: cached {file_type.upper()} as {cache_path.name}")
        return True
    except Exception as e:
        logging.warning(f"GDrive cache failed for {file_type}: {e}")
        return False


@eod_bp.route('/gdrive-process', methods=['POST'])
@require_api_key
def gdrive_process_endpoint():
    """Step 3: Process staged GDrive files (uses last-cache path, same as useLastCache=true)."""
    from werkzeug.datastructures import ImmutableMultiDict

    form_data = dict(request.form)
    form_data['useLastCache'] = 'true'
    form_data['useBackendDemand'] = 'true'

    original_form = request.form
    request.form = ImmutableMultiDict(form_data)
    try:
        return process_files_endpoint()
    finally:
        request.form = original_form


@eod_bp.route('/cache-file', methods=['POST'])
def cache_file_endpoint():
    """
    Convert uploaded Excel file to Parquet cache immediately at upload time.
    This makes the actual processing much faster since files are pre-cached.
    """
    try:
        import pandas as pd
        import hashlib

        file_type = request.form.get('type')

        if file_type not in ['par', 'collection']:
            return jsonify({'error': 'Invalid file type'}), 400

        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400

        file = request.files['file']
        tmp_path = save_upload_to_temp(file, prefix=f"cache_{file_type}_")

        try:
            # Calculate hash from disk
            file_hash = compute_file_hash(tmp_path)

            # Setup cache directory
            DB_CACHE_DIR.mkdir(parents=True, exist_ok=True)

            if file_type == 'par':
                cache_pattern = "daily_par_cache_*.parquet"
                cache_path = DB_CACHE_DIR / f"daily_par_cache_{file_hash}.parquet"
                cols_to_use = ['AccountID', 'Days Group', 'Days group', 'DaysGroup', 'DPD Group', 'DPD Days', 'DPDDays']
            else:
                cache_pattern = "daily_collection_cache_*.parquet"
                cache_path = DB_CACHE_DIR / f"daily_collection_cache_{file_hash}.parquet"
                cols_to_use = ['AccountID', 'CollectionTotal', 'Trxdate', 'ReverseTotal']

            progress_messages = []

            # Delete OLD caches (different hash) — keep the current hash if already cached
            deleted_count = 0
            for old_cache in DB_CACHE_DIR.glob(cache_pattern):
                if old_cache.name == cache_path.name:
                    continue  # same hash — keep it
                try:
                    old_cache.unlink()
                    deleted_count += 1
                    msg = f"Deleted old {file_type.upper()} cache: {old_cache.name}"
                    logging.info(msg)
                    progress_messages.append({'type': 'delete', 'message': msg})
                except Exception as e:
                    msg = f"Could not delete {old_cache.name}: {e}"
                    logging.warning(msg)
                    progress_messages.append({'type': 'warning', 'message': msg})

            if deleted_count > 0:
                msg = f"Cleared {deleted_count} old {file_type.upper()} cache file(s)"
                logging.info(msg)
                progress_messages.append({'type': 'cleanup', 'message': msg})

            # Read Excel and convert to Parquet
            msg = f"Caching {file_type.upper()} file (hash: {file_hash})..."
            logging.info(msg)
            progress_messages.append({'type': 'cache', 'message': msg})
            t_start = time.time()

            df = smart_read_excel(tmp_path)
            if cols_to_use:
                available = [c for c in cols_to_use if c in df.columns]
                if available:
                    df = df[available]

            # Fix mixed-type columns before Parquet conversion
            # (e.g. ReverseTotal may contain both ints and strings in Excel)
            numeric_cols = {'CollectionTotal', 'ReverseTotal'}
            for col in df.columns:
                if col in numeric_cols and df[col].dtype == 'object':
                    df[col] = pd.to_numeric(df[col], errors='coerce')

            df.to_parquet(cache_path, index=False)

            # Save full Excel copy for "Last Cache" quick-reload feature
            full_copy = DB_CACHE_DIR / f"daily_{file_type}_last.xlsx"
            shutil.copy(tmp_path, full_copy)
            # Preserve original filename for VBA date extraction
            meta_path = DB_CACHE_DIR / f"daily_{file_type}_last.meta.json"
            meta_path.write_text(json.dumps({'originalName': file.filename}))
            _append_cache_history(file_type, file.filename, full_copy)
            logging.info(f"Saved full Excel copy: {full_copy.name} (original: {file.filename})")

            elapsed = time.time() - t_start
            file_size = tmp_path.stat().st_size
            cache_size = cache_path.stat().st_size

            def fmt_size(b):
                return f"{b/1024:.1f}KB" if b < 1024 * 1024 else f"{b/1024/1024:.1f}MB"

            msg = f"Cached {file_type.upper()}: {cache_path.name} ({elapsed:.1f}s, {fmt_size(file_size)} -> {fmt_size(cache_size)})"
            logging.info(msg)
            progress_messages.append({'type': 'success', 'message': msg})

            return jsonify({
                'success': True,
                'cached': True,
                'hash': file_hash,
                'time': round(elapsed, 2),
                'originalSize': file_size,
                'cacheSize': cache_size,
                'message': f'{file_type.upper()} cached successfully in {elapsed:.1f}s',
                'progress': progress_messages
            })
        finally:
            tmp_path.unlink(missing_ok=True)

    except Exception as e:
        err = user_error(e, context='eod-cache-file')
        return jsonify({
            'error': err['user_message'],
            'suggestion': err['suggestion'],
        }), 500


@eod_bp.route('/process', methods=['POST'])
@require_api_key
def process_files_endpoint():
    """Main processing endpoint - merges PAR + Demand + Collection."""
    if not try_acquire_processing():
        return jsonify({
            'error': 'Server is busy processing another request. Please try again in a moment.'
        }), 503
    try:
        _proc_t0 = time.perf_counter()
        use_last_cache = request.form.get('useLastCache') == 'true'
        use_backend_demand = request.form.get('useBackendDemand') == 'true' or use_last_cache
        auto_fix_sheets = request.form.get('autoFixSheets') == 'true'

        cache_par = request.form.get('cachePar') == 'true'
        cache_collection = request.form.get('cacheCollection') == 'true'

        if cache_par or cache_collection:
            logging.info(f"Cache preferences: PAR={cache_par}, Collection={cache_collection}")

        # EOD date = the UPLOADED data's own date, so OverAll's "as on" label,
        # its source window (ftod_mask = Meeting Date in [first_of_month,
        # target_date]) AND the on-date Today/Tomorrow all follow the files you
        # actually uploaded. Priority: collection Trxdate (the real data date,
        # resolved below after the file is saved) > uploaded-filename
        # "as on DD-MM-YYYY". The frontend-sent "today" is only a last-resort
        # fallback. Upload 07-06 files -> 07-06 report; upload 03-06 -> 03-06.
        target_date = None
        import re as _re
        _eod_date_re = _re.compile(r'(?<!\d)(\d{2})[-_./](\d{2})[-_./](\d{4})(?!\d)')
        for _fobj in (request.files.get('collection'), request.files.get('par'),
                      request.files.get('demand')):
            _fname = getattr(_fobj, 'filename', '') if _fobj else ''
            if _fname:
                _m = _eod_date_re.search(_fname)
                if _m:
                    try:
                        target_date = datetime.strptime(
                            f"{_m.group(1)}-{_m.group(2)}-{_m.group(3)}", '%d-%m-%Y')
                        logging.info(f"EOD target_date from data filename '{_fname}': "
                                     f"{target_date.strftime('%d-%m-%Y')}")
                        break
                    except ValueError:
                        continue

        # Validate required files (skip if using last cache)
        if not use_last_cache:
            if 'par' not in request.files or 'collection' not in request.files:
                return jsonify({'error': 'Missing PAR or Collection files'}), 400

            if not use_backend_demand and 'demand' not in request.files:
                return jsonify({'error': 'Missing Demand file'}), 400

        par = request.files.get('par')
        collection = request.files.get('collection')

        # Clear log queue before new run
        with log_queue.mutex:
            log_queue.queue.clear()

        # Reset session for new archive folder
        reset_session()

        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            par_path = temp_path / "par.xlsx"
            collection_path = temp_path / "collection.xlsx"
            output_path = temp_path / "output.xlsx"

            # Check for pre-fixed files from /fix-sheets endpoint
            fixed_par_path = TEMP_DIR / 'fixed_par.xlsx'
            fixed_coll_path = TEMP_DIR / 'fixed_collection.xlsx'

            if use_last_cache:
                # Use saved Excel copies from last upload
                cached_par = DB_CACHE_DIR / 'daily_par_last.xlsx'
                cached_coll = DB_CACHE_DIR / 'daily_collection_last.xlsx'
                if not cached_par.exists() or not cached_coll.exists():
                    return jsonify({'error': 'No cached files found. Upload files at least once first.'}), 400
                shutil.copy(cached_par, par_path)
                shutil.copy(cached_coll, collection_path)
                logging.info("Using cached Excel files from last upload (Last Cache)")
            elif auto_fix_sheets and fixed_par_path.exists():
                shutil.copy(fixed_par_path, par_path)
                logging.info("Using pre-fixed PAR file (sheet renamed to 'Sheet1')")
                fixed_par_path.unlink()
                if fixed_coll_path.exists():
                    shutil.copy(fixed_coll_path, collection_path)
                    logging.info("Using pre-fixed Collection file (sheet renamed to 'Sheet1')")
                    fixed_coll_path.unlink()
                else:
                    collection.save(collection_path)
            else:
                par.save(par_path)
                collection.save(collection_path)

            # Always save full Excel copies for "Last Cache" quick-reload
            if not use_last_cache:
                try:
                    DB_CACHE_DIR.mkdir(parents=True, exist_ok=True)
                    shutil.copy(par_path, DB_CACHE_DIR / 'daily_par_last.xlsx')
                    shutil.copy(collection_path, DB_CACHE_DIR / 'daily_collection_last.xlsx')
                    # Save original filenames
                    par_name = par.filename if par else 'PAR'
                    coll_name = collection.filename if collection else 'Collection'
                    (DB_CACHE_DIR / 'daily_par_last.meta.json').write_text(
                        json.dumps({'originalName': par_name}))
                    (DB_CACHE_DIR / 'daily_collection_last.meta.json').write_text(
                        json.dumps({'originalName': coll_name}))
                    _append_cache_history('par', par_name, DB_CACHE_DIR / 'daily_par_last.xlsx')
                    _append_cache_history('collection', coll_name, DB_CACHE_DIR / 'daily_collection_last.xlsx')
                    logging.info("Saved Excel copies for Last Cache feature")
                except Exception as e:
                    logging.warning(f"Could not save Last Cache copies: {e}")

            # Handle demand file
            if use_backend_demand:
                demand_files = list(BACKEND_DATA_DIR.glob("Demand_Sheet_Master_*"))
                if not demand_files:
                    return jsonify({'error': 'No backend demand file found. Please upload one in Backend Data.'}), 400
                demand_path = demand_files[0]
            else:
                demand = request.files['demand']
                demand_path = temp_path / "demand.xlsx"
                demand.save(demand_path)

            # Authoritative EOD date: the Collection's own latest transaction
            # date (max Trxdate) IS the uploaded data's date. This drives the
            # OverAll "as on" label + data window AND the on-date Today/Tomorrow,
            # so everything reflects the files you uploaded. Overrides the
            # filename guess above.
            try:
                from services.eod_processor import parse_trxdate
                _coll_dates = smart_read_excel(collection_path, usecols=['Trxdate'])
                _coll_dates['Trxdate'] = parse_trxdate(_coll_dates['Trxdate'])
                _coll_max = _coll_dates['Trxdate'].dropna().max()
                if pd.notna(_coll_max):
                    _dd = pd.Timestamp(_coll_max)
                    target_date = datetime(_dd.year, _dd.month, _dd.day)
                    logging.info(f"EOD target_date from Collection Trxdate (data date): "
                                 f"{target_date.strftime('%d-%m-%Y')}")
            except Exception as _date_err:
                logging.warning(f"Collection Trxdate date-detect skipped: {_date_err}")
            if target_date is None:
                target_date = datetime.now()
                logging.info(f"EOD target_date fallback to today: {target_date.strftime('%d-%m-%Y')}")

            # Process using the imported logic (returns DataFrame + report path)
            # sheets_dir=None: skip inline sheet extraction (saves ~15-20s)
            # Individual sheets are extracted in a background thread below.
            _extraction_done.clear()
            sheets_dir = str(BACKEND_DATA_DIR / 'sheets')
            df_result, report_path = processor.process_files(
                demand_path, collection_path, par_path, output_path,
                auto_fix_sheets=auto_fix_sheets,
                db_manager=db_manager,
                target_date=target_date,
                sheets_dir=sheets_dir,
            )

            # Validate processing produced results (partial failure detection)
            if df_result is None or len(df_result) == 0:
                return jsonify({
                    'error': 'Processing completed but produced no results. Check input files.',
                    'suggestion': 'Verify that your PAR, Collection, and Demand files contain valid data.'
                }), 500

            # Invalidate employee merged DF cache (full flush -- demand data changed)
            invalidate_merged_df_cache()

            # Archive & sync
            logging.info("STEP 5: Archiving files & syncing outputs")
            archive_file(par_path, "PAR", is_output=False, is_monthly=False)
            archive_file(collection_path, "Collection", is_output=False, is_monthly=False)
            archive_file(demand_path, "Demand_Master", is_output=False, is_monthly=True)

            # Also archive Last Month PAR if it exists
            last_month_par_files = list(BACKEND_DATA_DIR.glob("Last_Month_PAR_*"))
            if last_month_par_files:
                archive_file(last_month_par_files[0], "Last_Month_PAR", is_output=False, is_monthly=True)

            # Save report file (VBA-replacement) to backend dir before temp cleanup
            if report_path and Path(report_path).exists():
                report_dest = BACKEND_DATA_DIR / 'EOD_Report_Latest.xlsx'
                try:
                    shutil.copy(report_path, report_dest)
                    logging.info(f"Report saved to {report_dest}")
                    # Update sheet manifest to reference the backend copy (not temp file)
                    _manifest_path = BACKEND_DATA_DIR / 'sheets' / 'manifest.json'
                    if _manifest_path.exists():
                        try:
                            _mdata = json.loads(_manifest_path.read_text())
                            _mdata['source'] = report_dest.name
                            _mdata['source_mtime'] = report_dest.stat().st_mtime
                            _manifest_path.write_text(json.dumps(_mdata, indent=2))
                        except Exception:
                            pass  # non-fatal — manifest validation may just skip
                    # Start background preload so email preview is instant
                    _preload_report_wb()
                    # Render body image for email attachment (background)
                    import threading as _thr
                    _thr.Thread(target=_render_body_image, daemon=True).start()
                    # Individual sheets were generated inline by build_report (sheets_dir param).
                    # Signal extraction done so email sending can proceed immediately.
                    _extraction_done.set()
                except Exception as rpt_err:
                    logging.warning(f"Report save failed (non-fatal): {rpt_err}")

            # Archive output file
            archive_file(output_path, "Output", is_output=True, is_monthly=False)

            # Auto-flow EOD output to HOURLY backend directory
            try:
                file_manager.save_file_to_backend(output_path, BACKEND_DATA_DIR)
                logging.info("Auto-flow: EOD output synced to HOURLY backend")
            except Exception as af_err:
                logging.warning(f"Auto-flow failed (non-fatal): {af_err}")

            # Auto-cache EOD output as Parquet for fast Hourly reads
            # Uses the in-memory DataFrame directly (avoids re-reading the Excel file)
            try:
                eod_backend_files = list(BACKEND_DATA_DIR.glob("EOD_Output_*"))
                if eod_backend_files and df_result is not None:
                    eod_file_path = eod_backend_files[0]

                    # Hash first 1MB of the backend file for cache key
                    hasher = hashlib.md5()
                    with open(eod_file_path, 'rb') as f:
                        hasher.update(f.read(1024 * 1024))
                    file_hash = hasher.hexdigest()[:16]

                    # Delete old hourly EOD caches
                    DB_CACHE_DIR.mkdir(parents=True, exist_ok=True)
                    for old_cache in DB_CACHE_DIR.glob("hourly_eod_cache_*.parquet"):
                        try:
                            old_cache.unlink()
                            logging.info(f"Auto-cache: deleted old hourly EOD cache: {old_cache.name}")
                        except OSError:
                            pass  # cache file cleanup

                    # Write Parquet directly from in-memory DataFrame (no Excel re-read)
                    # Mutate in-place instead of .copy() to avoid doubling memory
                    cache_path = DB_CACHE_DIR / f"hourly_eod_cache_{file_hash}.parquet"
                    obj_cols = df_result.select_dtypes(include='object').columns
                    df_result[obj_cols] = df_result[obj_cols].replace('', pd.NA)
                    df_result.to_parquet(cache_path, index=False)
                    logging.info(f"Auto-cache: EOD output cached as Parquet for Hourly ({cache_path.name})")
            except Exception as cache_err:
                logging.warning(f"Auto-cache for Hourly failed (non-fatal): {cache_err}")

            # Auto-generate Employee Report (aggregated only)
            try:
                emp_output = BACKEND_DATA_DIR / 'Employee_Report_Latest.xlsx'
                t_date = target_date or pd.Timestamp.now()
                if not isinstance(t_date, pd.Timestamp):
                    t_date = pd.Timestamp(t_date)

                logging.info("STEP 6: Building employee report...")
                emp_path = processor.build_employee_report(df_result, t_date, emp_output)
            except Exception as emp_err:
                logging.warning(f"Employee report generation failed (non-fatal): {emp_err}")

            # Free the large result DataFrame now that Excel + Parquet are written
            del df_result
            gc_checkpoint("eod-df-freed")

            # Cleanup & finalize
            logging.info("STEP 7: Cleanup & finalize")
            threading.Thread(target=cleanup_old_daily_caches, daemon=True).start()

            # Save main output to backend dir so it survives temp dir cleanup
            output_dest = BACKEND_DATA_DIR / 'EOD_Output_Latest.xlsx'
            shutil.copy(output_path, output_dest)

            # Persist target_date so VBA template injection can use it at bundle-save time
            if target_date:
                td = target_date if isinstance(target_date, datetime) else datetime.now()
                (BACKEND_DATA_DIR / '.target_date').write_text(td.strftime('%d-%m-%Y'))

            _proc_elapsed = time.perf_counter() - _proc_t0
            logging.info(f"ALL STEPS COMPLETED in {_proc_elapsed:.1f} seconds")

            # Return JSON with available downloads instead of auto-downloading
            available = ['eod']
            if (BACKEND_DATA_DIR / 'EOD_Report_Latest.xlsx').exists():
                available.append('report')
            if (BACKEND_DATA_DIR / 'Employee_Report_Latest.xlsx').exists():
                available.append('employee')

            # Sheets generated inline by build_report — ensure event is set for email waiters
            _extraction_done.set()

            result = {
                'status': 'success',
                'available': available,
                'message': 'Processing complete. Choose which file to download.'
            }

            return jsonify(result)

    except Exception as e:
        _extraction_done.set()  # Unblock any waiters on failure
        err = user_error(e, context='eod-process')
        return jsonify({
            'error': err['user_message'],
            'suggestion': err['suggestion'],
        }), 500
    finally:
        gc_checkpoint("eod-request-complete")
        release_processing()


@eod_bp.route('/download-output', methods=['GET'])
def download_output():
    """Download the main EOD output (Regular Demand Vs Collection)."""
    output_file = BACKEND_DATA_DIR / 'EOD_Output_Latest.xlsx'
    if not output_file.exists():
        return jsonify({'error': 'No output available. Run EOD processing first.'}), 404
    return send_file(
        output_file,
        as_attachment=True,
        download_name='Regular Demand Vs Collection.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@eod_bp.route('/download-report', methods=['GET'])
def download_report():
    """Download the latest generated report (VBA-replacement Excel)."""
    report_file = BACKEND_DATA_DIR / 'EOD_Report_Latest.xlsx'
    if not report_file.exists():
        return jsonify({'error': 'No report available. Run EOD processing first.'}), 404
    return send_file(
        report_file,
        as_attachment=True,
        download_name='EOD_Report.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


def _get_latest_mtime():
    """Get the modification time of the latest output file (used to detect same data)."""
    eod_output = BACKEND_DATA_DIR / 'EOD_Output_Latest.xlsx'
    if eod_output.exists():
        return eod_output.stat().st_mtime
    return None


def _find_existing_save(mtime):
    """Check if EOD_Bundle already has a save matching this processing run's mtime."""
    downloads = Path.home() / 'Downloads'
    if not downloads.exists():
        downloads = Path.home() / 'Desktop'
    bundle_dir = downloads / 'EOD_Bundle'
    marker_file = bundle_dir / '.last_save_mtime'
    if marker_file.exists():
        try:
            content = marker_file.read_text()
            saved_mtime = float(content.strip().split('\n')[0])
            saved_path = content.strip().split('\n')[1] if '\n' in content else ''
            if abs(saved_mtime - mtime) < 0.01 and saved_path and Path(saved_path).exists():
                return saved_path
        except (ValueError, IndexError):
            pass
    return None


def _save_bundle(save_dir):
    """Copy latest reports + VBA to the given directory."""
    save_dir = Path(save_dir)
    save_dir.mkdir(parents=True, exist_ok=True)

    saved = []
    eod_output = BACKEND_DATA_DIR / 'EOD_Output_Latest.xlsx'
    eod_report = BACKEND_DATA_DIR / 'EOD_Report_Latest.xlsx'
    if eod_output.exists():
        shutil.copy2(eod_output, save_dir / 'Regular Demand Vs Collection.xlsx')
        saved.append('Regular Demand Vs Collection.xlsx')
    if eod_report.exists():
        shutil.copy2(eod_report, save_dir / 'EOD_Report.xlsx')
        saved.append('EOD_Report.xlsx')

    # Read persisted target_date for VBA injection
    target_date_file = BACKEND_DATA_DIR / '.target_date'
    vba_date_injection = None
    if target_date_file.exists():
        try:
            td_str = target_date_file.read_text().strip()  # "dd-mm-yyyy"
            parts = td_str.split('-')
            if len(parts) == 3:
                dd, mm, yyyy = int(parts[0]), int(parts[1]), int(parts[2])
                vba_date_injection = (yyyy, mm, dd, td_str)
                logging.info(f"VBA date injection: {td_str} -> DateSerial({yyyy},{mm},{dd})")
        except Exception as e:
            logging.warning(f"Could not read target_date for VBA injection: {e}")

    vba_dir = Path(config.STATIC_DIR) / 'eod'
    for src_name, dst_name in [
        ('vba_template.js', 'VBA_Template.txt'),
        ('vba_template_month_end.js', 'VBA_Template_Month_End.txt'),
    ]:
        src = vba_dir / src_name
        if src.exists():
            content = src.read_text(encoding='utf-8')
            start = content.find('`')
            end = content.rfind('`')
            if start != -1 and end > start:
                vba_text = content[start + 1:end]
            else:
                vba_text = content

            # Inject target date into VBA (replace placeholder with DateSerial)
            if vba_date_injection:
                yyyy, mm, dd, td_str = vba_date_injection
                vba_text = vba_text.replace(
                    "dateWasInjected = False",
                    "dateWasInjected = True"
                )
                vba_text = vba_text.replace(
                    "' targetDate = DateSerial(YYYY, MM, DD) <-- Date will be filled here\n    targetDate = Date ' Default to Today\n    reportDate = Format(targetDate, \"dd-mm-yyyy\") ' Default",
                    f"targetDate = DateSerial({yyyy}, {mm}, {dd})\n    reportDate = \"{td_str}\""
                )

                # Inject dynamic FY label (replace hardcoded FY_25-26 with correct label)
                try:
                    from services.eod_processor import get_fy_label
                    fy_label = get_fy_label(datetime(int(yyyy), int(mm), int(dd)))
                    vba_text = vba_text.replace('FY_25-26', fy_label)
                except Exception as e:
                    logging.warning(f"Could not inject FY label into VBA: {e}")

            (save_dir / dst_name).write_text(vba_text, encoding='utf-8')
            saved.append(dst_name)

    # Save target_date metadata for the runner
    if vba_date_injection:
        (save_dir / '.target_date').write_text(vba_date_injection[3])

    # Write marker so we can detect duplicate saves
    downloads = Path.home() / 'Downloads'
    if not downloads.exists():
        downloads = Path.home() / 'Desktop'
    marker = downloads / 'EOD_Bundle' / '.last_save_mtime'
    mtime = _get_latest_mtime()
    if mtime is not None:
        marker.parent.mkdir(parents=True, exist_ok=True)
        marker.write_text(f"{mtime}\n{save_dir}")

    return saved, str(save_dir)


@eod_bp.route('/save-bundle-to-server', methods=['POST'])
def save_bundle_to_server():
    """Save the bundle to Downloads/EOD_Bundle/<timestamp>.

    Supports duplicate detection:
      - Default (no action): checks if same data already saved → returns prompt
      - action=replace: overwrite the existing folder
      - action=new: create a new timestamped folder
    """
    from datetime import datetime
    from flask import request as req

    action = req.json.get('action') if req.is_json else None
    mtime = _get_latest_mtime()

    downloads = Path.home() / 'Downloads'
    if not downloads.exists():
        downloads = Path.home() / 'Desktop'

    # Check for existing save of the same data
    if action is None and mtime is not None:
        existing = _find_existing_save(mtime)
        if existing:
            return jsonify(
                already_saved=True,
                existing_path=existing,
                existing_name=Path(existing).name
            )

    now = datetime.now()
    timestamp = f"{now.strftime('%Hh.%Mm.%Ss')} on {now.strftime('%d-%m-%Y')}"

    if action == 'replace':
        existing = _find_existing_save(mtime) if mtime else None
        if existing:
            save_dir = existing
        else:
            save_dir = str(downloads / 'EOD_Bundle' / timestamp)
    else:
        # action=new or first save
        save_dir = str(downloads / 'EOD_Bundle' / timestamp)

    saved, path = _save_bundle(save_dir)
    return jsonify(success=True, saved=saved, path=path)


@eod_bp.route('/download-employee-report', methods=['GET'])
def download_employee_report():
    """Download the employee report (raw data, 3 product sheets)."""
    emp_file = BACKEND_DATA_DIR / 'Employee_Report_Latest.xlsx'
    if not emp_file.exists():
        return jsonify({'error': 'No employee report available. Generate it first.'}), 404
    return send_file(
        emp_file,
        as_attachment=True,
        download_name='Employee_Report.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@eod_bp.route('/download-employee-report-accounts', methods=['GET'])
def download_employee_report_accounts():
    """Download the account-level employee report (raw data, 3 product sheets)."""
    emp_file = BACKEND_DATA_DIR / 'Employee_Report_Accounts_Latest.xlsx'
    if not emp_file.exists():
        return jsonify({'error': 'No account-level employee report available. Generate it first.'}), 404
    return send_file(
        emp_file,
        as_attachment=True,
        download_name='Employee_Report_Accounts.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


import re as _re

_EOD_PAR_DATE_RE_SEP = _re.compile(r'(?<!\d)(\d{2})[\-_./](\d{2})[\-_./](\d{4})(?!\d)')
_EOD_PAR_DATE_RE_PACKED = _re.compile(r'(?<!\d)(\d{2})(\d{2})(\d{4})(?!\d)')


def _derive_eod_par_date():
    """Parse YYYY-MM-DD from the last-cached PAR filename meta, else None.

    The EOD /process pipeline writes data/db/cache/daily_par_last.meta.json
    with {"originalName": "Par as on DD-MM-YYYY.xlsx"} on every run. This
    lets /sync-to-dashboard stamp the right date on the daily_performance
    row without the frontend needing to send one.
    """
    meta_path = config.DB_CACHE_DIR / 'daily_par_last.meta.json'
    if not meta_path.exists():
        return None
    try:
        original_name = json.loads(meta_path.read_text()).get('originalName') or ''
    except Exception:
        return None
    for rx in (_EOD_PAR_DATE_RE_SEP, _EOD_PAR_DATE_RE_PACKED):
        for m in rx.finditer(Path(original_name).name):
            day, month, year = m.group(1), m.group(2), m.group(3)
            try:
                return datetime.strptime(f"{day}-{month}-{year}", '%d-%m-%Y').strftime('%Y-%m-%d')
            except ValueError:
                continue
    return None


def _get_report_date_str():
    """Return dd-mm-yyyy date for EOD report title/filename.

    Source priority:
      1. .target_date file (persisted at processing time, already dd-mm-yyyy)
      2. PAR cache filename (parsed via _derive_eod_par_date -> YYYY-MM-DD)
      3. today (fallback)
    Ensures email subject + attachment filename match the data's date,
    not the date the email was sent.
    """
    target_date_file = BACKEND_DATA_DIR / '.target_date'
    if target_date_file.exists():
        try:
            td_str = target_date_file.read_text().strip()
            datetime.strptime(td_str, '%d-%m-%Y')
            return td_str
        except Exception:
            pass
    par_iso = _derive_eod_par_date()
    if par_iso:
        try:
            return datetime.strptime(par_iso, '%Y-%m-%d').strftime('%d-%m-%Y')
        except ValueError:
            pass
    return datetime.now().strftime('%d-%m-%Y')


@eod_bp.route('/sync-to-dashboard', methods=['POST'])
@require_api_key
def sync_to_dashboard():
    """Upload the latest Employee Report to the Coll_Db dashboard.

    Two endpoints hit in sequence:
      1. POST {COLLDB_URL}/api/upload        — aggregate collection tab
      2. POST {COLLDB_URL}/api/upload-daily  — daily_performance row for
         the date derived from the cached PAR filename, so the dashboard
         calendar lights up that date.
    Step 2 is best-effort: if no date can be derived or the endpoint errors,
    step 1's success is still reported (daily_status reflects what happened).
    """
    try:
        emp_output = BACKEND_DATA_DIR / 'Employee_Report_Latest.xlsx'
        if not emp_output.exists():
            return jsonify({
                'success': False,
                'message': 'Employee report not found. Run EOD processing first.',
            }), 404

        target_url = f"{config.COLLDB_URL.rstrip('/')}/api/upload"
        emp_bytes = emp_output.read_bytes()

        try:
            resp = http_requests.post(
                target_url,
                files={'file': ('Employee_Report.xlsx', io.BytesIO(emp_bytes),
                                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
                timeout=60,
            )
        except (http_requests.exceptions.ConnectionError,
                http_requests.exceptions.ConnectTimeout) as conn_err:
            logging.warning(f"EOD sync: Coll_Db not reachable at {config.COLLDB_URL}: {conn_err}")
            return jsonify({
                'success': False,
                'target': target_url,
                'message': (
                    f"Coll_Db server not reachable at {config.COLLDB_URL}. "
                    "Start it with `cd /Users/raghunandanmali/Desktop/Coll_Db/server && node index.js`"
                ),
            }), 503

        if resp.status_code == 200:
            try:
                db_result = resp.json()
            except ValueError:
                db_result = {}
            emp_count = db_result.get('employees') or db_result.get('empCount') or 0
            perf_count = db_result.get('performance') or db_result.get('perfCount') or 0

            # ── Step 2: daily upload (best-effort) ────────────────────
            daily_url = f"{config.COLLDB_URL.rstrip('/')}/api/upload-daily"
            par_date = _derive_eod_par_date()
            daily_status = {'attempted': False}
            daily_inserted = 0
            if not par_date:
                daily_status = {
                    'attempted': False,
                    'reason': 'Could not derive date from daily_par_last.meta.json',
                }
                logging.warning("EOD sync: skipping daily upload — no PAR date derivable")
            else:
                daily_status = {'attempted': True, 'date': par_date, 'target': daily_url}
                try:
                    daily_resp = http_requests.post(
                        daily_url,
                        files={'file': ('Employee_Report.xlsx', io.BytesIO(emp_bytes),
                                        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
                        data={'date': par_date},
                        timeout=60,
                    )
                    if daily_resp.status_code == 200:
                        try:
                            daily_json = daily_resp.json()
                        except ValueError:
                            daily_json = {}
                        daily_inserted = daily_json.get('inserted') or daily_json.get('count') or 0
                        daily_status.update({'ok': True, 'inserted': daily_inserted, 'stats': daily_json})
                        logging.info(f"EOD sync: daily upload OK for {par_date} ({daily_inserted} rows)")
                    else:
                        daily_status.update({
                            'ok': False,
                            'httpStatus': daily_resp.status_code,
                            'body': (daily_resp.text or '')[:300],
                        })
                        logging.warning(
                            f"EOD sync: daily upload failed HTTP {daily_resp.status_code} for {par_date}"
                        )
                except (http_requests.exceptions.ConnectionError,
                        http_requests.exceptions.ConnectTimeout) as daily_err:
                    daily_status.update({'ok': False, 'error': str(daily_err)})
                    logging.warning(f"EOD sync: daily upload connection error: {daily_err}")

            msg = f"Coll_Db collection tab updated — {emp_count} employees, {perf_count} records"
            if daily_status.get('ok'):
                msg += f" · daily row for {par_date}: {daily_inserted} inserted"
            elif daily_status.get('attempted'):
                msg += f" · daily upload failed for {par_date}"
            else:
                msg += " · daily upload skipped (no PAR date)"

            return jsonify({
                'success': True,
                'target': target_url,
                'message': msg,
                'stats': db_result,
                'daily': daily_status,
            })
        else:
            # Preserve body for diagnostics, but avoid dumping huge HTML error pages
            body_text = (resp.text or '')[:500]
            return jsonify({
                'success': False,
                'target': target_url,
                'message': f"Upload failed ({resp.status_code}): {body_text}",
            }), 502

    except Exception as e:
        err = user_error(e, context='eod-sync-dashboard')
        return jsonify({'success': False, 'message': err['user_message']}), 500


@eod_bp.route('/daily-check-date', methods=['GET'])
def daily_check_date():
    """Check if a date already has data in the daily_performance table."""
    try:
        date = request.args.get('date', '')
        if not date:
            return jsonify({'error': 'date required'}), 400
        # Coll_Db (growwithme) — EC2_UPLOAD_URL points at a stale Node box
        # whose Postgres lacks daily_performance.
        base_url = config.COLLDB_URL.rstrip('/')
        resp = http_requests.get(f'{base_url}/api/daily/check-date?date={date}', timeout=10)
        return jsonify(resp.json())
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@eod_bp.route('/sync-daily', methods=['POST'])
@require_api_key
def sync_daily():
    """Upload the latest EOD Employee Report to daily_performance with a specific date."""
    try:
        data = request.get_json() or {}
        report_date = data.get('date', '')
        if not report_date:
            return jsonify({'success': False, 'message': 'date is required (YYYY-MM-DD)'}), 400

        emp_output = BACKEND_DATA_DIR / 'Employee_Report_Latest.xlsx'
        if not emp_output.exists():
            emp_output = BACKEND_DATA_DIR / 'EOD_Report_Latest.xlsx'
        if not emp_output.exists():
            return jsonify({'success': False, 'message': 'No report found. Run processing first.'}), 404

        # Coll_Db (growwithme) — EC2_UPLOAD_URL points at a stale Node box
        # whose Postgres lacks daily_performance.
        base_url = config.COLLDB_URL.rstrip('/')
        daily_url = f'{base_url}/api/upload-daily'

        with open(emp_output, 'rb') as f:
            resp = http_requests.post(
                daily_url,
                files={'file': ('report.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
                data={'date': report_date},
                timeout=60,
            )

        if resp.status_code == 200:
            result = resp.json()
            return jsonify({
                'success': True,
                'message': f"Daily data uploaded for {report_date} — {result.get('inserted', 0)} records",
                'stats': result,
            })
        else:
            return jsonify({'success': False, 'message': f"Upload failed ({resp.status_code}): {resp.text}"}), 502

    except Exception as e:
        err = user_error(e, context='eod-sync-daily')
        return jsonify({'success': False, 'message': err['user_message']}), 500


@eod_bp.route('/generate-employee-report', methods=['POST'])
def generate_employee_report_endpoint():
    """Generate employee reports (aggregated + account-level) from the latest EOD output."""
    try:
        logging.info("EMPLOYEE REPORT: Starting on-demand generation")

        # 1. Read EOD output data — try parquet cache first, then Excel
        df = None
        target_date = None

        # Accept target_date from POST body (form or JSON)
        target_date_str = None
        if request.is_json:
            target_date_str = request.json.get('targetDate')
        else:
            target_date_str = request.form.get('targetDate')

        if target_date_str:
            try:
                target_date = pd.Timestamp(datetime.strptime(target_date_str, '%d-%m-%Y'))
                logging.info(f"EMPLOYEE REPORT: Using provided target_date: {target_date_str}")
            except ValueError:
                logging.warning(f"EMPLOYEE REPORT: Could not parse target_date '{target_date_str}'")

        # Try hourly EOD parquet cache first (fastest)
        if DB_CACHE_DIR.exists():
            parquet_caches = sorted(DB_CACHE_DIR.glob("hourly_eod_cache_*.parquet"),
                                     key=lambda f: f.stat().st_mtime, reverse=True)
            if parquet_caches:
                try:
                    logging.info(f"EMPLOYEE REPORT: Reading from parquet cache: {parquet_caches[0].name}")
                    df = pd.read_parquet(parquet_caches[0])
                except Exception as e:
                    logging.warning(f"EMPLOYEE REPORT: Parquet read failed: {e}")

        # Fallback to Excel
        if df is None:
            eod_excel = BACKEND_DATA_DIR / 'EOD_Output_Latest.xlsx'
            if not eod_excel.exists():
                return jsonify({'error': 'No EOD output available. Run EOD processing first.'}), 404
            logging.info("EMPLOYEE REPORT: Reading from EOD_Output_Latest.xlsx")
            df = smart_read_excel(eod_excel)

        if df is None or len(df) == 0:
            return jsonify({'error': 'EOD output is empty. Run EOD processing first.'}), 404

        # Derive target_date from data if not provided
        if target_date is None and 'Meeting Date' in df.columns:
            try:
                from services.eod_processor import parse_date_column
                meeting_dates = parse_date_column(df['Meeting Date'])
                max_date = meeting_dates.dropna().max()
                if pd.notna(max_date):
                    target_date = max_date
                    logging.info(f"EMPLOYEE REPORT: Derived target_date from data: {target_date.strftime('%d-%m-%Y')}")
            except Exception as e:
                logging.warning(f"EMPLOYEE REPORT: Could not derive target_date: {e}")

        if target_date is None:
            target_date = pd.Timestamp.now()
            logging.info(f"EMPLOYEE REPORT: Defaulting target_date to today: {target_date.strftime('%d-%m-%Y')}")

        # Check required columns
        if 'Emp ID' not in df.columns:
            return jsonify({'error': 'EOD output missing Emp ID column. Cannot generate employee report.'}), 400

        # 2. Generate aggregated employee report
        emp_output = BACKEND_DATA_DIR / 'Employee_Report_Latest.xlsx'
        logging.info("EMPLOYEE REPORT: Generating aggregated report...")
        emp_path = processor.build_employee_report(df, target_date, emp_output)

        # 3. Generate account-level employee report
        emp_acct_output = BACKEND_DATA_DIR / 'Employee_Report_Accounts_Latest.xlsx'
        logging.info("EMPLOYEE REPORT: Generating account-level report...")
        emp_acct_path = processor.build_employee_report_with_accounts(df, target_date, emp_acct_output)

        # Build response
        result = {
            'success': True,
            'employee': emp_path is not None,
            'employee_accounts': emp_acct_path is not None,
            'message': 'Employee reports generated successfully.'
        }

        if not emp_path and not emp_acct_path:
            result['success'] = False
            result['message'] = 'Both employee report generations failed.'

        logging.info(f"EMPLOYEE REPORT: Done. Aggregated={emp_path is not None}, Accounts={emp_acct_path is not None}")

        # 4. Auto-upload Employee_Report to EC2 PostgreSQL
        if emp_path:
            try:
                EC2_UPLOAD_URL = config.EC2_UPLOAD_URL
                with open(emp_output, 'rb') as f:
                    resp = http_requests.post(
                        EC2_UPLOAD_URL,
                        files={'file': ('Employee_Report.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
                        timeout=30
                    )
                if resp.status_code == 200:
                    db_result = resp.json()
                    logging.info(f"EMPLOYEE REPORT: Auto-uploaded to EC2 PostgreSQL — {db_result.get('employees', 0)} employees, {db_result.get('performance', 0)} records")
                    result['db_upload'] = True
                    result['db_stats'] = db_result
                    result['message'] += ' Data synced to dashboard database.'
                else:
                    logging.warning(f"EMPLOYEE REPORT: EC2 upload failed ({resp.status_code}): {resp.text}")
                    result['db_upload'] = False
                    result['db_upload_error'] = f"Server returned {resp.status_code}"
            except Exception as upload_err:
                logging.warning(f"EMPLOYEE REPORT: EC2 upload failed: {upload_err}")
                result['db_upload'] = False
                result['db_upload_error'] = str(upload_err)

        return jsonify(result)

    except Exception as e:
        err = user_error(e, context='eod-generate-employee-report')
        return jsonify({
            'error': err['user_message'],
            'suggestion': err['suggestion'],
        }), 500


@eod_bp.route('/save-backend-file', methods=['POST'])
@require_api_key
def save_backend_file():
    """Save backend file (Master Demand or Last Month PAR)."""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400

        file = request.files['file']
        file_type = request.form.get('type')

        if not file_type or file_type not in ['masterDemand', 'lastMonthPar']:
            return jsonify({'error': 'Invalid file type'}), 400

        # Validate file is Excel
        valid, msg = file_manager.validate_excel_file(file)
        if not valid:
            return jsonify({'error': msg}), 400

        BACKEND_DATA_DIR.mkdir(exist_ok=True)

        prefixes = {
            'masterDemand': 'Demand_Sheet_Master_',
            'lastMonthPar': 'Last_Month_PAR_'
        }
        prefix = prefixes[file_type]

        # Delete existing files with this prefix
        for existing_file in BACKEND_DATA_DIR.glob(f"{prefix}*"):
            existing_file.unlink()
            logging.info(f"Deleted existing file: {existing_file}")

        from werkzeug.utils import secure_filename
        original_name = secure_filename(file.filename) or 'uploaded_file.xlsx'
        new_filename = f"{prefix}{original_name}"
        save_path = BACKEND_DATA_DIR / new_filename

        file.save(save_path)
        logging.info(f"Saved new file: {save_path}")

        # Upload/ingestion separation:
        #   ingest=false  -> only save the file (fast upload; DB Module flow).
        #   default/true  -> save AND ingest into DuckDB (legacy behaviour).
        skip_ingest = request.form.get('ingest') == 'false'

        ingestion_msg = ""
        if db_manager and not skip_ingest:
            success = False
            msg = ""
            if file_type == 'masterDemand':
                success, msg = db_manager.ingest_demand_master(save_path)
            elif file_type == 'lastMonthPar':
                success, msg = db_manager.ingest_last_month_par(save_path)

            if success:
                ingestion_msg = f" & DB Ingestion: {msg}"
            else:
                ingestion_msg = f" & DB Ingestion Failed: {msg}"

        return jsonify({
            'success': True,
            'ingested': bool(db_manager and not skip_ingest),
            'message': f'File saved successfully{ingestion_msg}',
            'filename': new_filename
        })

    except Exception as e:
        err = user_error(e, context='eod-save-backend')
        return jsonify({
            'error': err['user_message'],
            'suggestion': err['suggestion'],
        }), 500


@eod_bp.route('/backend-files-status', methods=['GET'])
def get_backend_files_status():
    """Get backend files status."""
    try:
        BACKEND_DATA_DIR.mkdir(exist_ok=True)

        status = {
            'masterDemand': None,
            'lastMonthPar': None,
            # Optional per-file metadata (size in bytes + modified epoch seconds).
            # Added alongside the legacy string keys so existing callers keep working.
            'meta': {'masterDemand': None, 'lastMonthPar': None},
        }

        def _meta(f):
            try:
                st = f.stat()
                # Strip the storage prefix so the UI shows the user's original name.
                display = f.name
                for prefix in ('Demand_Sheet_Master_', 'Last_Month_PAR_'):
                    if display.startswith(prefix):
                        display = display[len(prefix):]
                        break
                return {'name': f.name, 'displayName': display or f.name,
                        'size': st.st_size, 'modified': st.st_mtime}
            except OSError:
                return None

        for f in BACKEND_DATA_DIR.iterdir():
            if f.name.startswith('Demand_Sheet_Master_'):
                status['masterDemand'] = f.name
                status['meta']['masterDemand'] = _meta(f)
            elif f.name.startswith('Last_Month_PAR_'):
                status['lastMonthPar'] = f.name
                status['meta']['lastMonthPar'] = _meta(f)

        return jsonify(status)

    except Exception as e:
        err = user_error(e, context='eod-backend-status')
        return jsonify({
            'error': err['user_message'],
            'suggestion': err['suggestion'],
        }), 500


@eod_bp.route('/db-status', methods=['GET'])
def get_db_status():
    """Get the current status of database tables."""
    try:
        status = {
            'dbAvailable': db_manager is not None,
            'demandMaster': {'loaded': False, 'rowCount': 0},
            'lastMonthPar': {'loaded': False, 'rowCount': 0}
        }

        if db_manager:
            con = db_manager.get_connection()

            try:
                count = con.execute("SELECT count(*) FROM Demand_Master").fetchone()[0]
                if count > 0:
                    status['demandMaster'] = {'loaded': True, 'rowCount': count}
            except (duckdb.CatalogException, duckdb.Error):
                pass

            try:
                count = con.execute("SELECT count(*) FROM Last_Month_PAR").fetchone()[0]
                if count > 0:
                    status['lastMonthPar'] = {'loaded': True, 'rowCount': count}
            except (duckdb.CatalogException, duckdb.Error):
                pass

        return jsonify(status)

    except Exception as e:
        err = user_error(e, context='eod-db-status')
        return jsonify({
            'error': err['user_message'],
            'suggestion': err['suggestion'],
        }), 500


@eod_bp.route('/ingest-to-db', methods=['POST'])
@require_api_key
def ingest_to_db():
    """Ingest files from BACKEND_DATA folder into the database."""
    try:
        if not db_manager:
            return jsonify({'error': 'Database not available'}), 500

        results = {
            'demandMaster': {'success': False, 'message': 'No file found'},
            'lastMonthPar': {'success': False, 'message': 'No file found'}
        }

        demand_files = list(BACKEND_DATA_DIR.glob("Demand_Sheet_Master_*"))
        if demand_files:
            success, msg = db_manager.ingest_demand_master(demand_files[0])
            results['demandMaster'] = {'success': success, 'message': msg}

        par_files = list(BACKEND_DATA_DIR.glob("Last_Month_PAR_*"))
        if par_files:
            success, msg = db_manager.ingest_last_month_par(par_files[0])
            results['lastMonthPar'] = {'success': success, 'message': msg}

        return jsonify(results)

    except Exception as e:
        err = user_error(e, context='eod-ingest-db')
        return jsonify({
            'error': err['user_message'],
            'suggestion': err['suggestion'],
        }), 500


@eod_bp.route('/ingest-single-to-db', methods=['POST'])
@require_api_key
def ingest_single_to_db():
    """Ingest a single file type to the database."""
    try:
        if not db_manager:
            return jsonify({'error': 'Database not available'}), 500

        data = request.get_json()
        file_type = data.get('type')

        if file_type == 'demand':
            demand_files = list(BACKEND_DATA_DIR.glob("Demand_Sheet_Master_*"))
            if demand_files:
                success, msg = db_manager.ingest_demand_master(demand_files[0])
                return jsonify({'success': success, 'message': msg})
            else:
                return jsonify({'success': False, 'message': 'No Demand file found in BACKEND_DATA'})

        elif file_type == 'lastMonth':
            par_files = list(BACKEND_DATA_DIR.glob("Last_Month_PAR_*"))
            if par_files:
                success, msg = db_manager.ingest_last_month_par(par_files[0])
                return jsonify({'success': success, 'message': msg})
            else:
                return jsonify({'success': False, 'message': 'No Last Month PAR file found in BACKEND_DATA'})

        return jsonify({'success': False, 'message': 'Unknown file type'})

    except Exception as e:
        err = user_error(e, context='eod-single-ingest')
        return jsonify({
            'error': err['user_message'],
            'suggestion': err['suggestion'],
        }), 500


@eod_bp.route('/clear-db', methods=['POST'])
@require_api_key
def clear_db():
    """Clear all data from the database tables."""
    try:
        if not db_manager:
            return jsonify({'error': 'Database not available'}), 500

        con = db_manager.get_connection()
        con.execute("DROP TABLE IF EXISTS Demand_Master")
        con.execute("DROP TABLE IF EXISTS Last_Month_PAR")

        db_manager._init_schema()

        return jsonify({'success': True, 'message': 'Database cleared'})

    except Exception as e:
        err = user_error(e, context='eod-clear-db')
        return jsonify({
            'error': err['user_message'],
            'suggestion': err['suggestion'],
        }), 500


@eod_bp.route('/report-sheet-names', methods=['GET'])
def report_sheet_names():
    """Return the sheet names from the latest EOD Report, categorised."""
    report_file = BACKEND_DATA_DIR / 'EOD_Report_Latest.xlsx'
    if not report_file.exists():
        return jsonify({'error': 'No EOD Report available. Run processing first.'}), 404

    # Fast path: check in-memory cache / JSON sidecar (< 1ms)
    mtime = report_file.stat().st_mtime
    with _cache_lock:
        if _sheet_json_cache.get('names') and _sheet_json_cache['mtime'] == mtime:
            return jsonify(_sheet_json_cache['names'])
    _ensure_sidecar_loaded()
    with _cache_lock:
        if _sheet_json_cache.get('names') and _sheet_json_cache['mtime'] == mtime:
            return jsonify(_sheet_json_cache['names'])

    # Slow fallback: open workbook
    try:
        wb = load_workbook(report_file, read_only=True, data_only=True)
        names = wb.sheetnames
        wb.close()

        summary = []
        regions = []
        divisions = []
        areas = []
        branches = []

        for n in names:
            nl = n.lower()
            if nl.startswith('region_'):
                regions.append(n)
            elif nl.startswith('division_'):
                divisions.append(n)
            elif nl.startswith('area_'):
                areas.append(n)
            elif nl.startswith('branch_'):
                branches.append(n)
            else:
                summary.append(n)

        result = {
            'success': True,
            'total': len(names),
            'summary': summary,
            'regions': regions,
            'divisions': divisions,
            'areas': areas,
            'branches': branches,
        }
        with _cache_lock:
            _sheet_json_cache['names'] = result
            _sheet_json_cache['mtime'] = mtime
        return jsonify(result)
    except Exception as e:
        err = user_error(e, context='eod-report-sheets')
        return jsonify({'error': err['user_message']}), 500


# ---------------------------------------------------------------------------
# Cached workbook for report-sheet-data (avoids reopening 166-sheet file)
# ---------------------------------------------------------------------------
import threading as _threading
import re as _re

_HEX_COLOR_RE = _re.compile(r'^[0-9A-Fa-f]{6}$')

_report_wb_cache = {'wb': None, 'mtime': 0}
_sheet_json_cache = {'mtime': 0, 'data': {}, 'names': None}
_wb_loading = False  # True while background thread is loading
_wb_lock = _threading.Lock()  # Thread safety for workbook cache
_cache_lock = _threading.Lock()  # Thread safety for sheet JSON + email body caches
_attach_lock = _threading.Lock()  # Thread safety for attachment cache
_extraction_done = _threading.Event()  # Signals when sheet extraction finishes
_extraction_done.set()  # Initially "done" (no extraction pending)
_JSON_CACHE_FILE = BACKEND_DATA_DIR / 'EOD_Report_Cache.json'
_EMAIL_RE = _re.compile(r'^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$')


def _get_report_wb():
    """Return a cached openpyxl workbook; reloads only when file changes."""
    report_file = BACKEND_DATA_DIR / 'EOD_Report_Latest.xlsx'
    if not report_file.exists():
        return None
    mtime = report_file.stat().st_mtime
    with _wb_lock:
        if _report_wb_cache['wb'] is None or _report_wb_cache['mtime'] != mtime:
            if _report_wb_cache['wb']:
                try:
                    _report_wb_cache['wb'].close()
                except OSError:
                    pass  # workbook close cleanup
            _report_wb_cache['wb'] = load_workbook(report_file, read_only=False, data_only=True)
            _report_wb_cache['mtime'] = mtime
            # Acquire _cache_lock INSIDE _wb_lock (lock ordering: _wb_lock -> _cache_lock)
            with _cache_lock:
                _sheet_json_cache['mtime'] = mtime
                _sheet_json_cache['data'] = {}
                # Invalidate email body cache so it rebuilds from fresh extraction
                _email_body_cache['mtime'] = 0
                _email_body_cache['html'] = ''
        return _report_wb_cache['wb']


def _sanitize_sheet_colors(sheet_data):
    """Remove corrupted colour values (e.g. openpyxl theme-colour error strings)
    from a cached sheet JSON structure.  Modifies *sheet_data* in-place."""
    if not sheet_data:
        return sheet_data
    dirty = False
    for row in sheet_data.get('rows', []):
        for i, cell in enumerate(row):
            if not isinstance(cell, dict) or 'f' not in cell:
                continue
            fmt = cell['f']
            for key in ('bg', 'fc'):
                val = fmt.get(key)
                if val and not _HEX_COLOR_RE.match(val.lstrip('#')):
                    del fmt[key]
                    dirty = True
            # If fmt is now empty, flatten back to plain value
            if not fmt:
                row[i] = cell.get('v', '')
    if dirty:
        logging.info("Sanitized corrupted colour values in sidecar sheet data")
    return sheet_data


def _ensure_sidecar_loaded():
    """Load JSON sidecar from disk into memory cache (called by endpoints)."""
    with _cache_lock:
        if _sheet_json_cache.get('names'):
            return  # already in memory
        if not _JSON_CACHE_FILE.exists():
            return
        try:
            report_file = BACKEND_DATA_DIR / 'EOD_Report_Latest.xlsx'
            mtime = report_file.stat().st_mtime if report_file.exists() else 0
            cache = json.loads(_JSON_CACHE_FILE.read_text())
            if cache.get('mtime') == mtime:
                _sheet_json_cache['names'] = cache.get('names')
                _sheet_json_cache['mtime'] = mtime
                for k, v in cache.get('sheets', {}).items():
                    if v:
                        _sanitize_sheet_colors(v)
                        _sheet_json_cache['data'][k] = v
                logging.info("Loaded JSON sidecar into memory cache")
        except (json.JSONDecodeError, OSError) as e:
            logging.debug(f"Could not load JSON sidecar: {e}")


def _preload_report_wb():
    """Preload workbook + OverAll JSON + JSON sidecar in a background thread."""
    import threading
    global _wb_loading
    if _wb_loading:
        return
    report_file = BACKEND_DATA_DIR / 'EOD_Report_Latest.xlsx'
    if not report_file.exists():
        return
    mtime = report_file.stat().st_mtime
    # Already loaded and current
    if _report_wb_cache['wb'] is not None and _report_wb_cache['mtime'] == mtime:
        return

    def _load():
        global _wb_loading
        _wb_loading = True
        try:
            wb = _get_report_wb()
            if wb is None:
                return

            # Categorise sheet names
            names = wb.sheetnames
            summary, regions, divisions, areas, branches = [], [], [], [], []
            for n in names:
                nl = n.lower()
                if nl.startswith('region_'):
                    regions.append(n)
                elif nl.startswith('division_'):
                    divisions.append(n)
                elif nl.startswith('area_'):
                    areas.append(n)
                elif nl.startswith('branch_'):
                    branches.append(n)
                else:
                    summary.append(n)
            names_data = {
                'success': True,
                'total': len(names),
                'summary': summary,
                'regions': regions,
                'divisions': divisions,
                'areas': areas,
                'branches': branches,
            }
            _sheet_json_cache['names'] = names_data

            # Extract OverAll
            if 'OverAll' in wb.sheetnames and 'OverAll' not in _sheet_json_cache['data']:
                _sheet_json_cache['data']['OverAll'] = _extract_sheet_json(wb['OverAll'])

            # Save JSON sidecar to disk (survives server restarts)
            sidecar = {
                'mtime': mtime,
                'names': names_data,
                'sheets': {'OverAll': _sheet_json_cache['data'].get('OverAll')},
            }
            _JSON_CACHE_FILE.write_text(json.dumps(sidecar))
            logging.info(f"JSON sidecar built: {_JSON_CACHE_FILE.name} "
                         f"({_JSON_CACHE_FILE.stat().st_size / 1024:.0f}KB)")
        except Exception as exc:
            logging.warning(f"Background preload failed: {exc}")
        finally:
            _wb_loading = False

    threading.Thread(target=_load, daemon=True).start()


def _extract_sheet_json(ws):
    """Convert an openpyxl worksheet to JSON-serialisable rows + merges."""
    rows = []
    merges = [{'r1': m.min_row - 1, 'c1': m.min_col - 1,
               'r2': m.max_row - 1, 'c2': m.max_col - 1}
              for m in ws.merged_cells.ranges]

    for row in ws.iter_rows():
        cells = []
        for cell in row:
            val = cell.value
            if val is None:
                val = ''
            elif isinstance(val, float):
                val = round(val, 4)

            fmt = {}
            fill = cell.fill
            try:
                fill_rgb = fill.fgColor.rgb if (fill and fill.fgColor and fill.fgColor.rgb) else None
                if fill_rgb and isinstance(fill_rgb, str) and fill_rgb != '00000000':
                    rgb = fill_rgb[2:] if len(fill_rgb) == 8 else fill_rgb
                    if _HEX_COLOR_RE.match(rgb) and rgb != '000000':
                        fmt['bg'] = f'#{rgb}'
            except (TypeError, AttributeError):
                pass
            font = cell.font
            if font:
                if font.bold:
                    fmt['b'] = True
                try:
                    font_rgb = font.color.rgb if (font.color and font.color.rgb) else None
                    if font_rgb and isinstance(font_rgb, str) and font_rgb != '00000000':
                        fc = font_rgb[2:] if len(font_rgb) == 8 else font_rgb
                        if _HEX_COLOR_RE.match(fc) and fc != '000000':
                            fmt['fc'] = f'#{fc}'
                except (TypeError, AttributeError):
                    pass
            if cell.alignment and cell.alignment.horizontal == 'center':
                fmt['a'] = 'c'
            nf = cell.number_format
            if nf and '0.00%' in nf:
                if isinstance(val, (int, float)) and val != '':
                    val = round(float(val) * 100, 2)
                    fmt['pct'] = True
            elif nf and '#,##0' in nf:
                fmt['num'] = True

            cells.append({'v': val, 'f': fmt} if fmt else val)
        rows.append(cells)

    return {'success': True, 'sheet': ws.title, 'rows': rows, 'merges': merges}


@eod_bp.route('/report-sheet-data', methods=['GET'])
def report_sheet_data():
    """Return data + cell formatting from a specific sheet (cached after first load)."""
    report_file = BACKEND_DATA_DIR / 'EOD_Report_Latest.xlsx'
    if not report_file.exists():
        return jsonify({'error': 'No EOD Report available.'}), 404

    sheet_name = request.args.get('sheet')
    if not sheet_name:
        return jsonify({'error': 'Missing sheet parameter.'}), 400

    # Invalidate cache if file changed
    mtime = report_file.stat().st_mtime
    with _cache_lock:
        if mtime != _sheet_json_cache['mtime']:
            _sheet_json_cache['mtime'] = mtime
            _sheet_json_cache['data'] = {}
            _sheet_json_cache['names'] = None

        # Fast path: in-memory cache (< 1ms)
        if sheet_name in _sheet_json_cache['data']:
            return jsonify(_sheet_json_cache['data'][sheet_name])

    # Try JSON sidecar (< 5ms, covers OverAll after restart)
    _ensure_sidecar_loaded()
    with _cache_lock:
        if sheet_name in _sheet_json_cache['data']:
            return jsonify(_sheet_json_cache['data'][sheet_name])

    # Slow fallback: open workbook
    try:
        wb = _get_report_wb()
        if wb is None:
            return jsonify({'error': 'No EOD Report available.'}), 404
        if sheet_name not in wb.sheetnames:
            return jsonify({'error': f'Sheet "{sheet_name}" not found.'}), 404

        result = _extract_sheet_json(wb[sheet_name])
        with _cache_lock:
            _sheet_json_cache['data'][sheet_name] = result
        return jsonify(result)
    except Exception as e:
        err = user_error(e, context='eod-report-sheet-data')
        return jsonify({'error': err['user_message']}), 500


@eod_bp.route('/send-sheet-email', methods=['POST'])
@require_api_key
def send_sheet_email():
    """Legacy single-recipient endpoint (kept for backward compat)."""
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.base import MIMEBase
    from email.mime.text import MIMEText
    from email import encoders

    data = request.json
    sheets = data.get('sheets', [])
    recipient = data.get('recipient')

    gmail_user = config.GMAIL_USER
    gmail_pass = config.GMAIL_APP_PASSWORD

    if not recipient:
        return jsonify({'success': False, 'message': 'Missing recipient email.'}), 400
    if not _EMAIL_RE.match(recipient):
        return jsonify({'success': False, 'message': f'Invalid email address: {recipient}'}), 400
    if not sheets:
        return jsonify({'success': False, 'message': 'No sheets selected.'}), 400

    report_file = BACKEND_DATA_DIR / 'EOD_Report_Latest.xlsx'
    if not report_file.exists():
        return jsonify({'success': False, 'message': 'No EOD Report available.'}), 404

    try:
        today_str = _get_report_date_str()
        body_html = _get_email_body_html()

        # Build attachment (uses cache)
        attachment_bytes = _get_or_build_attachment(report_file, sheets)
        if attachment_bytes is None:
            return jsonify({'success': False, 'message': 'None of the selected sheets exist.'}), 400

        from email.mime.image import MIMEImage

        msg = MIMEMultipart('mixed')
        msg['From'] = gmail_user
        msg['To'] = recipient
        msg['Subject'] = f'EOD Report - {today_str}'

        body_img = BACKEND_DATA_DIR / 'eod_body.png'
        if body_img.exists():
            related = MIMEMultipart('related')
            related.attach(MIMEText(
                '<html><body><img src="cid:eod_body"></body></html>', 'html'))
            img_part = MIMEImage(body_img.read_bytes(), _subtype='png')
            img_part.add_header('Content-ID', '<eod_body>')
            img_part.add_header('Content-Disposition', 'inline')
            related.attach(img_part)
            msg.attach(related)
        else:
            msg.attach(MIMEText(body_html, 'html'))

        part = MIMEBase('application',
                        'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        part.set_payload(attachment_bytes)
        encoders.encode_base64(part)
        part.add_header('Content-Disposition',
                        f'attachment; filename="EOD_Report_{today_str}.xlsx"')
        msg.attach(part)

        server = None
        try:
            server = smtplib.SMTP(config.SMTP_HOST, config.SMTP_PORT, timeout=60)
            server.starttls()
            server.login(gmail_user, gmail_pass)
            server.send_message(msg)
        finally:
            if server:
                try:
                    server.quit()
                except OSError:
                    pass  # SMTP cleanup

        return jsonify({'success': True,
                        'message': f'Email sent to {recipient} with {len(sheets)} sheet(s).'})

    except Exception as e:
        err = user_error(e, context='eod-send-email')
        return jsonify({'success': False, 'message': err['user_message']}), 500


# ---------------------------------------------------------------------------
# Attachment cache: (file_mtime, frozenset(sheets)) -> raw xlsx bytes
# Avoids rebuilding identical attachments across requests/recipients
# ---------------------------------------------------------------------------
_attachment_cache = {}


def _get_or_build_attachment(report_file, sheets):
    """Build or retrieve cached xlsx bytes for a given sheet set."""
    mtime = report_file.stat().st_mtime
    # Use cached sheet names if available, else quick probe
    with _cache_lock:
        if _sheet_json_cache.get('names') and _sheet_json_cache['mtime'] == mtime:
            available = set()
            for cat in ('summary', 'regions', 'divisions', 'areas', 'branches'):
                available.update(_sheet_json_cache['names'].get(cat, []))
        else:
            available = None

    if available is None:
        wb_probe = load_workbook(report_file, read_only=True, data_only=True)
        available = set(wb_probe.sheetnames)
        wb_probe.close()

    requested = [s for s in sheets if s in available]
    if not requested:
        return None

    key = (mtime, frozenset(requested))
    with _attach_lock:
        if key in _attachment_cache:
            return _attachment_cache[key]

        # Evict stale entries from previous file versions
        for old_key in list(_attachment_cache):
            if old_key[0] != mtime:
                del _attachment_cache[old_key]

        # Evict oldest entries if cache exceeds 50 entries
        if len(_attachment_cache) >= 50:
            oldest_key = min(_attachment_cache, key=lambda k: k[0])  # k[0] is mtime
            del _attachment_cache[oldest_key]
            logging.info("Evicted oldest attachment cache entry (limit: 50)")

    tmp_path = None
    try:
        sheets_to_remove = available - set(requested)

        # If no sheets to remove, just read the combined file directly
        if not sheets_to_remove:
            raw_bytes = report_file.read_bytes()
        else:
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                tmp_path = tmp.name
            shutil.copy(report_file, tmp_path)
            wb_copy = load_workbook(tmp_path, data_only=True)
            for sn in list(wb_copy.sheetnames):
                if sn not in set(requested):
                    del wb_copy[sn]
            wb_copy.save(tmp_path)
            wb_copy.close()
            with open(tmp_path, 'rb') as f:
                raw_bytes = f.read()

        with _attach_lock:
            _attachment_cache[key] = raw_bytes
        return raw_bytes
    finally:
        if tmp_path:
            try:
                Path(tmp_path).unlink(missing_ok=True)
            except OSError:
                pass  # temp file cleanup


# ---------------------------------------------------------------------------
# V2 Email Queue: Helper functions
# ---------------------------------------------------------------------------

def _sse_event(data):
    """Format a dict as a Server-Sent Event data line."""
    return f"data: {json.dumps(data)}\n\n"


def _get_sheet_manifest(sheets_dir):
    """Load manifest.json from pre-extracted sheets dir and validate freshness.

    Returns the parsed manifest dict if valid and current, else None.
    The manifest is considered stale if:
      - the source workbook's mtime has drifted (re-generated after extraction)
      - the source file no longer exists at the recorded path
      - no actual sheet files exist in the directory
    """
    manifest_path = sheets_dir / 'manifest.json'
    if not manifest_path.exists():
        return None
    try:
        manifest = json.loads(manifest_path.read_text())
        source_rel = manifest.get('source', '')
        if source_rel:
            source_path = BACKEND_DATA_DIR / source_rel
            if not source_path.exists():
                return None  # source file gone — manifest is orphaned
            if abs(source_path.stat().st_mtime - manifest.get('source_mtime', 0)) > 0.01:
                return None  # stale — workbook changed since extraction

        # Verify at least one sheet file actually exists on disk
        sheet_entries = manifest.get('sheets', {})
        if not sheet_entries:
            return None
        has_any = any(
            (sheets_dir / info.get('path', '')).exists()
            for info in sheet_entries.values()
            if isinstance(info, dict)
        )
        if not has_any:
            return None  # manifest exists but no sheet files — orphaned

        return manifest
    except Exception:
        return None


def _merge_extracted_sheets(sheets_dir, manifest, sheet_names, mode):
    """Merge individual pre-extracted .xlsx sheet files into one workbook.

    For *combined* mode returns raw bytes of a single merged workbook.
    For *separate* mode returns a dict ``{sheet_name: raw_bytes}`` — one
    single-sheet workbook per requested sheet.
    """
    from openpyxl import Workbook, load_workbook as _lw

    if mode == 'separate':
        result = {}
        for name in sheet_names:
            sheet_info = manifest.get('sheets', {}).get(name)
            if sheet_info:
                sheet_path = sheets_dir / sheet_info['path']
                if sheet_path.exists():
                    result[name] = sheet_path.read_bytes()
        return result

    # Combined mode — merge into one workbook
    merged_wb = Workbook()
    merged_wb.remove(merged_wb.active)  # remove default empty sheet

    for name in sheet_names:
        sheet_info = manifest.get('sheets', {}).get(name)
        if not sheet_info:
            continue
        sheet_path = sheets_dir / sheet_info['path']
        if not sheet_path.exists():
            continue

        src_wb = _lw(sheet_path, data_only=True)
        src_ws = src_wb.active

        dest_ws = merged_wb.create_sheet(title=name)

        # Copy cell values and styles
        for row in src_ws.iter_rows():
            for cell in row:
                dest_cell = dest_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    dest_cell.font = cell.font.copy()
                    dest_cell.fill = cell.fill.copy()
                    dest_cell.border = cell.border.copy()
                    dest_cell.alignment = cell.alignment.copy()
                    dest_cell.number_format = cell.number_format

        # Copy merged cell ranges
        for merge_range in src_ws.merged_cells.ranges:
            dest_ws.merge_cells(str(merge_range))

        # Copy column dimensions (widths)
        for col_letter, dim in src_ws.column_dimensions.items():
            dest_ws.column_dimensions[col_letter].width = dim.width

        # Copy row dimensions (heights) if explicitly set
        for row_idx, dim in src_ws.row_dimensions.items():
            if dim.height is not None:
                dest_ws.row_dimensions[row_idx].height = dim.height

        src_wb.close()

    buf = io.BytesIO()
    merged_wb.save(buf)
    merged_wb.close()
    return buf.getvalue()


def _build_fast_attachment(sheets, mode, today_str, *, _batch_cache=None):
    """Build attachment bytes using pre-extracted sheet files when available.

    Fast path: merge small individual .xlsx files from ``data/backend/sheets/``.
    Fallback:  use the legacy ``_get_or_build_attachment`` (copy full workbook,
               delete unwanted sheets).

    *_batch_cache* is an optional dict used within a single batch run to avoid
    rebuilding identical sheet-set attachments across recipients.

    Returns:
        - ``bytes`` for combined mode
        - ``dict[str, bytes]`` for separate mode
    """
    cache_key = (frozenset(sheets), mode)
    if _batch_cache is not None and cache_key in _batch_cache:
        return _batch_cache[cache_key]

    sheets_dir = BACKEND_DATA_DIR / 'sheets'
    manifest = _get_sheet_manifest(sheets_dir)

    report_file = BACKEND_DATA_DIR / 'EOD_Report_Latest.xlsx'
    if not report_file.exists():
        return {} if mode == 'separate' else None

    if manifest:
        # ── FAST PATH: merge pre-extracted sheets ──
        result = _merge_extracted_sheets(sheets_dir, manifest, sheets, mode)
    elif mode == 'combined':
        # ── SLOW FALLBACK: no manifest, copying full workbook ──
        logging.warning("Sheet manifest not found — using slow attachment fallback")
        result = _get_or_build_attachment(report_file, sheets)
    else:
        # ── SEPARATE mode without manifest: extract only requested sheets ──
        logging.warning("Sheet manifest not found — using slow separate-mode fallback")
        result = {}
        for sheet_name in sheets:
            raw = _get_or_build_attachment(report_file, [sheet_name])
            if raw:
                result[sheet_name] = raw

    if _batch_cache is not None:
        _batch_cache[cache_key] = result
    return result


def _build_mime_message(from_addr, to_addr, today_str, body_html,
                        attachment_data, sheets, mode, body_img_bytes=None):
    """Construct a complete MIME email with attachment(s).

    *attachment_data* is either raw ``bytes`` (combined mode) or a
    ``dict[str, bytes]`` (separate mode).
    *body_img_bytes* — pre-read PNG bytes for the inline body image (avoids
    re-reading from disk for every email).
    """
    from email.mime.multipart import MIMEMultipart
    from email.mime.base import MIMEBase
    from email.mime.text import MIMEText
    from email import encoders

    from email.mime.image import MIMEImage

    msg = MIMEMultipart('mixed')
    msg['From'] = from_addr
    msg['To'] = to_addr
    msg['Subject'] = f'EOD Report - {today_str}'

    # Body: inline image if available, else HTML table
    if body_img_bytes:
        related = MIMEMultipart('related')
        related.attach(MIMEText(
            '<html><body><img src="cid:eod_body"></body></html>', 'html'))
        img_part = MIMEImage(body_img_bytes, _subtype='png')
        img_part.add_header('Content-ID', '<eod_body>')
        img_part.add_header('Content-Disposition', 'inline')
        related.attach(img_part)
        msg.attach(related)
    else:
        msg.attach(MIMEText(body_html, 'html'))

    if mode == 'separate' and isinstance(attachment_data, dict):
        for sheet_name in sheets:
            raw = attachment_data.get(sheet_name)
            if not raw:
                continue
            part = MIMEBase(
                'application',
                'vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            part.set_payload(raw)
            encoders.encode_base64(part)
            part.add_header(
                'Content-Disposition',
                f'attachment; filename="{sheet_name}.xlsx"',
            )
            msg.attach(part)
    elif attachment_data and isinstance(attachment_data, bytes):
        part = MIMEBase(
            'application',
            'vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        part.set_payload(attachment_data)
        encoders.encode_base64(part)
        part.add_header(
            'Content-Disposition',
            f'attachment; filename="EOD_Report_{today_str}.xlsx"',
        )
        msg.attach(part)
    else:
        return None  # nothing to attach

    return msg


def _reconnect_smtp(gmail_user, gmail_pass, timeout):
    """Create a fresh authenticated SMTP connection."""
    import smtplib
    server = smtplib.SMTP(config.SMTP_HOST, config.SMTP_PORT, timeout=timeout)
    server.starttls()
    server.login(gmail_user, gmail_pass)
    return server


# ---------------------------------------------------------------------------
# V2 batch email endpoint (replaces legacy two-phase approach)
# ---------------------------------------------------------------------------

@eod_bp.route('/send-batch-email', methods=['POST'])
@require_api_key
def send_batch_email():
    """V2: One-by-one email sending with pre-extracted sheet merging.

    Builds each recipient's attachment on-the-fly from small pre-extracted
    sheet files (fast path) or falls back to the legacy full-workbook copy
    when sheets/ directory is unavailable.  Streams SSE progress per email.

    Request body: ``{recipients: [{email, sheets[], mode?}]}``
    """
    import smtplib
    import time as _time

    data = request.json
    recipients = data.get('recipients', [])

    gmail_user = config.GMAIL_USER
    gmail_pass = config.GMAIL_APP_PASSWORD

    if not recipients:
        return jsonify({'success': False, 'message': 'No recipients provided.'}), 400

    # ── Validate ALL upfront (fail fast) ──
    errors = []
    for i, r in enumerate(recipients):
        if not r.get('email') or not _EMAIL_RE.match(r['email']):
            errors.append(f"Recipient {i+1}: invalid email")
        if not r.get('sheets'):
            errors.append(f"Recipient {i+1}: no sheets")
    if errors:
        return jsonify({'success': False, 'message': '; '.join(errors[:5])}), 400

    report_file = BACKEND_DATA_DIR / 'EOD_Report_Latest.xlsx'
    if not report_file.exists():
        return jsonify({'success': False, 'message': 'No EOD Report available.'}), 404

    today_str = _get_report_date_str()
    body_html = _get_email_body_html()
    total = len(recipients)

    def generate():
        batch_start = _time.time()
        BATCH_DEADLINE_SECONDS = 30 * 60   # 30-minute hard ceiling
        EMAIL_TIMEOUT_SECONDS = 60         # per-connection SMTP timeout

        server = None
        sent = 0
        failed = 0
        fail_list = []      # [(orig_idx, email_addr, sheets, mode)]
        batch_cache = {}     # frozenset(sheets)+mode -> attachment bytes (per-batch)

        try:
            # ── Connect SMTP ──
            yield _sse_event({
                'phase': 'connecting',
                'message': 'Connecting to Gmail...',
                'total': total,
            })

            server = smtplib.SMTP(config.SMTP_HOST, config.SMTP_PORT,
                                  timeout=EMAIL_TIMEOUT_SECONDS)
            server.starttls()
            server.login(gmail_user, gmail_pass)

            yield _sse_event({
                'phase': 'connected',
                'message': 'Connected. Starting sends...',
                'total': total,
            })

            # ── Wait for sheet extraction if still running ──
            if not _extraction_done.is_set():
                yield _sse_event({
                    'phase': 'preparing',
                    'message': 'Waiting for sheet extraction to finish...',
                    'total': total,
                })
                # Wait up to 10 minutes, sending keepalives every 5s
                waited = 0
                while not _extraction_done.is_set() and waited < 600:
                    _extraction_done.wait(timeout=5)
                    waited += 5
                    if not _extraction_done.is_set():
                        yield _sse_event({
                            'phase': 'preparing',
                            'message': f'Sheet extraction in progress... ({waited}s)',
                            'total': total,
                        })
                if _extraction_done.is_set():
                    yield _sse_event({
                        'phase': 'connected',
                        'message': 'Sheets ready. Starting sends...',
                        'total': total,
                    })

            # Pre-read body image once for all emails
            _cached_body_img = None
            body_img_path = BACKEND_DATA_DIR / 'eod_body.png'
            if body_img_path.exists():
                _cached_body_img = body_img_path.read_bytes()
                logging.info(f"Body image cached: {len(_cached_body_img)//1024}KB")

            # ── Send one-by-one ──
            for idx, r in enumerate(recipients):
                # Deadline check
                if _time.time() - batch_start > BATCH_DEADLINE_SECONDS:
                    yield _sse_event({
                        'phase': 'deadline',
                        'message': 'Batch deadline exceeded (30 min)',
                        'sent': sent,
                        'failed': failed + (total - idx),
                        'total': total,
                        'done': True,
                    })
                    return

                email_addr = r['email']
                sheets = r['sheets']
                mode = r.get('mode', 'combined')
                email_start = _time.time()
                msg = None  # reset per-recipient

                # Phase: building attachment
                yield _sse_event({
                    'phase': 'building',
                    'index': idx + 1,
                    'total': total,
                    'email': email_addr,
                })

                try:
                    # Build attachment (fast path or fallback)
                    build_start = _time.time()
                    attachment_data = _build_fast_attachment(
                        sheets, mode, today_str, _batch_cache=batch_cache,
                    )
                    build_elapsed = _time.time() - build_start

                    # Skip if nothing could be built
                    if not attachment_data:
                        raise ValueError('No valid sheets found for this recipient')

                    # Build MIME message
                    mime_start = _time.time()
                    msg = _build_mime_message(
                        gmail_user, email_addr, today_str, body_html,
                        attachment_data, sheets, mode,
                        body_img_bytes=_cached_body_img,
                    )
                    mime_elapsed = _time.time() - mime_start
                    if msg is None:
                        raise ValueError('Failed to build email (no attachable data)')

                    # Phase: sending
                    yield _sse_event({
                        'phase': 'sending',
                        'index': idx + 1,
                        'total': total,
                        'email': email_addr,
                    })

                    send_start = _time.time()
                    server.send_message(msg)
                    send_elapsed = _time.time() - send_start

                    logging.info(f"Email {idx+1}/{total} to {email_addr}: "
                                 f"build={build_elapsed:.2f}s mime={mime_elapsed:.2f}s send={send_elapsed:.2f}s")

                    elapsed = round(_time.time() - email_start, 1)
                    sent += 1
                    yield _sse_event({
                        'phase': 'sent',
                        'index': idx + 1,
                        'total': total,
                        'email': email_addr,
                        'elapsed': elapsed,
                    })

                except smtplib.SMTPServerDisconnected:
                    # Attempt reconnect + retry this email
                    try:
                        server = _reconnect_smtp(
                            gmail_user, gmail_pass, EMAIL_TIMEOUT_SECONDS,
                        )
                        # Rebuild message if needed (msg may exist from above)
                        if msg is not None:
                            server.send_message(msg)
                            elapsed = round(_time.time() - email_start, 1)
                            sent += 1
                            yield _sse_event({
                                'phase': 'sent',
                                'index': idx + 1,
                                'total': total,
                                'email': email_addr,
                                'elapsed': elapsed,
                                'reconnected': True,
                            })
                        else:
                            raise ValueError('Message was not built before disconnect')
                    except Exception as reconn_err:
                        elapsed = round(_time.time() - email_start, 1)
                        failed += 1
                        fail_list.append((idx, email_addr, sheets, mode))
                        yield _sse_event({
                            'phase': 'failed',
                            'index': idx + 1,
                            'total': total,
                            'email': email_addr,
                            'error': str(reconn_err),
                            'elapsed': elapsed,
                        })

                except Exception as e:
                    elapsed = round(_time.time() - email_start, 1)
                    failed += 1
                    fail_list.append((idx, email_addr, sheets, mode))
                    yield _sse_event({
                        'phase': 'failed',
                        'index': idx + 1,
                        'total': total,
                        'email': email_addr,
                        'error': str(e),
                        'elapsed': elapsed,
                    })

            # ── Retry failed ones ONCE ──
            if fail_list:
                yield _sse_event({
                    'phase': 'retrying',
                    'message': f'Retrying {len(fail_list)} failed email(s)...',
                    'count': len(fail_list),
                })

                # Ensure SMTP connection is alive for retries
                try:
                    server.noop()
                except Exception:
                    try:
                        server = _reconnect_smtp(
                            gmail_user, gmail_pass, EMAIL_TIMEOUT_SECONDS,
                        )
                    except Exception as reconn_err:
                        yield _sse_event({
                            'phase': 'retry_connect_failed',
                            'message': f'Could not reconnect for retries: {reconn_err}',
                        })
                        fail_list_copy = list(fail_list)
                        fail_list = fail_list_copy  # keep original failures
                        # Skip retry loop — jump to summary
                        fail_list = []  # clear so loop below is skipped
                        # (failed count already includes these)

                retry_success = 0
                still_failed = []
                for (orig_idx, email_addr, sheets, mode) in fail_list:
                    retry_start = _time.time()
                    try:
                        attachment_data = _build_fast_attachment(
                            sheets, mode, today_str, _batch_cache=batch_cache,
                        )
                        msg = _build_mime_message(
                            gmail_user, email_addr, today_str, body_html,
                            attachment_data, sheets, mode,
                            )
                        if msg is None:
                            raise ValueError('Failed to build retry email')
                        server.send_message(msg)
                        elapsed = round(_time.time() - retry_start, 1)
                        retry_success += 1
                        yield _sse_event({
                            'phase': 'retry_sent',
                            'index': orig_idx + 1,
                            'total': total,
                            'email': email_addr,
                            'elapsed': elapsed,
                        })
                    except Exception as retry_err:
                        elapsed = round(_time.time() - retry_start, 1)
                        still_failed.append(email_addr)
                        yield _sse_event({
                            'phase': 'retry_failed',
                            'index': orig_idx + 1,
                            'total': total,
                            'email': email_addr,
                            'error': str(retry_err),
                            'elapsed': elapsed,
                        })

                # Adjust counters
                sent += retry_success
                failed -= retry_success

        except smtplib.SMTPAuthenticationError:
            yield _sse_event({
                'phase': 'auth_error',
                'message': 'Gmail authentication failed. Check credentials.',
                'done': True,
                'sent': sent,
                'failed': total - sent,
            })
            return
        except Exception as e:
            yield _sse_event({
                'phase': 'error',
                'message': str(e),
                'done': True,
                'sent': sent,
                'failed': total - sent,
            })
            return
        finally:
            if server:
                try:
                    server.quit()
                except OSError:
                    pass  # SMTP cleanup on generator exit

        total_elapsed = round(_time.time() - batch_start, 1)
        yield _sse_event({
            'phase': 'complete',
            'sent': sent,
            'failed': failed,
            'total': total,
            'total_elapsed': total_elapsed,
            'done': True,
        })

    return Response(
        generate(),
        mimetype='text/event-stream',
        headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'},
    )


# ---------------------------------------------------------------------------
# Legacy batch email (V1) — kept as internal fallback
# ---------------------------------------------------------------------------

def _send_batch_email_legacy():
    """V1 (legacy): Two-phase batch email — build ALL attachments first, then
    send all over ONE SMTP connection.  Kept for fallback if V2 encounters
    issues.  Not registered to any route by default.
    """
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.base import MIMEBase
    from email.mime.text import MIMEText
    from email import encoders

    data = request.json
    recipients = data.get('recipients', [])

    gmail_user = config.GMAIL_USER
    gmail_pass = config.GMAIL_APP_PASSWORD

    if not recipients:
        return jsonify({'success': False, 'message': 'No recipients provided.'}), 400

    # Validate ALL upfront (fail fast)
    errors = []
    for i, r in enumerate(recipients):
        if not r.get('email') or not _EMAIL_RE.match(r['email']):
            errors.append(f"Recipient {i+1}: invalid email")
        if not r.get('sheets'):
            errors.append(f"Recipient {i+1}: no sheets")
    if errors:
        return jsonify({'success': False, 'message': '; '.join(errors[:5])}), 400

    report_file = BACKEND_DATA_DIR / 'EOD_Report_Latest.xlsx'
    if not report_file.exists():
        return jsonify({'success': False, 'message': 'No EOD Report available.'}), 404

    today_str = _get_report_date_str()
    body_html = _get_email_body_html()

    # ── PHASE 1: Build unique attachments (deduplicated), per-recipient mode ──
    attachment_keys = set()
    for r in recipients:
        mode = r.get('mode', 'combined')
        sheets = r['sheets']
        if mode == 'separate':
            for sheet_name in sheets:
                attachment_keys.add(frozenset([sheet_name]))
        else:
            attachment_keys.add(frozenset(sheets))

    attachment_map = {}
    try:
        for sheet_set in attachment_keys:
            raw = _get_or_build_attachment(report_file, list(sheet_set))
            if raw:
                attachment_map[sheet_set] = raw
    except Exception as e:
        logging.warning(f"Batch attachment build failed: {e}")
        return jsonify({'success': False, 'message': f'Failed to prepare attachments: {e}'}), 500

    messages = []
    for r in recipients:
        email_addr = r['email']
        sheets = r['sheets']
        mode = r.get('mode', 'combined')

        from email.mime.image import MIMEImage

        msg = MIMEMultipart('mixed')
        msg['From'] = gmail_user
        msg['To'] = email_addr
        msg['Subject'] = f'EOD Report - {today_str}'

        body_img = BACKEND_DATA_DIR / 'eod_body.png'
        if body_img.exists():
            related = MIMEMultipart('related')
            related.attach(MIMEText(
                '<html><body><img src="cid:eod_body"></body></html>', 'html'))
            img_part = MIMEImage(body_img.read_bytes(), _subtype='png')
            img_part.add_header('Content-ID', '<eod_body>')
            img_part.add_header('Content-Disposition', 'inline')
            related.attach(img_part)
            msg.attach(related)
        else:
            msg.attach(MIMEText(body_html, 'html'))

        attached_any = False

        if mode == 'separate':
            for sheet_name in sheets:
                raw = attachment_map.get(frozenset([sheet_name]))
                if not raw:
                    continue
                part = MIMEBase('application',
                                'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                part.set_payload(raw)
                encoders.encode_base64(part)
                part.add_header('Content-Disposition',
                                f'attachment; filename="{sheet_name}.xlsx"')
                msg.attach(part)
                attached_any = True
        else:
            raw = attachment_map.get(frozenset(sheets))
            if raw:
                part = MIMEBase('application',
                                'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                part.set_payload(raw)
                encoders.encode_base64(part)
                part.add_header('Content-Disposition',
                                f'attachment; filename="EOD_Report_{today_str}.xlsx"')
                msg.attach(part)
                attached_any = True

        if attached_any:
            messages.append((email_addr, msg))

    total = len(messages)
    if total == 0:
        return jsonify({'success': False, 'message': 'No valid sheets to send.'}), 400

    # ── PHASE 2: Send all emails over ONE SMTP connection, streaming progress ──
    def generate():
        import time as _time
        batch_start = _time.time()
        BATCH_DEADLINE_SECONDS = 30 * 60
        EMAIL_TIMEOUT_SECONDS = 60

        server = None
        sent = 0
        failed = 0
        fail_list = []
        try:
            yield f"data: {json.dumps({'phase': 'connecting', 'message': 'Connecting to Gmail...', 'progress': 5})}\n\n"

            server = smtplib.SMTP(config.SMTP_HOST, config.SMTP_PORT, timeout=EMAIL_TIMEOUT_SECONDS)
            server.starttls()
            server.login(gmail_user, gmail_pass)

            yield f"data: {json.dumps({'phase': 'sending', 'message': 'Connected. Sending...', 'progress': 10})}\n\n"

            for idx, (email_addr, msg) in enumerate(messages):
                if _time.time() - batch_start > BATCH_DEADLINE_SECONDS:
                    yield f"data: {json.dumps({'phase': 'error', 'message': 'Batch deadline exceeded (30 min).', 'done': True, 'sent': sent, 'failed': total - sent})}\n\n"
                    return

                pct = 10 + int(85 * (idx + 1) / total)
                try:
                    server.send_message(msg)
                    sent += 1
                    yield f"data: {json.dumps({'phase': 'sending', 'index': idx + 1, 'total': total, 'email': email_addr, 'status': 'sent', 'progress': pct})}\n\n"
                except smtplib.SMTPServerDisconnected:
                    try:
                        server = smtplib.SMTP(config.SMTP_HOST, config.SMTP_PORT, timeout=EMAIL_TIMEOUT_SECONDS)
                        server.starttls()
                        server.login(gmail_user, gmail_pass)
                        server.send_message(msg)
                        sent += 1
                        yield f"data: {json.dumps({'phase': 'sending', 'index': idx + 1, 'total': total, 'email': email_addr, 'status': 'sent', 'progress': pct})}\n\n"
                    except Exception as retry_err:
                        failed += 1
                        fail_list.append(email_addr)
                        yield f"data: {json.dumps({'phase': 'sending', 'index': idx + 1, 'total': total, 'email': email_addr, 'status': 'failed', 'error': str(retry_err), 'progress': pct})}\n\n"
                except Exception as e:
                    failed += 1
                    fail_list.append(email_addr)
                    yield f"data: {json.dumps({'phase': 'sending', 'index': idx + 1, 'total': total, 'email': email_addr, 'status': 'failed', 'error': str(e), 'progress': pct})}\n\n"

        except smtplib.SMTPAuthenticationError:
            yield f"data: {json.dumps({'phase': 'error', 'message': 'Gmail authentication failed.', 'done': True, 'sent': sent, 'failed': total - sent})}\n\n"
            return
        except Exception as e:
            yield f"data: {json.dumps({'phase': 'error', 'message': str(e), 'done': True, 'sent': sent, 'failed': total - sent})}\n\n"
            return
        finally:
            if server:
                try:
                    server.quit()
                except OSError:
                    pass  # SMTP cleanup

        yield f"data: {json.dumps({'phase': 'complete', 'done': True, 'sent': sent, 'failed': failed, 'progress': 100})}\n\n"

    return Response(generate(), mimetype='text/event-stream',
                    headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'})


# ---------------------------------------------------------------------------
# Precomputed email body HTML (built from cached sheet JSON)
# ---------------------------------------------------------------------------
_email_body_cache = {'mtime': 0, 'html': ''}


def _build_email_html_from_json(sheet_data):
    """Convert _extract_sheet_json output into an inline-styled HTML table for email.

    Reuses the same cell data + formatting that powers the live preview,
    so colors, bold, merged cells, number formatting all carry over.
    """
    rows = sheet_data.get('rows', [])
    merges = sheet_data.get('merges', [])
    if not rows:
        return '<p>EOD Report attached.</p>'

    # Build merge lookup: {(r,c): {rs, cs}} for origins, {(r,c): skip} for hidden
    mm = {}
    for m in merges:
        mm[(m['r1'], m['c1'])] = {'rs': m['r2'] - m['r1'] + 1,
                                   'cs': m['c2'] - m['c1'] + 1}
        for r in range(m['r1'], m['r2'] + 1):
            for c in range(m['c1'], m['c2'] + 1):
                if r != m['r1'] or c != m['c1']:
                    mm[(r, c)] = {'skip': True}

    # Detect first-section boundary (stop at 2+ consecutive blank rows)
    blank_run = 0
    section_end = len(rows)
    found_data = False
    for ri, row in enumerate(rows):
        all_empty = all(
            (cell == '' or cell is None or (isinstance(cell, dict) and (cell.get('v', '') == '' or cell.get('v') is None)))
            for cell in row
        )
        if all_empty:
            blank_run += 1
            if blank_run >= 2 and found_data:
                section_end = ri - blank_run + 1
                break
        else:
            blank_run = 0
            found_data = True

    html_parts = [
        '<table style="border-collapse:collapse;font-family:Calibri,Arial,sans-serif;'
        'font-size:11px;white-space:nowrap;">'
    ]

    for ri in range(section_end):
        row = rows[ri]
        html_parts.append('<tr>')
        for ci, cell in enumerate(row):
            key = (ri, ci)
            if key in mm and mm[key].get('skip'):
                continue

            # Parse cell value + format
            v = ''
            fmt = {}
            if isinstance(cell, dict) and 'v' in cell:
                v = cell['v']
                fmt = cell.get('f', {})
            else:
                v = cell
            if v is None:
                v = ''

            # Number formatting
            is_pct = fmt.get('pct', False)
            is_num = fmt.get('num', False)
            if is_pct and isinstance(v, (int, float)):
                v = f'{v:.2f}%'
            elif is_num and isinstance(v, (int, float)):
                v = f'{v:,.0f}'
            elif isinstance(v, float):
                if v == int(v) and abs(v) >= 1:
                    v = f'{int(v):,}'
                else:
                    v = f'{v:.2f}'
            v = str(v)

            # Inline styles — validate color values (sidecar may contain
            # corrupted theme-colour strings from older openpyxl extractions)
            parts = []
            bg = fmt.get('bg')
            if bg and _HEX_COLOR_RE.match(bg.lstrip('#')):
                parts.append(f'background:{bg}')
            fc = fmt.get('fc')
            if fc and _HEX_COLOR_RE.match(fc.lstrip('#')):
                parts.append(f'color:{fc}')
            else:
                parts.append('color:#000')
            if fmt.get('b'):
                parts.append('font-weight:bold')
            if fmt.get('a') == 'c':
                parts.append('text-align:center')
            parts.append('border:1px solid #bbb')
            parts.append('padding:3px 6px')

            style = ';'.join(parts)
            attrs = f' style="{style}"'

            mg = mm.get(key)
            if mg:
                if mg.get('rs', 1) > 1:
                    attrs += f' rowspan="{mg["rs"]}"'
                if mg.get('cs', 1) > 1:
                    attrs += f' colspan="{mg["cs"]}"'

            # Escape HTML
            v = v.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            html_parts.append(f'<td{attrs}>{v}</td>')

        html_parts.append('</tr>')

    html_parts.append('</table>')
    return ''.join(html_parts)


def _get_email_body_html():
    """Return precomputed email body HTML (uses cache)."""
    report_file = BACKEND_DATA_DIR / 'EOD_Report_Latest.xlsx'
    if not report_file.exists():
        return '<p>EOD Report attached.</p>'
    mtime = report_file.stat().st_mtime
    with _cache_lock:
        if _email_body_cache['mtime'] == mtime and _email_body_cache['html']:
            return _email_body_cache['html']

    # Try sidecar / in-memory cache first (avoids opening workbook)
    _ensure_sidecar_loaded()
    with _cache_lock:
        sheet_data = _sheet_json_cache['data'].get('OverAll')

    if not sheet_data:
        # Fallback: open workbook (acquires _wb_lock internally, then _cache_lock -- correct order)
        wb = _get_report_wb()
        if wb is None or 'OverAll' not in wb.sheetnames:
            return '<p>EOD Report attached.</p>'
        sheet_data = _extract_sheet_json(wb['OverAll'])
        with _cache_lock:
            _sheet_json_cache['data']['OverAll'] = sheet_data

    html = _build_email_html_from_json(sheet_data)
    with _cache_lock:
        _email_body_cache['mtime'] = mtime
        _email_body_cache['html'] = html
    return html


def _render_body_image():
    """Render the OverAll sheet to data/backend/eod_body.png using Pillow.

    Deletes any existing image first so at most 1 file exists.
    """
    from PIL import Image, ImageDraw, ImageFont

    # Get sheet data
    _ensure_sidecar_loaded()
    with _cache_lock:
        sheet_data = _sheet_json_cache['data'].get('OverAll')
    if not sheet_data:
        wb = _get_report_wb()
        if wb is None or 'OverAll' not in wb.sheetnames:
            return
        sheet_data = _extract_sheet_json(wb['OverAll'])

    rows = sheet_data.get('rows', [])
    merges = sheet_data.get('merges', [])
    if not rows:
        return

    out_path = BACKEND_DATA_DIR / 'eod_body.png'
    try:
        if out_path.exists():
            out_path.unlink()
    except OSError:
        pass

    try:
        # Build merge lookup
        mm = {}
        for m in merges:
            mm[(m['r1'], m['c1'])] = {'rs': m['r2'] - m['r1'] + 1, 'cs': m['c2'] - m['c1'] + 1}
            for r in range(m['r1'], m['r2'] + 1):
                for c in range(m['c1'], m['c2'] + 1):
                    if r != m['r1'] or c != m['c1']:
                        mm[(r, c)] = {'skip': True}

        # Detect first section (stop at 2+ blank rows)
        blank_run = 0
        section_end = len(rows)
        found_data = False
        for ri, row in enumerate(rows):
            all_empty = all(
                (cell == '' or cell is None or
                 (isinstance(cell, dict) and (cell.get('v', '') == '' or cell.get('v') is None)))
                for cell in row
            )
            if all_empty:
                blank_run += 1
                if blank_run >= 2 and found_data:
                    section_end = ri - blank_run + 1
                    break
            else:
                blank_run = 0
                found_data = True

        font = ImageFont.truetype('Arial', 11)
        font_bold = ImageFont.truetype('Arial Bold', 11)
        pad_x, pad_y = 5, 3

        # Parse all cell values first to measure column widths
        grid = []  # grid[ri][ci] = {text, bold, bg, fc, skip, rs, cs}
        num_cols = max(len(r) for r in rows[:section_end]) if rows[:section_end] else 0
        for ri in range(section_end):
            row = rows[ri]
            row_cells = []
            for ci in range(num_cols):
                cell = row[ci] if ci < len(row) else ''
                key = (ri, ci)
                if key in mm and mm[key].get('skip'):
                    row_cells.append({'text': '', 'skip': True})
                    continue

                v, fmt = '', {}
                if isinstance(cell, dict) and 'v' in cell:
                    v, fmt = cell['v'], cell.get('f', {})
                else:
                    v = cell
                if v is None:
                    v = ''

                if fmt.get('pct') and isinstance(v, (int, float)):
                    v = f'{v:.2f}%'
                elif fmt.get('num') and isinstance(v, (int, float)):
                    v = f'{v:,.0f}'
                elif isinstance(v, float):
                    v = f'{int(v):,}' if v == int(v) and abs(v) >= 1 else f'{v:.2f}'
                v = str(v)

                mg = mm.get(key, {})
                row_cells.append({
                    'text': v, 'bold': fmt.get('b', False),
                    'bg': fmt.get('bg'), 'fc': fmt.get('fc', '#000'),
                    'rs': mg.get('rs', 1), 'cs': mg.get('cs', 1), 'skip': False,
                })
            grid.append(row_cells)

        # Measure column widths
        col_widths = [0] * num_cols
        for ri, row_cells in enumerate(grid):
            for ci, cell in enumerate(row_cells):
                if cell.get('skip'):
                    continue
                cs = cell.get('cs', 1)
                f = font_bold if cell.get('bold') else font
                tw = f.getlength(cell['text']) if cell['text'] else 0
                w = int(tw) + pad_x * 2
                if cs == 1:
                    col_widths[ci] = max(col_widths[ci], w)
                # For merged cells, distribute width later

        # Ensure min width
        col_widths = [max(w, 30) for w in col_widths]

        row_height = 18
        total_w = sum(col_widths) + 1
        total_h = section_end * row_height + 1

        img = Image.new('RGB', (total_w, total_h), 'white')
        draw = ImageDraw.Draw(img)

        # Draw cells
        y = 0
        for ri, row_cells in enumerate(grid):
            x = 0
            for ci, cell in enumerate(row_cells):
                cw = col_widths[ci]
                if cell.get('skip'):
                    x += cw
                    continue

                cs = cell.get('cs', 1)
                rs = cell.get('rs', 1)
                cell_w = sum(col_widths[ci:ci + cs])
                cell_h = rs * row_height

                # Background
                bg = cell.get('bg')
                if bg and _HEX_COLOR_RE.match(bg.lstrip('#')):
                    if not bg.startswith('#'):
                        bg = '#' + bg
                    draw.rectangle([x, y, x + cell_w, y + cell_h], fill=bg)

                # Border
                draw.rectangle([x, y, x + cell_w, y + cell_h], outline='#bbbbbb')

                # Text
                txt = cell['text']
                if txt:
                    f = font_bold if cell.get('bold') else font
                    fc = cell.get('fc', '#000')
                    if not (fc and _HEX_COLOR_RE.match(fc.lstrip('#'))):
                        fc = '#000'
                    if not fc.startswith('#'):
                        fc = '#' + fc
                    draw.text((x + pad_x, y + pad_y), txt, fill=fc, font=f)

                x += cw
            y += row_height

        img.save(str(out_path), 'PNG')
        size_kb = out_path.stat().st_size / 1024
        logging.info(f"Email body image saved: {out_path} ({size_kb:.0f}KB)")
    except Exception as e:
        logging.warning(f"Body image render failed (non-fatal): {e}")


@eod_bp.route('/precompute-email-body', methods=['POST'])
def precompute_email_body():
    """Precompute the OverAll HTML for email body. Called when email page opens."""
    html = _get_email_body_html()
    # Also render body image in background
    import threading
    threading.Thread(target=_render_body_image, daemon=True).start()
    size_kb = len(html.encode('utf-8')) / 1024
    return jsonify({'success': True, 'size_kb': round(size_kb, 1)})


@eod_bp.route('/email-preview')
def email_preview():
    """Serve the email preview page. Kicks off background preload of workbook."""
    _preload_report_wb()
    resp = send_from_directory(str(STATIC_EOD_DIR), 'email_preview.html')
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    return resp


@eod_bp.route('/email-config', methods=['GET'])
def get_email_config():
    """Load email recipients + sheet connections from CSV."""
    if not EMAIL_CONFIG_CSV.exists():
        return jsonify({'cards': [], 'conns': [], 'cid': 0})

    try:
        cards = {}   # id -> email
        modes = {}   # id -> mode ('combined' or 'separate')
        conns = []
        cid = 0
        with open(EMAIL_CONFIG_CSV, newline='') as f:
            reader = csv.DictReader(f)
            for row in reader:
                card_id = row['card_id']
                email = row['email']
                sheet = row.get('sheet', '')
                cards[card_id] = email
                modes[card_id] = row.get('mode', 'combined') or 'combined'
                num = int(card_id.lstrip('c') or '0')
                if num > cid:
                    cid = num
                if sheet:
                    conns.append({'id': f"x{card_id}_{hashlib.md5(sheet.encode()).hexdigest()[:8]}",
                                  'sheet': sheet, 'cardId': card_id})

        cards_list = [{'id': k, 'email': v, 'mode': modes.get(k, 'combined')} for k, v in cards.items()]
        return jsonify({'cards': cards_list, 'conns': conns, 'cid': cid})
    except Exception as e:
        return jsonify({'cards': [], 'conns': [], 'cid': 0,
                        'warning': str(e)})


@eod_bp.route('/email-config', methods=['POST'])
def save_email_config():
    """Save email recipients + sheet connections to CSV."""
    try:
        data = request.json
        cards = data.get('cards', [])
        conns = data.get('conns', [])

        BACKEND_DATA_DIR.mkdir(parents=True, exist_ok=True)
        with open(EMAIL_CONFIG_CSV, 'w', newline='') as f:
            w = csv.writer(f)
            w.writerow(['card_id', 'email', 'sheet', 'timestamp', 'mode'])
            ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # Build card lookups
            card_map = {c['id']: c['email'] for c in cards}
            mode_map = {c['id']: c.get('mode', 'combined') for c in cards}

            # Write one row per connection
            written_cards = set()
            for conn in conns:
                cid = conn.get('cardId', '')
                email = card_map.get(cid, '')
                sheet = conn.get('sheet', '')
                mode = mode_map.get(cid, 'combined')
                w.writerow([cid, email, sheet, ts, mode])
                written_cards.add(cid)

            # Write cards with no connections (so we don't lose them)
            for card in cards:
                if card['id'] not in written_cards:
                    mode = card.get('mode', 'combined')
                    w.writerow([card['id'], card['email'], '', ts, mode])

        return jsonify({'success': True, 'message': 'Email config saved.'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@eod_bp.route('/auto-assign-branches', methods=['POST'])
def auto_assign_branches():
    """Auto-assign sheets to emails from email_sheet_config.xlsx final_combination.

    Reads the final_combination sheet which has columns:
      email_id, attached_sheet, format_type (separate/combine)
    Builds cards + conns for the email UI and saves to email_config.csv.
    """
    config_xlsx = config.DATA_DIR / 'email_sheet_config.xlsx'
    if not config_xlsx.exists():
        return jsonify({
            'success': False,
            'message': 'email_sheet_config.xlsx not found in data/',
        }), 404

    # Read final_combination sheet
    try:
        df = pd.read_excel(str(config_xlsx), sheet_name='final_combination')
        df.columns = [str(c).strip().lower() for c in df.columns]
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error reading config xlsx: {e}'}), 500

    # Identify columns
    email_col = next((c for c in df.columns if 'email' in c), None)
    sheet_col = next((c for c in df.columns if 'sheet' in c or 'attach' in c), None)
    fmt_col = next((c for c in df.columns if 'format' in c or 'type' in c), None)
    if not email_col or not sheet_col:
        return jsonify({
            'success': False,
            'message': f'Expected email_id and attached_sheet columns, found: {list(df.columns)}',
        }), 400

    # Validate sheets exist in report
    report_file = BACKEND_DATA_DIR / 'EOD_Report_Latest.xlsx'
    valid_sheets = set()
    if report_file.exists():
        try:
            from openpyxl import load_workbook as _lwb
            wb = _lwb(str(report_file), read_only=True)
            valid_sheets = set(wb.sheetnames)
            wb.close()
        except Exception:
            pass

    # Group rows by email → list of (sheet, format_type)
    email_groups = {}  # email → {'sheets': [...], 'mode': 'separate'|'combined'}
    skipped = []
    for _, row in df.iterrows():
        email = str(row.get(email_col, '')).strip()
        sheet = str(row.get(sheet_col, '')).strip()
        fmt = str(row.get(fmt_col, 'separate')).strip().lower() if fmt_col else 'separate'
        if not email or not sheet or '@' not in email or sheet == 'nan' or email == 'nan':
            continue
        if valid_sheets and sheet not in valid_sheets:
            skipped.append({'email': email, 'sheet': sheet, 'reason': 'sheet not in report'})
            continue
        if email not in email_groups:
            email_groups[email] = {'sheets': [], 'fmt': fmt}
        email_groups[email]['sheets'].append(sheet)
        # If any row for this email says 'combine', set mode to combined
        if fmt == 'combine':
            email_groups[email]['fmt'] = 'combine'

    # Build cards and conns
    cards = []
    conns = []
    allocated = []
    cid_counter = 0

    for email, info in sorted(email_groups.items()):
        cid_counter += 1
        card_id = f'c{cid_counter}'
        mode = 'combined' if info['fmt'] == 'combine' else 'separate'
        cards.append({'id': card_id, 'email': email, 'mode': mode})
        for sheet in info['sheets']:
            conn_id = f'x{card_id}_{hashlib.md5(sheet.encode()).hexdigest()[:8]}'
            conns.append({'id': conn_id, 'sheet': sheet, 'cardId': card_id})
            allocated.append({'email': email, 'sheet': sheet, 'mode': mode})

    # Save to email_config.csv
    try:
        BACKEND_DATA_DIR.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        with open(EMAIL_CONFIG_CSV, 'w', newline='') as f:
            w = csv.writer(f)
            w.writerow(['card_id', 'email', 'sheet', 'timestamp', 'mode'])
            for conn in conns:
                card = next(c for c in cards if c['id'] == conn['cardId'])
                w.writerow([card['id'], card['email'], conn['sheet'], ts, card['mode']])
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saving config: {e}'}), 500

    return jsonify({
        'success': True,
        'message': f'{len(allocated)} assignments loaded ({len(email_groups)} recipients)',
        'allocated': allocated,
        'unallocated': skipped,
        'cards': cards,
        'conns': conns,
        'cid': cid_counter,
    })


@eod_bp.route('/cache-history', methods=['GET'])
def get_cache_history():
    """Return cache history from CSV."""
    if not CACHE_HISTORY_CSV.exists():
        return jsonify({'rows': []})
    try:
        rows = []
        with open(CACHE_HISTORY_CSV, newline='') as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append(row)
        return jsonify({'rows': rows})
    except Exception as e:
        return jsonify({'rows': [], 'warning': str(e)})


@eod_bp.route('/api/health')
def health():
    """Health check endpoint."""
    return jsonify({'status': 'ok', 'message': 'EOD server is running'})


@eod_bp.route('/get-bucket-metadata', methods=['POST'])
def get_bucket_metadata():
    """
    Extract unique DPD bucket values from PAR file and Last Month PAR.
    Returns JSON with bucket mappings for both current and last month.
    OPTIMIZED: Uses Parquet cache when available for 90%+ faster extraction.
    """
    try:
        import pandas as pd

        result = {
            'success': True,
            'currentMonth': {},
            'lastMonth': {},
            'message': ''
        }

        # Process PAR file if provided
        if 'par' in request.files:
            par = request.files['par']
            par_tmp = save_upload_to_temp(par, prefix="dpd_par_")

            try:
                par_hash = compute_file_hash(par_tmp)
                par_cache = DB_CACHE_DIR / f"daily_par_cache_{par_hash}.parquet"

                if par_cache.exists():
                    logging.info(f"Using PAR cache for bucket extraction: {par_cache.name}")
                    try:
                        df_par = pd.read_parquet(par_cache, columns=['AccountID', 'DPD Days', 'LoanStatus', 'Days Group'])
                    except (ValueError, KeyError) as e:
                        logging.debug(f"Column-selective parquet read failed, falling back: {e}")
                        df_par = pd.read_parquet(par_cache)
                else:
                    existing_par_caches = list(DB_CACHE_DIR.glob("daily_par_cache_*.parquet")) if DB_CACHE_DIR.exists() else []
                    if existing_par_caches:
                        existing_par_caches.sort(key=lambda f: f.stat().st_mtime, reverse=True)
                        logging.info(f"Using existing PAR cache for bucket extraction: {existing_par_caches[0].name}")
                        try:
                            df_par = pd.read_parquet(existing_par_caches[0], columns=['AccountID', 'DPD Days', 'LoanStatus', 'Days Group'])
                        except (ValueError, KeyError) as e:
                            logging.debug(f"Column-selective parquet read failed, falling back: {e}")
                            df_par = pd.read_parquet(existing_par_caches[0])
                    else:
                        logging.info("Reading PAR from Excel for bucket extraction (first time)...")
                        try:
                            df_par = pd.read_excel(par_tmp, engine='calamine')
                        except (ImportError, ValueError) as e:
                            logging.debug(f"Calamine engine unavailable, falling back: {e}")
                            df_par = pd.read_excel(par_tmp)

                dpd_column = None
                possible_names = ['Days Group', 'Days group', 'DaysGroup', 'Daysgroup',
                                 'DPD Group', 'DPD Days', 'DPDDays']
                for col in df_par.columns:
                    if col in possible_names:
                        dpd_column = col
                        break

                if dpd_column:
                    bucket_mapping = processor.extract_dpd_buckets(df_par, dpd_column)
                    result['currentMonth'] = bucket_mapping
                    logging.info(f"Extracted current month buckets: {bucket_mapping}")
                else:
                    result['message'] += 'DPD column not found in PAR file. '

            except Exception as e:
                result['message'] += f'Error reading PAR: {str(e)}. '
                logging.error(f"Error reading PAR file: {e}")
            finally:
                par_tmp.unlink(missing_ok=True)

        # Process Last Month PAR
        last_month_cache = DB_CACHE_DIR / "last_month_par_cache.parquet" if DB_CACHE_DIR.exists() else None
        last_month_files = list(BACKEND_DATA_DIR.glob("Last_Month_PAR_*"))

        try:
            df_last = None
            try:
                if db_manager and db_manager.get_connection():
                    count = db_manager.get_connection().execute("SELECT count(*) FROM Last_Month_PAR").fetchone()[0]
                    if count > 0:
                        logging.info("Using Last Month PAR from DuckDB for bucket extraction")
                        df_last = db_manager.get_connection().execute("SELECT * FROM Last_Month_PAR").df()
            except (duckdb.CatalogException, duckdb.Error):
                pass

            if df_last is None and last_month_cache and last_month_cache.exists():
                logging.info("Using Last Month PAR cache for bucket extraction")
                try:
                    df_last = pd.read_parquet(last_month_cache, columns=['AccountID', 'DPD Days', 'LoanStatus'])
                except (ValueError, KeyError) as e:
                    logging.debug(f"Column-selective parquet read failed, falling back: {e}")
                    df_last = pd.read_parquet(last_month_cache)
            elif df_last is None and last_month_files:
                logging.info("Reading Last Month PAR from Excel...")
                try:
                    df_last = pd.read_excel(last_month_files[0], sheet_name=0, engine='calamine')
                except (ImportError, ValueError) as e:
                    logging.debug(f"Calamine engine unavailable, falling back: {e}")
                    df_last = pd.read_excel(last_month_files[0], sheet_name=0)
            elif df_last is None:
                result['message'] += 'No Last Month PAR file found. '

            if df_last is not None:
                dpd_column = None
                possible_names = ['Days Group', 'Days group', 'DaysGroup', 'Daysgroup',
                                 'DPD Group', 'DPD Days', 'DPDDays']
                for col in df_last.columns:
                    if col in possible_names:
                        dpd_column = col
                        break

                if dpd_column:
                    bucket_mapping = processor.extract_dpd_buckets(df_last, dpd_column)
                    result['lastMonth'] = bucket_mapping
                    logging.info(f"Extracted last month buckets: {bucket_mapping}")
                else:
                    result['message'] += 'DPD column not found in Last Month PAR. '

        except Exception as e:
            result['message'] += f'Error reading Last Month PAR: {str(e)}. '
            logging.error(f"Error reading Last Month PAR: {e}")

        if not result['currentMonth'] and not result['lastMonth']:
            result['success'] = False
            result['message'] = result['message'] or 'No bucket mappings could be extracted.'

        return jsonify(result)

    except Exception as e:
        err = user_error(e, context='eod-bucket-metadata')
        return jsonify({
            'success': False,
            'currentMonth': {},
            'lastMonth': {},
            'message': err['user_message'],
            'suggestion': err['suggestion'],
        }), 500


@eod_bp.route('/download-daily-report', methods=['GET'])
def download_daily_report():
    """Download the pre-computed Daily report."""
    report_file = BACKEND_DATA_DIR / 'Daily_Report_Latest.xlsx'
    if not report_file.exists():
        return jsonify({'error': 'No daily report available. Generate it first.'}), 404
    # Include date in download filename
    date_str = request.args.get('date', '')
    dl_name = f'Daily_Report_{date_str}.xlsx' if date_str else 'Daily_Report.xlsx'
    return send_file(
        report_file,
        as_attachment=True,
        download_name=dl_name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@eod_bp.route('/download-hourly-report', methods=['GET'])
def download_hourly_report():
    """Download the pre-computed Hourly report."""
    report_file = BACKEND_DATA_DIR / 'Hourly_Report_Latest.xlsx'
    if not report_file.exists():
        return jsonify({'error': 'No hourly report available. Generate it first.'}), 404
    # Include date in download filename
    date_str = request.args.get('date', '')
    dl_name = f'Hourly_Report_{date_str}.xlsx' if date_str else 'Hourly_Report.xlsx'
    return send_file(
        report_file,
        as_attachment=True,
        download_name=dl_name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@eod_bp.route('/generate-daily-hourly-report', methods=['POST'])
def generate_daily_hourly_report_endpoint():
    """Generate Daily + Hourly pre-computed reports from the latest EOD output.

    Two distinct Excel files:
      1. Daily_Report  - pre-computed report from the EOD output (daily collection)
      2. Hourly_Report - pre-computed report from EOD output merged with the
                         HourlyDaily collection file (hourly collection replaces
                         daily Collection values, recalculates Partial Amount, etc.)
    """
    try:
        import numpy as np
        from services.daily_report_builder import build_daily_report
        from services.column_matcher import find_column

        logging.info("DAILY+HOURLY REPORT: Starting on-demand generation")

        # Accept target_date from POST body
        target_date_str = None
        if request.is_json:
            target_date_str = request.json.get('targetDate')
        else:
            target_date_str = request.form.get('targetDate')

        target_date = None
        if target_date_str:
            try:
                target_date = pd.Timestamp(datetime.strptime(target_date_str, '%d-%m-%Y'))
                logging.info(f"DAILY+HOURLY REPORT: Using target_date: {target_date_str}")
            except ValueError:
                logging.warning(f"DAILY+HOURLY REPORT: Could not parse target_date '{target_date_str}'")

        # ── Load the base EOD output (used by both Daily and Hourly) ──
        df_eod = None

        # Try hourly EOD parquet cache first (it's the full EOD output cached by hourly module)
        if DB_CACHE_DIR.exists():
            for pattern in ["hourly_eod_cache_*.parquet", "daily_eod_cache_*.parquet"]:
                caches = sorted(DB_CACHE_DIR.glob(pattern),
                                key=lambda f: f.stat().st_mtime, reverse=True)
                if caches:
                    try:
                        df_eod = pd.read_parquet(caches[0])
                        logging.info(f"DAILY+HOURLY: Loaded EOD from {caches[0].name}")
                        break
                    except Exception as e:
                        logging.warning(f"DAILY+HOURLY: Parquet read failed ({pattern}): {e}")

        if df_eod is None:
            eod_excel = BACKEND_DATA_DIR / 'EOD_Output_Latest.xlsx'
            if eod_excel.exists():
                try:
                    df_eod = pd.read_excel(eod_excel, engine='calamine')
                except (ImportError, ValueError) as e:
                    logging.debug(f"Calamine engine unavailable, falling back: {e}")
                    df_eod = pd.read_excel(eod_excel)
                logging.info("DAILY+HOURLY: Loaded EOD from EOD_Output_Latest.xlsx")

        if df_eod is None or len(df_eod) == 0:
            return jsonify({'error': 'No EOD data available. Run EOD processing first.'}), 404

        # Derive target_date from data if not provided
        if target_date is None and 'Meeting Date' in df_eod.columns:
            try:
                from services.eod_processor import parse_date_column
                meeting_dates = parse_date_column(df_eod['Meeting Date'])
                max_date = meeting_dates.dropna().max()
                if pd.notna(max_date):
                    target_date = max_date
            except (ValueError, TypeError) as e:
                logging.debug(f"Could not derive target_date from Meeting Date: {e}")
        if target_date is None:
            target_date = pd.Timestamp.now()

        has_officer = 'Emp ID' in df_eod.columns
        daily_path = None
        hourly_path = None

        # ═══════════════════════════════════════════════════════════════
        # DAILY REPORT - use EOD output as-is (daily collection already in it)
        # ═══════════════════════════════════════════════════════════════
        try:
            precomp_daily = processor._compute_precomputed_sheets(df_eod, target_date)
            if precomp_daily and '_precomp' in precomp_daily:
                daily_output = BACKEND_DATA_DIR / 'Daily_Report_Latest.xlsx'
                build_daily_report(precomp_daily['_precomp'], daily_output, target_date, has_officer)
                daily_path = daily_output
                logging.info(f"DAILY+HOURLY: Daily report OK -> {daily_output.name}")
        except Exception as e:
            logging.warning(f"DAILY+HOURLY: Daily report failed: {e}")

        # ═══════════════════════════════════════════════════════════════
        # HOURLY REPORT - merge hourly collection onto EOD, then precomp
        # ═══════════════════════════════════════════════════════════════
        try:
            # Step 1: Find the hourly collection data
            df_hourly_coll = None

            # Try hourly collection parquet cache first
            if DB_CACHE_DIR.exists():
                h_caches = sorted(DB_CACHE_DIR.glob("hourly_collection_cache_*.parquet"),
                                  key=lambda f: f.stat().st_mtime, reverse=True)
                if h_caches:
                    try:
                        df_hourly_coll = pd.read_parquet(h_caches[0])
                        logging.info(f"DAILY+HOURLY: Hourly collection from cache: {h_caches[0].name}")
                    except Exception as e:
                        logging.warning(f"DAILY+HOURLY: Hourly collection cache read failed: {e}")

            # Fallback: read HourlyDaily_* file from backend dir
            if df_hourly_coll is None:
                from services import file_manager
                hd_file = file_manager.find_file_by_pattern(BACKEND_DATA_DIR, 'HourlyDaily_*')
                if hd_file and not hd_file.name.startswith('~$'):
                    try:
                        df_hourly_coll = pd.read_excel(hd_file, engine='calamine')
                    except (ImportError, ValueError) as e:
                        logging.debug(f"Calamine engine unavailable, falling back: {e}")
                        df_hourly_coll = pd.read_excel(hd_file)
                    logging.info(f"DAILY+HOURLY: Hourly collection from {hd_file.name}")

            if df_hourly_coll is None or len(df_hourly_coll) == 0:
                logging.warning("DAILY+HOURLY: No hourly collection data found, skipping hourly report")
            else:
                # Step 2: Filter ReverseTotal == 0 (same as hourly.py /process)
                col_reverse = find_column(df_hourly_coll, 'ReverseTotal', 'Reverse Total')
                if col_reverse:
                    df_hourly_coll = df_hourly_coll[df_hourly_coll[col_reverse] == 0]

                # Step 3: Pivot by AccountID -> sum CollectionTotal
                col_acct = find_column(df_hourly_coll, 'AccountID', 'Account ID')
                col_coll_total = find_column(df_hourly_coll, 'CollectionTotal', 'Collection Total')

                if col_acct and col_coll_total:
                    pivot = df_hourly_coll.groupby(col_acct)[col_coll_total].sum()
                    hourly_lookup = pivot.to_dict()
                    logging.info(f"DAILY+HOURLY: Hourly pivot: {len(hourly_lookup)} unique accounts")

                    # Step 4: Clone EOD data and replace Collection with hourly values
                    df_hourly_merged = df_eod.copy()
                    col_eod_acct = find_column(df_hourly_merged, 'Account ID', 'AccountID')

                    if col_eod_acct:
                        # Replace Collection column with hourly collection values
                        df_hourly_merged['Collection'] = df_hourly_merged[col_eod_acct].map(hourly_lookup)

                        # Recalculate Partial Amount (same logic as eod_processor)
                        reg_demand = pd.to_numeric(df_hourly_merged.get('Regular Demand', 0), errors='coerce').fillna(0)
                        collection = pd.to_numeric(df_hourly_merged.get('Collection', 0), errors='coerce')
                        difference = reg_demand - collection.fillna(0)

                        df_hourly_merged['Partial Amount'] = 'Not Collected'
                        has_col = collection.notna()
                        df_hourly_merged.loc[has_col & (difference <= 0), 'Partial Amount'] = 'Full EMI Paid'
                        df_hourly_merged.loc[has_col & (difference > 0), 'Partial Amount'] = 'Partial Amount'

                        # Recalculate installment - collected columns
                        inst_amt = pd.to_numeric(df_hourly_merged.get('Installment Amount', 0), errors='coerce').fillna(0)
                        df_hourly_merged['installment - collected amt'] = inst_amt - collection.fillna(0)
                        df_hourly_merged['installment - collected value'] = (
                            df_hourly_merged['installment - collected amt'] <= 0
                        ).astype(int)

                        df_hourly_merged['Remark2'] = 'Not Collected'
                        df_hourly_merged.loc[has_col & (difference <= 0), 'Remark2'] = 'Full Collected'
                        df_hourly_merged.loc[has_col & (difference > 0), 'Remark2'] = 'Partially Collected'

                        matched = has_col.sum()
                        logging.info(f"DAILY+HOURLY: Hourly merge matched {matched}/{len(df_hourly_merged)} rows")

                        # Step 5: Precompute and build report
                        precomp_hourly = processor._compute_precomputed_sheets(df_hourly_merged, target_date)
                        if precomp_hourly and '_precomp' in precomp_hourly:
                            hourly_output = BACKEND_DATA_DIR / 'Hourly_Report_Latest.xlsx'
                            build_daily_report(precomp_hourly['_precomp'], hourly_output, target_date, has_officer, hourly_mode=True)
                            hourly_path = hourly_output
                            logging.info(f"DAILY+HOURLY: Hourly report OK -> {hourly_output.name}")
                    else:
                        logging.warning("DAILY+HOURLY: Account ID column not found in EOD output")
                else:
                    logging.warning(f"DAILY+HOURLY: Required columns not found in hourly collection "
                                    f"(AccountID={col_acct}, CollectionTotal={col_coll_total})")

        except Exception as e:
            logging.warning(f"DAILY+HOURLY: Hourly report failed: {e}")

        if not daily_path and not hourly_path:
            return jsonify({
                'error': 'No EOD data available or hourly collection file missing. '
                         'Run EOD processing first and upload a HourlyDaily collection file.',
            }), 404

        result = {
            'success': True,
            'daily': daily_path is not None,
            'hourly': hourly_path is not None,
            'reportDate': target_date.strftime('%d-%m-%Y'),
            'message': 'Daily + Hourly reports generated successfully.'
        }
        logging.info(f"DAILY+HOURLY: Done. Daily={daily_path is not None}, Hourly={hourly_path is not None}")
        return jsonify(result)

    except Exception as e:
        err = user_error(e, context='eod-generate-daily-hourly-report')
        return jsonify({
            'error': err['user_message'],
            'suggestion': err['suggestion'],
        }), 500


# ============================================================================
# VBA Runner
# ============================================================================

@eod_bp.route('/vba-runner')
def vba_runner():
    """Serve the VBA runner page."""
    return send_from_directory(str(STATIC_EOD_DIR), 'vba_runner.html')


@eod_bp.route('/whatsapp-sender')
def whatsapp_sender():
    """Serve the WhatsApp sender page."""
    return send_from_directory(str(STATIC_EOD_DIR), 'whatsapp_sender.html')


@eod_bp.route('/whatsapp-contacts', methods=['GET'])
def whatsapp_contacts_get():
    """Return the list of WhatsApp contacts."""
    csv_path = config.DATA_DIR / 'whatsapp_contacts.csv'
    names = []
    if csv_path.exists():
        import csv
        with open(csv_path, 'r', encoding='utf-8') as f:
            for row in csv.DictReader(f):
                n = row.get('name', '').strip()
                if n:
                    names.append(n)
    return jsonify({'contacts': names})


@eod_bp.route('/whatsapp-contacts', methods=['POST'])
def whatsapp_contacts_save():
    """Save the list of WhatsApp contacts."""
    data = request.get_json(silent=True) or {}
    names = data.get('contacts', [])
    csv_path = config.DATA_DIR / 'whatsapp_contacts.csv'
    import csv
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['name'])
        for n in names:
            n = n.strip()
            if n:
                writer.writerow([n])
    return jsonify({'success': True, 'contacts': names})


@eod_bp.route('/whatsapp-open', methods=['POST'])
def whatsapp_open():
    """Open WhatsApp Web in Chromium (browser stays open)."""
    from services.whatsapp_sender import open_whatsapp
    result = open_whatsapp()
    return jsonify(result), 200 if result['success'] else 500


@eod_bp.route('/whatsapp-send', methods=['POST'])
def whatsapp_send():
    """Send a file from a bundle via WhatsApp."""
    from services.whatsapp_sender import send_file_to_contact

    data = request.get_json(silent=True) or {}
    bundle_path = data.get('bundle_path', '')
    filename = data.get('filename', '')

    if not bundle_path or not filename:
        return jsonify({'success': False, 'error': 'Missing bundle_path or filename'}), 400

    file_path = Path(bundle_path) / filename
    if not file_path.exists():
        return jsonify({'success': False, 'error': f'File not found: {file_path}'}), 404

    result = send_file_to_contact(str(file_path))
    status_code = 200 if result['success'] else 500
    return jsonify(result), status_code


@eod_bp.route('/vba-runner/bundles')
def vba_runner_bundles():
    """List all EOD_Bundle folders sorted newest-first."""
    bundle_root = Path.home() / 'Downloads' / 'EOD_Bundle'
    if not bundle_root.exists():
        return jsonify({'bundles': []})

    bundles = []
    for d in sorted(bundle_root.iterdir(), reverse=True):
        if not d.is_dir() or d.name.startswith('.'):
            continue
        files = [f.name for f in d.iterdir() if f.is_file() and not f.name.startswith('~$') and not f.name.startswith('.')]
        # Read target_date if saved
        td_file = d / '.target_date'
        td = td_file.read_text().strip() if td_file.exists() else None
        bundles.append({
            'name': d.name,
            'path': str(d),
            'files': sorted(files),
            'target_date': td,
        })

    return jsonify({'bundles': bundles})


@eod_bp.route('/vba-runner/run', methods=['POST'])
def vba_runner_run():
    """Run actual VBA macro on the Excel file in a bundle.

    POST JSON: {bundle_path: str, script: "daily"|"month_end"}

    Windows: Uses COM automation (win32com) to open Excel, inject VBA, and run macro.
    Mac:     Uses AppleScript to automate Excel UI.
    """
    import platform as _plat

    data = request.get_json(force=True)
    bundle_path = Path(data.get('bundle_path', ''))
    script_type = data.get('script', 'daily')

    if not bundle_path.exists() or not bundle_path.is_dir():
        return jsonify({'error': 'Bundle folder not found'}), 404

    raw_xlsx = bundle_path / 'Regular Demand Vs Collection.xlsx'
    if not raw_xlsx.exists():
        return jsonify({'error': 'Regular Demand Vs Collection.xlsx not found in bundle'}), 404

    # Pick the right VBA file
    if script_type == 'month_end':
        vba_file = bundle_path / 'VBA_Template_Month_End.txt'
    else:
        vba_file = bundle_path / 'VBA_Template.txt'

    macro_name = 'CreateThreePivots'

    if not vba_file.exists():
        return jsonify({'error': f'{vba_file.name} not found in bundle'}), 404

    if _plat.system() == 'Windows':
        # Kill any existing/zombie Excel processes so COM gets a fresh instance
        import subprocess as _sp_check
        _sp_check.run(
            ['taskkill', '/F', '/IM', 'EXCEL.EXE'],
            capture_output=True, timeout=10,
        )
        import time as _tw
        _tw.sleep(1)
        return _vba_runner_windows(raw_xlsx, vba_file, macro_name)
    else:
        return _vba_runner_mac(raw_xlsx, vba_file, macro_name)


def _vba_runner_windows(raw_xlsx, vba_file, macro_name):
    """Run VBA macro via VBScript (cscript.exe).

    Bypasses pywin32 entirely — VBScript uses native Windows COM
    with no gen_py cache, no binding issues, no thread problems.
    """
    import subprocess as _sp
    import tempfile

    xlsx_path = str(raw_xlsx.resolve())
    vba_path = str(vba_file.resolve())

    # Write VBScript to a temp file — uses native Windows COM, no pywin32
    vbs_lines = [
        'On Error Resume Next',
        'Dim xlApp, wb, vbMod, fso, ts, vbaCode, modName',
        'Set fso = CreateObject("Scripting.FileSystemObject")',
        'Set xlApp = CreateObject("Excel.Application")',
        'If Err.Number <> 0 Then',
        '    WScript.StdErr.WriteLine "ERROR: Could not start Excel"',
        '    WScript.Quit 1',
        'End If',
        'On Error GoTo 0',
        '',
        'xlApp.DisplayAlerts = False',
        'xlApp.Visible = True',
        'xlApp.AutomationSecurity = 1',  # msoAutomationSecurityLow — allow macros
        '',
        # Open workbook, save as .xlsm FIRST (xlsx can't run macros)
        'Dim origPath, xlsmPath',
        f'origPath = "{xlsx_path}"',
        'xlsmPath = Replace(origPath, ".xlsx", ".xlsm")',
        'If fso.FileExists(xlsmPath) Then fso.DeleteFile xlsmPath',
        '',
        'Set wb = xlApp.Workbooks.Open(origPath)',
        'If wb Is Nothing Then',
        '    WScript.StdErr.WriteLine "ERROR: Could not open workbook"',
        '    xlApp.Quit',
        '    WScript.Quit 1',
        'End If',
        'wb.SaveAs xlsmPath, 52',
        '',
        # Read VBA code (ASCII encoding)
        f'Set ts = fso.OpenTextFile("{vba_path}", 1, False, 0)',
        'vbaCode = ts.ReadAll',
        'ts.Close',
        '',
        # Inject VBA into the .xlsm workbook
        'On Error Resume Next',
        'Set vbMod = wb.VBProject.VBComponents.Add(1)',
        'If Err.Number <> 0 Then',
        '    WScript.StdErr.WriteLine "ERROR: VBA access blocked"',
        '    xlApp.DisplayAlerts = True',
        '    xlApp.Quit',
        '    WScript.Quit 1',
        'End If',
        'Err.Clear',
        'vbMod.CodeModule.AddFromString vbaCode',
        'If Err.Number <> 0 Then',
        '    WScript.StdErr.WriteLine "ERROR: Failed to inject VBA: " & Err.Description',
        '    xlApp.DisplayAlerts = True',
        '    xlApp.Quit',
        '    WScript.Quit 1',
        'End If',
        'On Error GoTo 0',
        '',
        # Run the macro
        f'xlApp.Run "{macro_name}"',
        '',
        # Save back as .xlsx and clean up
        'wb.VBProject.VBComponents.Remove vbMod',
        'wb.SaveAs origPath, 51',
        'If fso.FileExists(xlsmPath) Then fso.DeleteFile xlsmPath',
        '',
        'xlApp.DisplayAlerts = True',
        '',
        'WScript.StdOut.WriteLine "OK"',
    ]
    vbs_content = '\r\n'.join(vbs_lines) + '\r\n'

    tmp_vbs = None
    try:
        t0 = time.perf_counter()
        logging.info(f"VBA-RUNNER [Windows/VBS]: Launching cscript for {raw_xlsx.name}")

        tmp_vbs = tempfile.NamedTemporaryFile(
            mode='w', suffix='.vbs', delete=False, encoding='utf-8'
        )
        tmp_vbs.write(vbs_content)
        tmp_vbs.close()

        result = _sp.run(
            ['cscript.exe', '//Nologo', tmp_vbs.name],
            capture_output=True, text=True, timeout=900,
        )

        logging.info(f"VBA-RUNNER [Windows/VBS] stdout: {result.stdout.strip()}")
        if result.stderr.strip():
            logging.warning(f"VBA-RUNNER [Windows/VBS] stderr: {result.stderr.strip()}")

        elapsed = time.perf_counter() - t0
        output = result.stdout.strip()
        err_output = result.stderr.strip()

        if result.returncode == 0 and output == 'OK':
            logging.info(f"VBA-RUNNER [Windows/VBS]: Completed in {elapsed:.1f}s")
            return jsonify({
                'success': True,
                'output': 'Macro executed on ' + raw_xlsx.name,
                'elapsed': round(elapsed, 1),
                'message': f'{macro_name} completed successfully',
            })
        else:
            error_msg = output or err_output or 'Unknown error from cscript'
            if 'VBA access blocked' in error_msg:
                return jsonify({
                    'error': 'Excel blocked VBA access. Fix: Excel > File > Options > '
                             'Trust Center > Trust Center Settings > Macro Settings > '
                             'check "Trust access to the VBA project object model", then retry.'
                }), 403
            return jsonify({'error': error_msg}), 500

    except _sp.TimeoutExpired:
        logging.error("VBA-RUNNER [Windows/VBS]: Timed out after 15 minutes")
        return jsonify({
            'error': 'Excel automation timed out (15 min). The macro may still be running in Excel.'
        }), 504
    except Exception as e:
        logging.exception(f"VBA-RUNNER [Windows/VBS]: Failed: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if tmp_vbs:
            import os
            try:
                os.unlink(tmp_vbs.name)
            except OSError:
                pass


def _vba_runner_mac(raw_xlsx, vba_file, macro_name):
    """Run VBA macro using AppleScript (macOS only)."""
    import subprocess as _sp

    try:
        t0 = time.perf_counter()
        logging.info(f"VBA-RUNNER [Mac]: Opening Excel and running VBA on {raw_xlsx.name}")

        xlsx_path = str(raw_xlsx)

        def _osa(script, label, timeout=30):
            r = _sp.run(['osascript', '-e', script],
                        capture_output=True, text=True, timeout=timeout)
            if r.returncode != 0:
                err = r.stderr.strip()
                logging.error(f"VBA-RUNNER [{label}]: {err}")
                raise RuntimeError(f'{label}: {err}')
            logging.info(f"VBA-RUNNER [{label}]: OK - {r.stdout.strip()}")
            return r.stdout.strip()

        _osa('''
tell application "Microsoft Excel"
    activate
    try
        close every workbook without saving
    end try
end tell
return "closed"
''', 'close-all')

        import time as _time
        _time.sleep(1)

        _osa(f'''
tell application "Microsoft Excel"
    open "{xlsx_path}"
end tell
delay 5
tell application "Microsoft Excel"
    return name of active workbook
end tell
''', 'open-file', timeout=60)

        vba_code = vba_file.read_bytes()
        _sp.run(['pbcopy'], input=vba_code, timeout=10)
        logging.info("VBA-RUNNER [clipboard]: VBA code copied")

        _osa('''
tell application "System Events"
    tell process "Microsoft Excel"
        click menu item "Record New Macro..." of menu 1 of menu item "Macro" of menu "Tools" of menu bar 1
    end tell
end tell
delay 1.5
tell application "System Events"
    keystroke return
end tell
delay 1
tell application "System Events"
    tell process "Microsoft Excel"
        click menu item "Stop Recording" of menu 1 of menu item "Macro" of menu "Tools" of menu bar 1
    end tell
end tell
return "recorded"
''', 'record-macro')

        _time.sleep(1)

        _osa('''
tell application "System Events"
    tell process "Microsoft Excel"
        click menu item "Macros..." of menu 1 of menu item "Macro" of menu "Tools" of menu bar 1
    end tell
end tell
delay 2
tell application "System Events"
    tell process "Microsoft Excel"
        click button "Edit" of front window
    end tell
end tell
delay 2
return "vbe-open"
''', 'open-vbe-edit')

        _sp.run(['pbcopy'], input=vba_code, timeout=10)

        _osa('''
tell application "System Events"
    tell process "Microsoft Excel"
        click menu item "Select All" of menu "Edit" of menu bar 1
        delay 0.5
        click menu item "Paste" of menu "Edit" of menu bar 1
        delay 5
    end tell
end tell
return "pasted"
''', 'paste-vba', timeout=30)

        _osa(f'''
tell application "Microsoft Excel"
    activate
end tell
delay 1
tell application "Microsoft Excel"
    run VB macro "{macro_name}"
end tell
tell application "Microsoft Excel"
    save active workbook
end tell
return "done"
''', 'run-macro', timeout=900)

        elapsed = time.perf_counter() - t0
        logging.info(f"VBA-RUNNER [Mac]: Macro completed in {elapsed:.1f}s")

        return jsonify({
            'success': True,
            'output': 'Macro executed on ' + raw_xlsx.name,
            'elapsed': round(elapsed, 1),
            'message': f'{macro_name} completed successfully',
        })

    except _sp.TimeoutExpired:
        logging.error("VBA-RUNNER [Mac]: Timed out after 15 minutes")
        return jsonify({'error': 'Excel automation timed out (15 min). The macro may still be running in Excel.'}), 504
    except Exception as e:
        logging.exception(f"VBA-RUNNER [Mac]: Failed: {e}")
        return jsonify({'error': str(e)}), 500
