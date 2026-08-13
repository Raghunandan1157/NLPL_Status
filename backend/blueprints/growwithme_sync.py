"""
GrowWithMe Sync Blueprint (Phase 2 — local MySQL API)
=====================================================
Pushes EOD daily, Quick hourly, Disbursement and Portfolio data into the
``growwithme-local`` Node/Express API (now deployed on AWS EC2, MariaDB
``Growwithme_NEWDB``; base URL in GROWWITHME_API_URL).

The growwithme-local ``/sync`` endpoints expect rows **already exploded** into
DPD buckets + NPA actions:

  POST {GROWWITHME_API_URL}/api/collection/sync         (EOD daily — grain 2)
  POST {GROWWITHME_API_URL}/api/hourly/sync             (Quick hourly — grain 1)
  POST {GROWWITHME_API_URL}/api/disbursement/sync       (Disbursement — monthly)
  POST {GROWWITHME_API_URL}/api/disbursement/sync-daily (Disbursement — per-day)
  POST {GROWWITHME_API_URL}/api/portfolio/sync          (Portfolio POS — monthly)
  POST {GROWWITHME_API_URL}/api/portfolio/sync-accounts (Portfolio Total Account)
  POST {GROWWITHME_API_URL}/api/clients/sync            (Customer detail — per account)
  POST {GROWWITHME_API_URL}/api/overview/compute        (Overview board figures, from the snapshots)

This blueprint reuses the shared report parsers (``blueprints.report_parsers``)
and transforms each flat row into the bucketed shape using the GrowwithmeDB ids:

  dpd_bucket : 1=regular 2=1_30 3=31_60 4=pnpa 5=on_date (6=61_90 derived, skipped)
  npa_action : 1=activation 2=closure
  product_type: 1=IGL 2=FIG 3=IL

Config (env, no engine-config change needed):
  GROWWITHME_API_URL    base URL of the Node API (default http://localhost:4000)
  GROWWITHME_API_TOKEN  optional `Authorization: Token <t>`; the /sync push
                        endpoints are intentionally open, so this is usually blank.

NOTE on semantics (differs from the Supabase path):
  The growwithme-local /sync endpoints now do a whole-scope OVERRIDE on the API
  side (collection per-date, hourly full-snapshot, disbursement per-month,
  disbursement/sync-daily per-date, portfolio per-month), so re-running a sync
  REPLACES that scope rather than appending. Disbursement is pushed at BOTH
  grains: monthly (db_month = first-of-month) for the Disbursement tab and daily
  (disb_date) for its Daily tab — from the same per-day aggregate. Portfolio
  pushes POS amounts and, when the POS sheet carries an account column, Total
  Account counts (pos_status 'total_acc').
"""

import io
import os
import json
import logging
import threading
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime

from flask import Blueprint, jsonify, request
import requests as http_requests

import config

# Reuse the exact parsers the Supabase sync uses, so AWS / Supabase / GrowwithmeDB
# all ingest identical numbers from the same report files.
from blueprints.report_parsers import (
    _employee_report_path,
    _parse_report,
    _quick_report_path,
    _parse_quick_report,
    _norm,
    _num,
    _safe_num,
    _DISB_PRODUCT_TYPE_ID,
)
from openpyxl import load_workbook
import re
import zipfile
from xml.etree import ElementTree as ET

logger = logging.getLogger(__name__)
growwithme_bp = Blueprint('growwithme_sync', __name__)

GROWWITHME_API_URL = (os.environ.get('GROWWITHME_API_URL') or 'http://localhost:4000').rstrip('/')
GROWWITHME_API_TOKEN = os.environ.get('GROWWITHME_API_TOKEN', '')

# Flat metric column -> (dpd_bucket_id). Each bucket pulls four metric columns:
# (demand_count, demand_amt, collection_count, collection_amt).
_DPD_MAP = [
    (1, 'regular_demand',  'regular_demand_amt',  'regular_collection',  'regular_collection_amt'),
    (2, 'demand_1_30',     'demand_1_30_amt',     'collection_1_30',     'collection_1_30_amt'),
    (3, 'demand_31_60',    'demand_31_60_amt',    'collection_31_60',    'collection_31_60_amt'),
    (4, 'pnpa_demand',     'pnpa_demand_amt',     'pnpa_collection',     'pnpa_collection_amt'),
    (5, 'on_date_demand',  'on_date_demand_amt',  'on_date_collection',  'on_date_collection_amt'),
]
# npa_action_id -> (accounts column, amount column).
_NPA_MAP = [
    (1, 'npa_act_acc', 'npa_act_amt'),  # activation
    (2, 'npa_clo_acc', 'npa_clo_amt'),  # closure
]

# ── Rupee amounts: sourced from the EOD workbook's `_precomp` sheet ────────
# The Daily Collection Report carries NO rupee columns — its DPD blocks are
# DEMAND / COLLECTION / BALANCE / COLLECTION %, all ACCOUNT COUNTS (verified
# against the raw 03-08-2026 workbook). That is why every `*_amt` used to be
# pushed as 0 and the MIS Collection screen's Amount view had nothing to show.
#
# The amounts DO exist, one file over: `EOD_Output_Latest.xlsx` ("Regular Demand
# vs Collection"), whose hidden `_precomp` sheet is what eod_processor already
# aggregates for the EOD report. Crucially its amount columns are computed from
# the SAME masks as the count columns it sits beside (eod_processor: reg_demand_amt
# uses ftod_mask, dem_130_amt uses mask_130_base, …), so the rupees MATCH the
# accounts row for row — no reconciliation drift between the two views.
#
# Grain: filter_type 'All_EmpID' -> group_value is the NL##### employee_code, the
# very key /api/collection/sync resolves. scope 'OA' = overall (the daily grain);
# 'FY' rows are financial-year-to-date and are ignored here.
#
# metric field pushed by _explode -> `_precomp` column name.
_PRECOMP_AMT_MAP = {
    'regular_demand_amt':     'reg_demand_amt',
    'regular_collection_amt': 'reg_collection_amt',
    'demand_1_30_amt':        'dem_130_amt',
    'collection_1_30_amt':    'col_130_amt',
    'demand_31_60_amt':       'dem_3160_amt',
    'collection_31_60_amt':   'col_3160_amt',
    'pnpa_demand_amt':        'pnpa_demand_amt',
    'pnpa_collection_amt':    'pnpa_collection_amt',
    'on_date_demand_amt':     'od_demand_amt',
    'on_date_collection_amt': 'od_collection_amt',
}
# `_precomp` product label -> product_type_id, matching report_parsers._PRODUCT_TYPE_ID
# (VVY is the source label for IL). 'ALL' is the product-combined row, keyed None.
_PRECOMP_PRODUCT_ID = {'ALL': None, 'IGL': 1, 'FIG': 2, 'IL': 3, 'VVY': 3}

# Reuse a connection-pooled session — daily syncs POST one row per employee.
# Per THREAD, because the append chunks below are pushed concurrently and
# requests.Session is not documented as thread-safe.
_local = threading.local()
_session = http_requests.Session()   # kept for /ping and other single-shot calls


def _thread_session():
    s = getattr(_local, 'session', None)
    if s is None:
        s = http_requests.Session()
        _local.session = s
    return s


def _headers():
    h = {'Content-Type': 'application/json'}
    if GROWWITHME_API_TOKEN:
        h['Authorization'] = f'Token {GROWWITHME_API_TOKEN}'
    return h


# Seconds to wait for a sync chunk. 60 WAS TOO SHORT (2026-08-12: chunks 2, 3, 4,
# 10, 11, 13 and 15 of 15 timed out, leaving the date partially written).
#
# A chunk is not slow because of its size — it is slow because the API inserts
# ROW BY ROW: one INSERT for the period, one per dpd bucket, one per npa action.
# 100 rows is therefore ~1000 sequential round-trips to MySQL, and at even 50 ms
# each that is a minute of work for one chunk.
#
# The real fix is the batched INSERT in the API (routes/collection.js and
# routes/hourly.js, deploy pending) which collapses those ~1000 round-trips into
# 3. Until that is on the server, waiting longer is what gets the day written.
_POST_TIMEOUT = int(os.environ.get('GROWWITHME_POST_TIMEOUT') or 300)


def _post(path, payload):
    """POST to the growwithme-local API. Returns (ok, result_or_errmsg)."""
    url = f'{GROWWITHME_API_URL}{path}'
    try:
        resp = _thread_session().post(url, json=payload, headers=_headers(), timeout=_POST_TIMEOUT)
    except http_requests.exceptions.RequestException as e:
        # A body the server refuses mid-upload looks IDENTICAL to a dead server
        # here: the socket is closed before any response, so requests raises
        # SSLEOFError / "Max retries exceeded" either way. Calling that "not
        # reachable" sent people to check whether EC2 was down when the API was
        # up and answering. Name the likely cause when the body was big.
        size = len(json.dumps(payload)) if payload is not None else 0
        if isinstance(e, http_requests.exceptions.ReadTimeout):
            return False, (
                f'growwithme-api did not answer within {_POST_TIMEOUT}s. The server is UP '
                f'and reading the request — it is just slow, because it inserts one row '
                f'at a time. Deploy api/routes/collection.js and api/routes/hourly.js '
                f'(batched INSERTs) and restart the API, or raise '
                f'GROWWITHME_POST_TIMEOUT / lower GROWWITHME_CHUNK_ROWS. '
                f'NOTE: a timed-out chunk may still have committed on the server.')
        if size > 100_000 and 'EOF' in str(e):
            return False, (
                f'growwithme-api closed the connection on a {size // 1024} KB request. '
                f'The server is most likely UP — it just refuses bodies this large. '
                f'Deploy api/index.js (it sets express.json limit 50mb; without it '
                f'express caps bodies at 100 KB) and restart the API, or lower '
                f'GROWWITHME_CHUNK_ROWS. Underlying error: {e}')
        return False, f'growwithme-api not reachable: {e}'
    if resp.status_code not in (200, 201, 204):
        return False, f'{path} failed ({resp.status_code}): {(resp.text or "")[:300]}'
    try:
        return True, resp.json()
    except ValueError:
        return True, None


def _explode(rec, period_date=None, period_hour=None):
    """Turn one flat 25-metric row into one growwithme /sync batch row.

    The whole-scope override lives on the API side, so the batch carries the
    exploded employee rows; date/hour identify the scope being replaced.
    """
    dpd = [
        {
            'bucket_id': bk,
            'demand_count': rec.get(dc, 0),
            'demand_amt': rec.get(da, 0),
            'collection_count': rec.get(cc, 0),
            'collection_amt': rec.get(ca, 0),
        }
        for bk, dc, da, cc, ca in _DPD_MAP
    ]
    npa = [
        {'action_id': aid, 'accounts': rec.get(acc, 0), 'amount': rec.get(amt, 0)}
        for aid, acc, amt in _NPA_MAP
    ]
    # Quick hourly rows carry the combined product id 0 (no product split); send
    # NULL rather than 0 since GrowwithmeDB product_type has no id 0.
    pt = rec.get('product_type_id')
    row = {
        'emp_id': rec['emp_id'],
        'product_type_id': pt if pt else None,
        'npa_cases': rec.get('npa_cases', 0),
        'dpd': dpd,
        'npa': npa,
    }
    if period_date is not None:
        row['period_date'] = period_date
    if period_hour is not None:
        row['period_hour'] = period_hour
    return row


# Rows per HTTP request. A full day's collection push is thousands of rows, each
# carrying ~25 metrics plus dpd/npa arrays, which serialises well past the body
# limits in front of the route. Splitting keeps every request small regardless of
# how the server is configured, so this works without touching the server.
#
# 300 WAS TOO BIG — that is the 2026-08-12 "SSLEOFError / not reachable" failure.
# There are TWO limits in the path, and they fail differently:
#
#   nginx  client_max_body_size ~1 MB  -> a clean "413 Request Entity Too Large"
#   express  express.json()   100 KB   -> the socket is closed WHILE the client is
#                                         still uploading, so requests never sees
#                                         a response. It raises SSLEOFError, which
#                                         _post() reports as "not reachable" —
#                                         even though the server is perfectly up
#                                         and answers a small request in 0.3 s.
#
# 300 rows serialises to just over 100 KB, so it tripped the express limit and
# every chunk died at "chunk 1/5 failed ... not reachable (0 row(s) written)".
# Measured against the live API on 2026-08-12: 99 KB is answered, 102 KB is cut off.
#
# 100 rows is ~35 KB — comfortably under the limit even on a server still running
# the old index.js. Once index.js (express.json limit 50mb) is deployed, this can
# go back up; it is read from the environment so it can be tuned without an edit.
_CHUNK_ROWS = int(os.environ.get('GROWWITHME_CHUNK_ROWS') or 100)

# Concurrent append requests. The first chunk (the whole-scope replace) is always
# sent alone before these start, so raising this only widens the append fan-out.
#
# DEFAULT 1 (SEQUENTIAL) ON PURPOSE — do not raise it without reading this.
# /api/collection/sync allocates its primary key as
#     SELECT COALESCE(MAX(period_id),0)+1 ... then bumps a local counter
# which is NOT atomic. Two concurrent appends read the SAME max and insert
# overlapping period_id values, so all but one die on a duplicate key and the
# client sees `500 {"error":"server error"}` on most chunks — with the date left
# PARTIALLY written. Observed 2026-08-07: chunks 3, 4 and 5 of 5 failed exactly
# this way, 600 rows landed.
#
# Sequential costs a few round-trips on a push that is already several seconds;
# a partially written day costs a re-run and a phone call. Raise this ONLY once
# the API is on a build whose period_id allocation is safe under concurrency
# (see routes/collection.js — it now takes a named lock around the id block).
_PUSH_WORKERS = int(os.environ.get('GROWWITHME_PUSH_WORKERS') or 1)


def _push_batch(path, payload):
    """POST `payload` to `path`, splitting rows across several requests when needed.

    The sync endpoints are whole-scope overrides (they delete the date/month, then
    insert). Only the FIRST chunk may do that — every later chunk sends
    replace:false and appends. Sending them all with the default replace:true
    would leave only the final chunk in the database.

    Returns the same (response_dict, http_status) shape as a single push.
    """
    rows = payload.get('rows')
    if not isinstance(rows, list) or len(rows) <= _CHUNK_ROWS:
        ok, res = _post(path, payload)
        if not ok:
            return {'success': False, 'message': res, 'inserted': 0}, 502
        res = res or {}
        inserted = int(res.get('inserted') or 0)
        skipped = int(res.get('skipped') or 0)
        msg = f'{inserted} rows synced to growwithme-local'
        if skipped:
            msg += f' · {skipped} skipped (employee not in GrowwithmeDB)'
        return {'success': True, 'inserted': inserted, 'skipped': skipped, 'message': msg}, 200

    total_inserted = total_skipped = total_unresolved = 0
    chunks = [rows[i:i + _CHUNK_ROWS] for i in range(0, len(rows), _CHUNK_ROWS)]
    logger.info(f'GrowwithmeDB push: {len(rows)} rows -> {len(chunks)} chunks of <= {_CHUNK_ROWS}')

    # Chunk 0 goes ALONE and FIRST. It is the one that carries replace:true, so it
    # deletes the scope before anything is appended — running it beside the others
    # would let an append land before the wipe and be erased.
    ok, res = _post(path, {**payload, 'rows': chunks[0], 'replace': True})
    if not ok:
        return {'success': False, 'inserted': 0,
                'message': f'chunk 1/{len(chunks)} failed: {res} (0 row(s) written)'}, 502
    res = res or {}
    # An older API build ignores `replace` and wipes the whole date/month on
    # EVERY request — so chunk 2 would erase chunk 1 and only the final batch
    # would survive ("only a few rows uploaded"). A build that honours the flag
    # echoes replace_supported:true.
    #
    # When it doesn't, DON'T continue chunking and DON'T just fail: fall back to
    # the original behaviour and send the whole payload in one request. That is
    # exactly what an old server expects (it wipes and reloads the scope), so
    # the result is correct — it only risks the proxy's body limit, which is a
    # clear 413 rather than silent data loss.
    if not res.get('replace_supported'):
        logger.warning('GrowwithmeDB push: API does not support batching — '
                       'falling back to a single unchunked request.')
        ok1, res1 = _post(path, {**payload, 'replace': True})
        if not ok1:
            return {'success': False, 'inserted': 0, 'message': (
                f'{res1}\n\nThe GrowwithmeDB API is an older build that cannot accept batched '
                f'uploads, so the whole day had to be sent at once and the server refused it. '
                f'Deploy api/index.js and api/routes/collection.js and restart the API '
                f'(or raise nginx client_max_body_size), then re-run this sync.')}, 502
        res1 = res1 or {}
        ins, skp = int(res1.get('inserted') or 0), int(res1.get('skipped') or 0)
        msg = f'{ins} rows synced in a single request (API does not support batching yet)'
        if skp:
            msg += f' · {skp} skipped (employee not in GrowwithmeDB)'
        return {'success': True, 'inserted': ins, 'skipped': skp,
                'unresolved_officers': int(res1.get('unresolved_officers') or 0),
                'message': msg}, 200
    total_inserted += int(res.get('inserted') or 0)
    total_skipped += int(res.get('skipped') or 0)
    total_unresolved += int(res.get('unresolved_officers') or 0)

    # Every remaining chunk is a pure append (replace:false) with no ordering
    # relationship to the others, so they go out concurrently. This is where the
    # wall-clock went: a day's push is many sequential round-trips to the EC2, and
    # the trip time — not the row count — dominates.
    rest = chunks[1:]
    if rest:
        def _send(i_chunk):
            i, chunk = i_chunk
            return i, _post(path, {**payload, 'rows': chunk, 'replace': False})

        with ThreadPoolExecutor(max_workers=min(_PUSH_WORKERS, len(rest))) as pool:
            results = list(pool.map(_send, enumerate(rest, start=1)))

        failed = [(i, r) for i, (ok_i, r) in results if not ok_i]
        for i, (ok_i, r) in results:
            if not ok_i:
                continue
            r = r or {}
            total_inserted += int(r.get('inserted') or 0)
            total_skipped += int(r.get('skipped') or 0)
            total_unresolved += int(r.get('unresolved_officers') or 0)
        if failed:
            # The wipe already ran, so the scope is half-written either way.
            which = ', '.join(str(i + 1) for i, _ in failed)
            return {'success': False, 'inserted': total_inserted,
                    'message': (f'chunk(s) {which} of {len(chunks)} failed: {failed[0][1]} '
                                f'({total_inserted} row(s) written) — the date/month is now '
                                f'PARTIALLY written, re-run this sync')}, 502

    msg = f'{total_inserted} rows synced to growwithme-local (in {len(chunks)} batches)'
    if total_skipped:
        msg += f' · {total_skipped} skipped (employee not in GrowwithmeDB)'
    return {'success': True, 'inserted': total_inserted, 'skipped': total_skipped,
            'unresolved_officers': total_unresolved, 'message': msg}, 200


@growwithme_bp.route('/ping', methods=['GET'])
def ping():
    """Read-only reachability check — hits the open GET /api index."""
    try:
        resp = _session.get(f'{GROWWITHME_API_URL}/api', headers=_headers(), timeout=15)
    except http_requests.exceptions.RequestException as e:
        return jsonify({'success': False, 'url': GROWWITHME_API_URL,
                        'reachable': False, 'message': f'growwithme-api not reachable: {e}'}), 502
    body = {}
    try:
        body = resp.json()
    except ValueError:
        pass
    return jsonify({
        'success': resp.status_code == 200,
        'url': GROWWITHME_API_URL,
        'reachable': True,
        'httpStatus': resp.status_code,
        'database': body.get('database'),
        'message': 'growwithme-local API reachable.' if resp.status_code == 200
                   else f'Reachable, unexpected response (HTTP {resp.status_code}).',
    }), 200


# ── Input helpers (support both "sync the latest generated report" and "upload
#    a custom file") ─────────────────────────────────────────────────────────
def _param(name):
    """Read a field from the multipart form (file-upload requests) or JSON body."""
    if request.form and name in request.form:
        return (request.form.get(name) or '').strip()
    data = request.get_json(silent=True) or {}
    return str(data.get(name) or '').strip()


def _uploaded_file():
    """If the request carries an uploaded 'file', save it to a temp path and return
    (path, True). Otherwise (None, False). The caller deletes the temp path when the
    second value is True. Raises ValueError for a non-Excel upload."""
    f = request.files.get('file')
    if not f or not f.filename:
        return None, False
    if not f.filename.lower().endswith(('.xlsx', '.xls')):
        raise ValueError('Expected an Excel (.xlsx) file')
    import tempfile
    suffix = os.path.splitext(f.filename)[1] or '.xlsx'
    fd, tmp = tempfile.mkstemp(prefix='gwm_upload_', suffix=suffix)
    os.close(fd)
    f.save(tmp)
    return tmp, True


def _cleanup(path, is_temp):
    if is_temp and path:
        try:
            os.unlink(path)
        except OSError:
            pass


def _uploaded_named(field):
    """Same as _uploaded_file() but for an explicitly named multipart field, so the
    Daily tab can carry three workbooks in one request: `file` = the EOD/Daily
    Collection Report (accounts), `channel_file` = the Collection report whose
    CollectionChannel column drives Mode of Collection, `amount_file` = the EOD
    'Regular Demand vs Collection' output whose `_precomp` sheet carries the
    rupee amounts."""
    f = request.files.get(field)
    if not f or not f.filename:
        return None, False
    if not f.filename.lower().endswith(('.xlsx', '.xls')):
        raise ValueError(f'{field}: expected an Excel (.xlsx) file')
    import tempfile
    suffix = os.path.splitext(f.filename)[1] or '.xlsx'
    fd, tmp = tempfile.mkstemp(prefix=f'gwm_{field}_', suffix=suffix)
    os.close(fd)
    f.save(tmp)
    return tmp, True


# ── Mode of Collection (payment channel) ──────────────────────────────────
# Source: "Collection report as on <date>.xlsx", one row per receipt. Only four
# columns matter — Trxdate, OfficerID/OfficerName, AccountID, CollectionTotal —
# plus CollectionChannel itself. Everything else is ignored.

def _excel_date(v):
    """Trxdate arrives as an Excel serial (46238 = 2026-08-04). Also accepts a real
    date or an ISO-ish string. Returns 'YYYY-MM-DD' or None."""
    import datetime
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    if s.isdigit():
        # Excel's 1900 epoch, with its leap-year quirk: serial 1 == 1900-01-01.
        return (datetime.date(1899, 12, 30) + datetime.timedelta(days=int(s))).strftime('%Y-%m-%d')
    return s[:10] if len(s) >= 10 and s[4] == '-' else None


def _officer_employee_code(officer_id, officer_name):
    """Map the report's officer to a GrowwithmeDB employee_code candidate.

    The two identifier spaces differ: the report prints 'CHE030013064' /
    'CHETAN B HULAMANI(030013064)', while employee_code is 'NL13064'. The numeric
    part in the NAME's parentheses is authoritative; its tail matches the digits of
    employee_code. Returns the DIGITS, which the API matches against employee_code's
    numeric tail — resolving ~96% of officers, with the rest stored unmapped
    rather than dropped.
    """
    import re
    m = re.search(r'\((\d+)\)', str(officer_name or ''))
    digits = m.group(1) if m else re.sub(r'\D', '', str(officer_id or ''))
    return digits.lstrip('0') or None


def _parse_collection_channels(path, date):
    """Aggregate a Collection report's CollectionChannel column for ONE date.

    Returns [{officer_code, officer_digits, channel, product_id, accounts, amount,
              reversed}] where `accounts` counts DISTINCT AccountID (a client can
    pay more than once a day) and `amount` sums CollectionTotal.

    The workbook typically holds several days; rows outside `date` are ignored, so
    uploading the same file on consecutive days stays correct.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        header = [str(c).strip() if c is not None else '' for c in next(it)]
        idx = {h: i for i, h in enumerate(header)}
        if 'CollectionChannel' not in idx:
            raise ValueError('No CollectionChannel column — is this the Collection report?')
        i_ch = idx['CollectionChannel']
        i_dt, i_acc = idx.get('Trxdate'), idx.get('AccountID')
        i_oid, i_onm = idx.get('OfficerID'), idx.get('OfficerName')
        i_tot, i_rev, i_pid = idx.get('CollectionTotal'), idx.get('ReverseTotal'), idx.get('ProductID')

        def cell(row, i):
            return row[i] if i is not None and i < len(row) else None

        agg, seen_dates, matched = {}, set(), 0
        for row in it:
            if not row or i_ch >= len(row) or row[i_ch] is None:
                continue
            d = _excel_date(cell(row, i_dt))
            if d:
                seen_dates.add(d)
            # No Trxdate at all -> keep the row (the file is already date-scoped by
            # its filename); otherwise take only the requested day.
            if d and d != date:
                continue
            matched += 1
            channel = str(row[i_ch]).strip().upper()
            if not channel:
                continue
            oid = str(cell(row, i_oid) or '').strip()
            key = (oid, channel, str(cell(row, i_pid) or '').strip() or None)
            slot = agg.get(key)
            if slot is None:
                slot = agg[key] = {
                    'officer_code': oid or None,
                    'officer_digits': _officer_employee_code(oid, cell(row, i_onm)),
                    'channel': channel,
                    'product_id': key[2],
                    'accounts': set(),
                    'amount': 0.0,
                    'reversed': 0.0,
                }
            acc = cell(row, i_acc)
            if acc is not None:
                slot['accounts'].add(str(acc))
            slot['amount'] += _num(cell(row, i_tot))
            slot['reversed'] += _num(cell(row, i_rev))

        if not agg:
            raise ValueError(
                f'No rows for {date}. The file covers {sorted(seen_dates) or "no recognisable dates"}.')
        logger.info(f'GrowwithmeDB channel parse: {matched} receipt rows for {date} '
                    f'-> {len(agg)} officer×channel groups (file covers {sorted(seen_dates)}).')
        return [{**v, 'accounts': len(v['accounts'])} for v in agg.values()]
    finally:
        wb.close()


def _parse_collection_channels_by_date(path, default_date):
    """Like _parse_collection_channels, but aggregates EVERY Trxdate in the file:
    {date: [rows]}. Rows with no usable Trxdate land under `default_date` (the
    file is date-scoped by its name). Raises when `default_date` has no rows —
    same contract as the single-date parser.

    Exists because the Collection report spans several days and receipts for an
    earlier date keep appearing in later exports (late postings). Pushing every
    date the file carries — each a whole-date replace on the API side — keeps the
    Mode of Collection page complete without anyone re-running old dates by hand.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        header = [str(c).strip() if c is not None else '' for c in next(it)]
        idx = {h: i for i, h in enumerate(header)}
        if 'CollectionChannel' not in idx:
            raise ValueError('No CollectionChannel column — is this the Collection report?')
        i_ch = idx['CollectionChannel']
        i_dt, i_acc = idx.get('Trxdate'), idx.get('AccountID')
        i_oid, i_onm = idx.get('OfficerID'), idx.get('OfficerName')
        i_tot, i_rev, i_pid = idx.get('CollectionTotal'), idx.get('ReverseTotal'), idx.get('ProductID')

        def cell(row, i):
            return row[i] if i is not None and i < len(row) else None

        agg, seen_dates = {}, set()   # (date, oid, channel, pid) -> slot
        for row in it:
            if not row or i_ch >= len(row) or row[i_ch] is None:
                continue
            d = _excel_date(cell(row, i_dt))
            if d:
                seen_dates.add(d)
            day = d or default_date
            channel = str(row[i_ch]).strip().upper()
            if not channel:
                continue
            oid = str(cell(row, i_oid) or '').strip()
            key = (day, oid, channel, str(cell(row, i_pid) or '').strip() or None)
            slot = agg.get(key)
            if slot is None:
                slot = agg[key] = {
                    'officer_code': oid or None,
                    'officer_digits': _officer_employee_code(oid, cell(row, i_onm)),
                    'channel': channel,
                    'product_id': key[3],
                    'accounts': set(),
                    'amount': 0.0,
                    'reversed': 0.0,
                }
            acc = cell(row, i_acc)
            if acc is not None:
                slot['accounts'].add(str(acc))
            slot['amount'] += _num(cell(row, i_tot))
            slot['reversed'] += _num(cell(row, i_rev))

        out = {}
        for (day, *_rest), v in agg.items():
            out.setdefault(day, []).append({**v, 'accounts': len(v['accounts'])})
        if not out.get(default_date):
            raise ValueError(
                f'No rows for {default_date}. The file covers {sorted(seen_dates) or "no recognisable dates"}.')
        logger.info(f'GrowwithmeDB channel parse: {len(out)} date(s) in file '
                    f'({sorted(out)}), syncing {default_date} + refreshing the rest.')
        return out
    finally:
        wb.close()


# ── Daily Collection Report (raw) → daily collection sync ─────────────────
# The daily sync normally ingests a generated EOD Employee Report (per-product
# IGL/FIG/VVY sheets, via _parse_report). A raw "Daily Collection Report" instead
# carries one combined per-employee row in an 'Employee Data' sheet — same bucket
# column layout the hourly parser (_parse_employee_data_sheet) already handles, but
# count-only (no rupee amounts, like the hourly grain) and with the employee CODE
# in a different column than the label 'EMP ID' (that column can hold a name).
_EMP_CODE_RE = re.compile(r'^[A-Za-z]{1,3}\d{3,}$')  # e.g. NL10838


def _has_employee_data_sheet(path):
    """True if the workbook has an 'Employee Data' sheet (a raw Daily Collection
    Report) rather than the per-product EOD Employee Report sheets."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            return 'Employee Data' in wb.sheetnames
        finally:
            wb.close()
    except Exception:
        return False


def _parse_daily_collection(path):
    """Parse a raw Daily Collection Report's 'Employee Data' sheet into the same row
    dicts _parse_report returns ({emp_id, product_type_id, <25 metrics>}), so it can
    feed the daily collection sync. Combined across products (product_type_id=None);
    amounts are 0 (this report is count-only, matching the hourly grain).

    The employee CODE column is detected by content (values like NL10838), because
    in this report the column literally headed 'EMP ID' can hold the officer NAME
    while the code sits under 'EMP Name'. Bucket columns match the hourly layout:
    Regular D/C = 7/8, 1-30 = 11/12, 31-60 = 15/16, PNPA = 19/20, NPA cases = 27,
    activation acc/amt = 28/29, closure acc/amt = 30/31.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb['Employee Data']
        rows = list(ws.iter_rows(values_only=True))
        # Header row = the one that contains an 'EMP ID' cell.
        hdr_idx = -1
        for r, row in enumerate(rows):
            if row and any(c is not None and str(c).strip().upper() == 'EMP ID' for c in row):
                hdr_idx = r
                break
        if hdr_idx < 0:
            raise ValueError("Daily Collection Report 'Employee Data' sheet has no 'EMP ID' header.")

        data = rows[hdr_idx + 1:]
        # Candidate emp columns: headers mentioning 'emp'. Pick the one whose values
        # best match the employee-code pattern (NL#####).
        header = rows[hdr_idx]
        candidates = [i for i, c in enumerate(header) if c is not None and 'emp' in str(c).strip().lower()]
        if not candidates:
            candidates = list(range(min(4, len(header))))

        def _code_hits(ci):
            hits = 0
            for row in data[:40]:
                if row and ci < len(row) and row[ci] is not None and _EMP_CODE_RE.match(str(row[ci]).strip()):
                    hits += 1
            return hits

        emp_col = max(candidates, key=_code_hits)
        if _code_hits(emp_col) == 0:
            # No code-like column found — fall back to the literal 'EMP ID' column.
            emp_col = next((i for i, c in enumerate(header)
                            if c is not None and str(c).strip().upper() == 'EMP ID'), candidates[0])

        def g(row, i):
            return _safe_num(row[i]) if i < len(row) else 0

        out = []
        for row in data:
            if not row or emp_col >= len(row) or row[emp_col] is None:
                continue
            emp = str(row[emp_col]).strip()
            if not emp or emp.upper() in ('EMP ID', 'GRAND TOTAL', 'TOTAL'):
                continue
            out.append({
                'emp_id': emp,
                'product_type_id': None,  # combined report — no product split
                'regular_demand': g(row, 7), 'regular_collection': g(row, 8),
                'demand_1_30': g(row, 11), 'collection_1_30': g(row, 12),
                'demand_31_60': g(row, 15), 'collection_31_60': g(row, 16),
                'pnpa_demand': g(row, 19), 'pnpa_collection': g(row, 20),
                'npa_cases': g(row, 27),
                'npa_act_acc': g(row, 28), 'npa_act_amt': g(row, 29),
                'npa_clo_acc': g(row, 30), 'npa_clo_amt': g(row, 31),
                'on_date_demand': 0, 'on_date_collection': 0,
                'regular_demand_amt': 0, 'regular_collection_amt': 0,
                'demand_1_30_amt': 0, 'collection_1_30_amt': 0,
                'demand_31_60_amt': 0, 'collection_31_60_amt': 0,
                'pnpa_demand_amt': 0, 'pnpa_collection_amt': 0,
                'on_date_demand_amt': 0, 'on_date_collection_amt': 0,
            })
        return out
    finally:
        wb.close()


# ── Product-split daily source (preferred) ────────────────────────────────
# The raw Daily Collection Report's FIRST sheet ('OverAll') carries BRANCH +
# OFFICER blocks split by product — titled 'BRANCH + OFFICER NAME WISE
# COLLECTION REPORT (OverAll - IGL / - FIG / - IL)'. Parsing those recovers the
# product_type_id the combined 'Employee Data' sheet drops (it hardcodes None),
# so the frontend's IGL/FIG/IL tabs populate instead of going blank. Count-only
# (amounts 0), same daily grain + whole-date override as _parse_daily_collection.
_OVERALL_PRODUCT_BLOCKS = [('(OVERALL - IGL)', 1), ('(OVERALL - FIG)', 2), ('(OVERALL - IL)', 3)]
# OverAll officer-block column indices (0-based within the row tuple).
_OA = {
    'name': 1, 'on_date_demand': 2, 'on_date_collection': 3,
    'regular_demand': 5, 'regular_collection': 6,
    'demand_1_30': 9, 'collection_1_30': 10,
    'demand_31_60': 13, 'collection_31_60': 14,
    'pnpa_demand': 17, 'pnpa_collection': 18,
    'npa_cases': 25, 'npa_act_acc': 26, 'npa_act_amt': 27,
    'npa_clo_acc': 28, 'npa_clo_amt': 29,
}


def _overall_block_title_rows(rows):
    """Map product_type_id -> 0-based index (into `rows`) of each product block's
    'BRANCH + OFFICER NAME WISE ... (OverAll - IGL/FIG/IL)' title, by scanning
    column A. The combined '(OverAll)' block (no product suffix) is ignored."""
    found = {}
    for ri, row in enumerate(rows):
        a = row[0] if row else None
        if a is None:
            continue
        s = str(a).strip().upper()
        if not s.startswith('BRANCH + OFFICER'):
            continue
        for suffix, pt in _OVERALL_PRODUCT_BLOCKS:
            if s.endswith(suffix):
                found[pt] = ri
    return found


def _has_overall_product_blocks(path):
    """True if the workbook's 'OverAll' sheet carries all three product-split
    officer blocks (IGL/FIG/IL) — the preferred, product-aware daily source."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            if 'OverAll' not in wb.sheetnames:
                return False
            rows = list(wb['OverAll'].iter_rows(values_only=True))
            return len(_overall_block_title_rows(rows)) == 3
        finally:
            wb.close()
    except Exception:
        return False


def _officer_code(label):
    """Extract the officer code (e.g. NL10838) from an officer block label, or
    None for area/branch subtotal + header rows.

    Both label orders appear in the wild — the report source switched from
    'OFFICER NAME - NL#####' to 'NL##### - OFFICER NAME' (seen 02-08-2026) — so
    every '-' separated segment is checked rather than just the tail. Names can
    themselves contain hyphens, hence the scan instead of a fixed position."""
    if label is None:
        return None
    for part in re.split(r'[-–]', str(label)):
        part = part.strip()
        if _EMP_CODE_RE.match(part):
            return part
    return None


def _officer_name(label, code):
    """The officer's name from a block label, i.e. everything except the code.
    Works for both label orders ('NAME - CODE' and 'CODE - NAME')."""
    parts = [p.strip() for p in re.split(r'[-–]', str(label or ''))]
    return ' - '.join(p for p in parts if p and p != code).strip()


# Non-officer labels inside a block: the area subtotal ('BADAMI AREA'), the
# branch heading ('BADAMI'), the column header and the grand total. Only the
# branch heading is worth keeping — it names the branch the officers below sit in.
def _branch_label(label):
    """Return `label` when it reads as a branch heading, else None."""
    s = str(label or '').strip()
    u = s.upper()
    if not s or u.endswith(' AREA') or u.startswith('BRANCH') or u.startswith('GRAND'):
        return None
    return s


def _parse_daily_collection_products(path):
    """Parse the raw Daily Collection Report's 'OverAll' sheet product officer
    blocks into flat metric rows carrying product_type_id (1 IGL / 2 FIG / 3 IL),
    so _explode feeds the daily collection sync WITH the product split. Only
    officer rows (label carries an NL##### code) are kept; area/branch subtotal
    and header rows are skipped. Count-only (amounts 0), like the hourly grain.

    Each row also carries the officer's `emp_name` and the `branch` heading it
    sits under (the nearest preceding non-officer, non-AREA label). _explode
    ignores both; _onboard_missing uses them to create employees the target DB
    doesn't know yet, so a new joiner's numbers aren't silently skipped."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        rows = list(wb['OverAll'].iter_rows(values_only=True))
    finally:
        wb.close()
    titles = _overall_block_title_rows(rows)
    if not titles:
        raise ValueError("Daily Collection Report 'OverAll' sheet has no product blocks.")
    # Each block runs from its title to the next block title (in sheet order).
    starts = sorted(titles.values())
    end_of = {s: (starts[i + 1] if i + 1 < len(starts) else len(rows))
              for i, s in enumerate(starts)}

    def g(row, key):
        i = _OA[key]
        return _safe_num(row[i]) if i < len(row) else 0

    out = []
    for pt, start in titles.items():
        branch = None
        for ri in range(start + 1, end_of[start]):
            row = rows[ri]
            if not row:
                continue
            label = row[_OA['name']] if _OA['name'] < len(row) else None
            code = _officer_code(label)
            if not code:
                branch = _branch_label(label) or branch
                continue
            out.append({
                'emp_id': code, 'product_type_id': pt,
                'emp_name': _officer_name(label, code), 'branch': branch,
                'regular_demand': g(row, 'regular_demand'), 'regular_collection': g(row, 'regular_collection'),
                'demand_1_30': g(row, 'demand_1_30'), 'collection_1_30': g(row, 'collection_1_30'),
                'demand_31_60': g(row, 'demand_31_60'), 'collection_31_60': g(row, 'collection_31_60'),
                'pnpa_demand': g(row, 'pnpa_demand'), 'pnpa_collection': g(row, 'pnpa_collection'),
                'on_date_demand': g(row, 'on_date_demand'), 'on_date_collection': g(row, 'on_date_collection'),
                'npa_cases': g(row, 'npa_cases'),
                'npa_act_acc': g(row, 'npa_act_acc'), 'npa_act_amt': g(row, 'npa_act_amt'),
                'npa_clo_acc': g(row, 'npa_clo_acc'), 'npa_clo_amt': g(row, 'npa_clo_amt'),
                'regular_demand_amt': 0, 'regular_collection_amt': 0,
                'demand_1_30_amt': 0, 'collection_1_30_amt': 0,
                'demand_31_60_amt': 0, 'collection_31_60_amt': 0,
                'pnpa_demand_amt': 0, 'pnpa_collection_amt': 0,
                'on_date_demand_amt': 0, 'on_date_collection_amt': 0,
            })
    return out


# ── Rupee amounts from the EOD workbook's `_precomp` sheet ────────────────
def _eod_output_path(date=None):
    """Locate the EOD 'Regular Demand vs Collection' workbook (the one carrying
    `_precomp`) for `date`.

    The per-date ARCHIVE is preferred over EOD_Output_Latest.xlsx, and that matters
    for correctness, not tidiness: *_Latest holds whichever date EOD last ran for.
    Syncing 04-08 the morning after a 05-08 EOD run would otherwise staple 05-08's
    rupees onto 04-08's accounts — silently, since both files parse fine. Only when
    no archived run exists for the date do we fall back to *_Latest.

    Returns (path_or_None, matched_the_date). The caller reports a False on the
    second element, because that is precisely the case where the rupees may belong
    to another day and nothing else would say so.
    """
    if date:
        day = str(date)[:10]
        root = config.DATA_DIR / 'reports_archive' / day
        if root.is_dir():
            # Run folders are HH-MM-SS[-n]; the newest is the authoritative one.
            for run in sorted((p for p in root.iterdir() if p.is_dir()),
                              key=lambda p: p.name, reverse=True):
                p = run / 'Regular Demand vs Collection.xlsx'
                if p.exists():
                    return p, True
    p = config.BACKEND_DATA_DIR / 'EOD_Output_Latest.xlsx'
    return (p, not date) if p.exists() else (None, False)


_XL_NS = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
_XL_REL_NS = '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}'
_PKG_REL_NS = '{http://schemas.openxmlformats.org/package/2006/relationships}'
_CELL_REF_RE = re.compile(r'^([A-Z]+)')


def _col_index(ref):
    """'A1' / 'BC7' -> 0-based column index."""
    letters = _CELL_REF_RE.match(ref or '')
    if not letters:
        return None
    n = 0
    for ch in letters.group(1):
        n = n * 26 + (ord(ch) - 64)
    return n - 1


def _iter_sheet_rows(path, sheet_name):
    """Yield each row of `sheet_name` as a list of values, reading the .xlsx zip directly.

    openpyxl is not used here for a blunt reason: `load_workbook(read_only=True)`
    on the 56 MB EOD output takes ~40 s before a single cell is read, because the
    workbook also holds a 261k-row / 422 MB Sheet1. Pulling just this one sheet's
    XML out of the zip does the same job in under a second, which is the whole
    difference between a slow and a quick Local DB Sync.

    Yields nothing when the sheet is absent. Cells are str/float/None; sparse rows
    are padded so column indexes stay aligned with the header.
    """
    with zipfile.ZipFile(path) as z:
        names = set(z.namelist())
        if 'xl/workbook.xml' not in names:
            return
        wb_root = ET.fromstring(z.read('xl/workbook.xml'))
        rid = None
        for el in wb_root.iter(_XL_NS + 'sheet'):
            if (el.get('name') or '') == sheet_name:
                rid = el.get(_XL_REL_NS + 'id')
                break
        if rid is None:
            return
        target = None
        if 'xl/_rels/workbook.xml.rels' in names:
            rels = ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))
            for el in rels.iter(_PKG_REL_NS + 'Relationship'):
                if el.get('Id') == rid:
                    target = el.get('Target')
                    break
        if not target:
            return
        target = target.lstrip('/')
        member = target if target in names else f'xl/{target}'
        if member not in names:
            return

        # Shared strings: every text cell is an index into this table.
        shared = []
        if 'xl/sharedStrings.xml' in names:
            for _, si in ET.iterparse(io.BytesIO(z.read('xl/sharedStrings.xml')),
                                      events=('end',)):
                if si.tag == _XL_NS + 'si':
                    shared.append(''.join(t.text or '' for t in si.iter(_XL_NS + 't')))
                    si.clear()

        for _, row in ET.iterparse(io.BytesIO(z.read(member)), events=('end',)):
            if row.tag != _XL_NS + 'row':
                continue
            values = []
            for c in row.iter(_XL_NS + 'c'):
                idx = _col_index(c.get('r'))
                if idx is None:
                    idx = len(values)
                while len(values) < idx:
                    values.append(None)
                ctype = c.get('t')
                if ctype == 'inlineStr':
                    val = ''.join(t.text or '' for t in c.iter(_XL_NS + 't'))
                else:
                    v = c.find(_XL_NS + 'v')
                    raw = None if v is None else v.text
                    if raw is None:
                        val = None
                    elif ctype == 's':
                        try:
                            val = shared[int(raw)]
                        except (ValueError, IndexError):
                            val = None
                    elif ctype in ('str', 'e'):
                        val = raw
                    elif ctype == 'b':
                        val = raw == '1'
                    else:
                        try:
                            val = float(raw)
                        except ValueError:
                            val = raw
                values.append(val)
            row.clear()
            yield values


def _parse_precomp_amounts(path):
    """Read per-officer rupee amounts from the EOD workbook's `_precomp` sheet.

    Returns {(EMP_CODE, product_type_id): {metric_field: amount}} for the overall
    ('OA') scope only, where product_type_id is None for the combined 'ALL' row.
    Returns {} when the sheet is absent, so an older workbook simply leaves the
    amounts at 0 instead of failing the sync.

    Reads the sheet straight out of the .xlsx zip (see `_iter_sheet_rows`); the
    openpyxl path is kept only as a fallback for anything that reader chokes on.
    """
    try:
        return _parse_precomp_amounts_fast(path)
    except Exception as e:
        logger.warning(f'_precomp fast reader failed ({e}); falling back to openpyxl')
        return _parse_precomp_amounts_openpyxl(path)


def _precomp_rows_to_amounts(it):
    """Shared row->amounts reduction for both `_precomp` readers.

    `it` yields the sheet's rows as sequences, header row first.
    """
    try:
        header = [str(c).strip() if c is not None else '' for c in next(it)]
    except StopIteration:
        return {}
    idx = {h: i for i, h in enumerate(header)}
    # Nothing to take if this build predates the amount columns.
    if 'reg_demand_amt' not in idx:
        return {}
    i_ft, i_grp, i_sc, i_pr = (idx.get('filter_type'), idx.get('group_value'),
                               idx.get('scope'), idx.get('product'))
    if None in (i_ft, i_grp, i_sc, i_pr):
        return {}

    def cell(row, i):
        return row[i] if i is not None and i < len(row) else None

    out = {}
    for row in it:
        if not row or str(cell(row, i_ft) or '') != 'All_EmpID':
            continue
        if str(cell(row, i_sc) or '').strip().upper() != 'OA':
            continue
        emp = str(cell(row, i_grp) or '').strip().upper()
        # 'All_EmpID' also carries a 'Grand Total' row holding the whole day.
        # It matches no real officer today, but letting it through would put the
        # entire day's rupees on one row the moment a parser emitted that label.
        if not emp or not _EMP_CODE_RE.match(emp):
            continue
        label = str(cell(row, i_pr) or 'ALL').strip().upper()
        if label not in _PRECOMP_PRODUCT_ID:
            continue
        out[(emp, _PRECOMP_PRODUCT_ID[label])] = {
            field: _safe_num(cell(row, idx[col]))
            for field, col in _PRECOMP_AMT_MAP.items() if col in idx
        }
    return out


def _parse_precomp_amounts_fast(path):
    """`_precomp` read straight from the .xlsx zip — the normal path."""
    return _precomp_rows_to_amounts(_iter_sheet_rows(path, '_precomp'))


# ── Officer -> branch/name directory, also from `_precomp` ────────────────
# Why this exists: the auto-onboard below used to create employees with an EMPTY
# name and branch whenever the sync had parsed the 'Employee Data' sheet, which
# carries neither. /api/employees/onboard then made an `employee` row with NO
# employee_assignment — and an officer with no assignment resolves to a NULL
# region, so every by-<level> endpoint (they all end `HAVING ... IS NOT NULL`)
# drops them. That is the drill table reading LOWER than the summary card: the
# card counts those officers, the breakdown cannot place them.
#
# The EOD workbook already knows where they sit. `_precomp` carries
# filter_type='BranchName' rows whose filter_value is the BRANCH and whose
# group_value is the EMPLOYEE CODE, plus officer_name — exactly the two fields
# onboarding was missing. It is the same sheet already being read for amounts.
def _precomp_rows_to_directory(it):
    """{EMP_CODE: {'name': str, 'branch': str}} from `_precomp` BranchName rows."""
    try:
        header = [str(c).strip() if c is not None else '' for c in next(it)]
    except StopIteration:
        return {}
    idx = {h: i for i, h in enumerate(header)}
    i_ft, i_fv, i_grp = idx.get('filter_type'), idx.get('filter_value'), idx.get('group_value')
    i_nm = idx.get('officer_name')
    if None in (i_ft, i_fv, i_grp):
        return {}

    def cell(row, i):
        return row[i] if i is not None and i < len(row) else None

    out = {}
    for row in it:
        if not row or str(cell(row, i_ft) or '') != 'BranchName':
            continue
        emp = str(cell(row, i_grp) or '').strip().upper()
        # group_value is the branch's own subtotal on non-officer rows; keep only
        # real employee codes.
        if not emp or not _EMP_CODE_RE.match(emp) or emp in out:
            continue
        branch = str(cell(row, i_fv) or '').strip()
        if not branch:
            continue
        out[emp] = {'name': str(cell(row, i_nm) or '').strip(), 'branch': branch}
    return out


def _parse_precomp_directory(path):
    """Officer -> {name, branch} from the EOD workbook. {} if unavailable."""
    try:
        return _precomp_rows_to_directory(_iter_sheet_rows(path, '_precomp'))
    except Exception as e:
        logger.warning(f'_precomp directory read failed ({e}); auto-onboard will '
                       f'fall back to whatever the report row carried')
        return {}


def _parse_precomp_amounts_openpyxl(path):
    """`_precomp` read via openpyxl — correct but slow; fallback only."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if '_precomp' not in wb.sheetnames:
            return {}
        return _precomp_rows_to_amounts(wb['_precomp'].iter_rows(values_only=True))
    finally:
        wb.close()


def _apply_precomp_amounts(rows, amounts):
    """Fill each parsed officer row's `*_amt` fields from `amounts`, in place.

    Matches on (employee_code, product_type_id). A product-split row falls back to
    the combined 'ALL' entry only when its own product has no `_precomp` row, so
    the per-product tabs never silently inherit another product's rupees.

    FILL-ONLY: a row that already carries real rupees is left untouched. The EOD
    Employee Report path (_parse_report) reads genuine `* amt` columns straight
    from its sheets; only the Daily Collection Report parsers, which have no rupee
    columns to read, zero-fill them. So this tops up what's missing instead of
    overwriting what the report already knew.

    Returns the number of rows that received a non-zero amount — the caller
    surfaces it, because a 0 here means the Amount view will stay empty.
    """
    if not amounts:
        return 0
    fields = list(_PRECOMP_AMT_MAP)
    filled = 0
    for r in rows:
        emp = str(r.get('emp_id') or '').strip().upper()
        if not emp:
            continue
        if any(r.get(f) for f in fields):
            continue  # already has rupees from the report itself
        pt = r.get('product_type_id')
        amt = amounts.get((emp, pt))
        if amt is None and pt is not None:
            amt = amounts.get((emp, None))
        if not amt:
            continue
        r.update(amt)
        if any(v for v in amt.values()):
            filled += 1
    return filled


# ── Auto-onboard employees the target DB doesn't know ─────────────────────
# /api/collection/sync resolves each emp_id against GrowwithmeDB's `employee`
# table and SKIPS the row when the code isn't there, so a new joiner present in
# the report but not yet onboarded loses their numbers for the date. The report
# carries the officer's name and branch, which is everything /api/employees/
# onboard needs to create them (scope 'self', default must-change login — same
# as an access-sheet upload), so the sync creates them and pushes again.
def _onboard_missing(rows, roster=None):
    """Create report employees that GrowwithmeDB doesn't have yet.

    Returns (created_count, message_or_None). Never raises — a failure here
    leaves the push exactly as it was (those rows stay skipped).

    `roster` is the {EMP_CODE: {name, branch}} directory from the EOD workbook
    (_parse_precomp_directory). It exists because creating an employee with NO
    BRANCH is actively harmful: onboarding makes an `employee` row but no
    `employee_assignment`, the hierarchy join then yields a NULL region, and
    every by-<level> endpoint drops that officer while the summary card still
    counts them. That is exactly the mismatch between the drill table and the
    card. The 'Employee Data' sheet carries no name or branch at all, so a sync
    from that parser used to mint one broken employee per new joiner, every day.

    Any officer still lacking a branch after the roster lookup is REPORTED by
    code rather than created silently — they are the only ones a human has to
    place by hand.
    """
    roster = roster or {}
    directory = {}
    for r in rows:
        code = r.get('emp_id')
        if not code or code in directory:
            continue
        ref = roster.get(str(code).strip().upper()) or {}
        directory[code] = {
            'emp_id': code,
            # Prefer what the report row itself carried; fall back to the EOD
            # workbook, which knows every officer's branch even when the parsed
            # sheet does not.
            'name': r.get('emp_name') or ref.get('name') or '',
            'branch': r.get('branch') or ref.get('branch') or '',
            'scope': 'self',
        }
    if not directory:
        return 0, 0, None

    from concurrent.futures import ThreadPoolExecutor

    def state(code):
        """(exists, has_branch). On a lookup error claim both, so a transient
        failure leaves the row skipped rather than creating a duplicate."""
        ok, res = _post('/api/employees/lookup', {'emp_id': code})
        if not ok:
            return True, True
        res = res or {}
        return bool(res.get('found')), bool(str(res.get('branch') or '').strip())

    codes = list(directory)
    try:
        with ThreadPoolExecutor(max_workers=8) as pool:
            states = list(pool.map(state, codes))
    except Exception as e:
        logger.warning(f'GrowwithmeDB daily sync: employee lookup failed: {e}')
        return 0, 0, None

    # Two groups, both sent to /api/employees/onboard — it creates an employee
    # that does not exist, and for one that does it UPDATE-or-INSERTs the
    # assignment, so the same call repairs either.
    #
    # REPAIRING EXISTING OFFICERS IS THE POINT: an employee created by an earlier
    # sync (before the roster lookup existed) sits in `employee` with NO
    # assignment. Onboarding used to skip them entirely — they "exist", so
    # nothing was wrong as far as it could tell — and they stayed invisible to
    # every region/branch drill for good. Now a re-upload heals them, instead of
    # needing a hand-written employee_assignment INSERT.
    new = [directory[c] for c, (ex, _) in zip(codes, states) if not ex]
    repair = [directory[c] for c, (ex, has_b) in zip(codes, states)
              if ex and not has_b and directory[c]['branch']]
    missing = new + repair
    if not missing:
        return 0, 0, None

    # Still branchless after the roster lookup. These are created anyway —
    # dropping them would silently understate the day's collection, which is
    # worse than an officer who is counted but not yet drillable — but they are
    # named in the response so somebody assigns them instead of finding out
    # months later via a total that does not add up.
    no_branch = [m['emp_id'] for m in missing if not m['branch']]

    ok, res = _post('/api/employees/onboard', {'rows': missing})
    if not ok:
        logger.warning(f'GrowwithmeDB daily sync: auto-onboard failed: {res}')
        return 0, 0, f'{len(missing)} unknown employee(s) could not be created: {res}'
    res = res or {}
    created = int(res.get('created') or 0)
    assigned = int(res.get('assignments_set') or 0)
    logger.info(f"GrowwithmeDB daily sync: auto-onboarded {created} new employee(s), "
                f"repaired {len(repair)} existing without a branch, "
                f"{assigned} assignment(s) set: {', '.join(m['emp_id'] for m in missing)}")
    if repair:
        # Report it: this is the branch that closes the drill-table-vs-card gap,
        # and it is worth knowing a re-upload did it rather than a DBA.
        logger.info('GrowwithmeDB daily sync: branch assigned to existing officer(s): '
                    + ', '.join(sorted(m['emp_id'] for m in repair)))
    if no_branch:
        shown = ', '.join(sorted(no_branch)[:15])
        more = f' (+{len(no_branch) - 15} more)' if len(no_branch) > 15 else ''
        warn = (f'{len(no_branch)} new employee(s) had NO BRANCH in the report or the '
                f'EOD workbook, so they have no assignment and will NOT appear in any '
                f'region/branch drill until one is set: {shown}{more}')
        logger.warning(f'GrowwithmeDB daily sync: {warn}')
        return created, len(repair), warn
    return created, len(repair), None


@growwithme_bp.route('/sync-daily', methods=['POST'])
def sync_daily():
    """Push an EOD Employee Report into GrowwithmeDB (collection grain 2).

    Body: {"date": "YYYY-MM-DD"} as JSON, or multipart form with the same `date`
    field plus an optional `file` (an Employee Report .xlsx). With a file, that
    file is parsed; without one, the latest generated report is used.
    NOTE: insert-only — re-running for the same date appends new rows.
    """
    date = _param('date')
    if not date:
        return jsonify({'success': False, 'message': 'date is required (YYYY-MM-DD)'}), 400

    try:
        up, is_temp = _uploaded_file()
    except ValueError as e:
        return jsonify({'success': False, 'message': str(e)}), 400
    path = up or _employee_report_path()
    if not path:
        return jsonify({'success': False,
                        'message': 'No EOD Employee Report found. Run EOD processing first, or upload one.'}), 404
    try:
        if _has_overall_product_blocks(path):
            # Raw Daily Collection Report — parse its 'OverAll' sheet product
            # officer blocks (IGL/FIG/IL) so rows carry product_type_id and the
            # frontend's product tabs populate. Preferred over the combined
            # 'Employee Data' sheet, which drops the product split.
            logger.info('GrowwithmeDB daily sync: parsing raw Daily Collection Report (OverAll product blocks).')
            rows = _parse_daily_collection_products(path)
        elif _has_employee_data_sheet(path):
            # Fallback: older report with only the combined 'Employee Data' sheet
            # (count-only, product-combined). Same daily grain, whole-date override.
            logger.info('GrowwithmeDB daily sync: parsing raw Daily Collection Report (Employee Data sheet, product-combined).')
            rows = _parse_daily_collection(path)
        else:
            rows = _parse_report(path)
    except Exception as e:
        logger.warning(f'GrowwithmeDB daily sync: report parse failed: {e}')
        return jsonify({'success': False, 'message': f'Report parse failed: {e}'}), 500
    finally:
        _cleanup(up, is_temp)
    if not rows:
        return jsonify({'success': False, 'message': 'No employee rows found in report.'}), 400

    # ── Rupee amounts ────────────────────────────────────────────────────
    # The report parsed above is count-only, so without this every *_amt goes up
    # as 0 and the MIS Amount view has nothing to render. Pull the amounts from
    # the EOD workbook's `_precomp` sheet — the `amount_file` upload (Daily tab
    # box 3) when given, else the EOD run archived for this date — and match them
    # on to these rows by employee_code. Best-effort on purpose: a missing or stale
    # workbook degrades to the previous count-only behaviour rather than failing
    # the whole sync, but it always says so in the response.
    amt_filled, amt_note = 0, None
    amt_up, amt_temp = None, False
    try:
        amt_up, amt_temp = _uploaded_named('amount_file')
    except ValueError as e:
        amt_note = str(e)
    found, dated = _eod_output_path(date)
    amt_path = amt_up or found
    # Officer -> {name, branch} from the SAME workbook, read here because the
    # uploaded copy is deleted in the `finally` below, long before _onboard_missing
    # runs. Without it a new joiner is created with no branch and no assignment,
    # and then never appears in a region/branch drill. See _onboard_missing.
    roster = {}
    if amt_path and not amt_note:
        try:
            roster = _parse_precomp_directory(amt_path)
            if roster:
                logger.info(f'GrowwithmeDB daily sync: officer roster from EOD workbook '
                            f'— {len(roster)} officer(s) with a branch.')
            amt_filled = _apply_precomp_amounts(rows, _parse_precomp_amounts(amt_path))
            if amt_filled and not amt_up and not dated:
                # Fell back to EOD_Output_Latest.xlsx because no archived EOD run
                # exists for this date. It may hold a DIFFERENT day's rupees.
                amt_note = (f'amounts came from EOD_Output_Latest.xlsx, not an EOD run archived '
                            f'for {date} — CHECK they belong to {date}')
                logger.warning(f'GrowwithmeDB daily sync: {amt_note}')
            if amt_filled:
                logger.info(f'GrowwithmeDB daily sync: rupee amounts applied to {amt_filled}/{len(rows)} rows.')
            else:
                amt_note = ('the EOD workbook has no _precomp amounts for these officers '
                            '— Amount view will stay empty')
        except Exception as e:
            amt_note = f'amount lookup failed ({e})'
            logger.warning(f'GrowwithmeDB daily sync: {amt_note}')
        finally:
            _cleanup(amt_up, amt_temp)
    elif not amt_note:
        amt_note = 'no EOD_Output_Latest.xlsx found — pushed counts only, no rupee amounts'
        logger.warning(f'GrowwithmeDB daily sync: {amt_note}')

    batch = {'period_date': date, 'rows': [_explode(r) for r in rows]}
    body, status = _push_batch('/api/collection/sync', batch)
    # Runs on EVERY successful push, not only when rows were skipped.
    #
    # `skipped` counts employees GrowwithmeDB has never heard of. It does NOT
    # count an officer who EXISTS but has no employee_assignment — their rows
    # land fine, so skipped stays 0 — and that officer is exactly the one who
    # vanishes from every region/branch drill while still counting toward the
    # summary card. Gating this on `skipped` meant the one case that needed
    # repairing was the one case that never triggered it.
    if body.get('success'):
        created, repaired, warn = _onboard_missing(rows, roster)
        if created:
            # Only NEW employees need a re-push: their rows were skipped the
            # first time. A repaired officer's rows already landed.
            body, status = _push_batch('/api/collection/sync', batch)
            if body.get('success'):
                body['created_employees'] = created
                body['message'] += f' · {created} new employee(s) created from the report'
        if repaired:
            body['repaired_assignments'] = repaired
            body['message'] += (f' · {repaired} existing officer(s) given their branch '
                                f'(they were missing from the region drill)')
        # A branchless new joiner is counted in the headline but cannot be
        # drilled into, so say so on the same line rather than in a log
        # nobody reads.
        if warn:
            body['unassigned_employees'] = warn
            body['message'] += f' · WARNING: {warn}'
    body.setdefault('date', date)
    # Say plainly whether rupees went up. Counts-only used to look like a clean
    # success while the MIS Amount toggle stayed hidden, which is the bug this
    # reports on rather than repeats.
    body['amount_rows'] = amt_filled
    if amt_note:
        body['amount_warning'] = amt_note
    if body.get('success'):
        if amt_filled:
            body['message'] += f' · rupee amounts on {amt_filled} row(s)'
            # A wrong-date fallback still fills rows, so this warning must ride
            # alongside the success line rather than instead of it.
            if amt_note:
                body['message'] += f' · WARNING: {amt_note}'
        elif amt_note:
            body['message'] += f' · NO rupee amounts: {amt_note}'
        logger.info(f"GrowwithmeDB daily sync: {body['inserted']} rows for {date}")

    # ── Mode of Collection (second, optional workbook) ────────────────────
    # Independent of the collection push above: a channel-file problem must not
    # fail an otherwise-good daily sync, so it reports alongside rather than
    # replacing the result.
    # Preferred path: the browser already aggregated the workbook to officer×channel
    # totals and sent just those (~440 KB of JSON instead of a 5 MB file). Falling
    # back to parsing a raw `channel_file` upload keeps the API usable directly.
    ch_rows = None
    # Other dates carried by the same channel workbook ({date: rows}). Each is
    # re-pushed after the main date, so late-posted receipts for earlier days are
    # picked up automatically (the API push is a whole-date replace — safe).
    ch_extra = {}
    raw_rows = (request.form.get('channel_rows') or '').strip()
    if raw_rows:
        try:
            ch_rows = json.loads(raw_rows)
            if not isinstance(ch_rows, list) or not ch_rows:
                raise ValueError('channel_rows must be a non-empty array')
        except Exception as e:
            body['channels_ok'] = False
            body['message'] = f"{body.get('message', '')} · Mode of Collection FAILED: bad channel_rows ({e})".strip(' ·')
            return jsonify(body), status
        logger.info(f'GrowwithmeDB channel sync: {len(ch_rows)} pre-aggregated rows for {date}')
        raw_extra = (request.form.get('channel_rows_extra') or '').strip()
        if raw_extra:
            try:
                parsed = json.loads(raw_extra)
                if isinstance(parsed, dict):
                    ch_extra = {
                        str(d): r for d, r in parsed.items()
                        if re.match(r'^\d{4}-\d{2}-\d{2}$', str(d)) and str(d) != date
                        and isinstance(r, list) and r
                    }
            except Exception as e:
                # Best-effort: a bad extras blob must not fail the day's sync.
                logger.warning(f'GrowwithmeDB channel sync: channel_rows_extra ignored ({e})')

    if ch_rows is None:
        try:
            ch_path, ch_temp = _uploaded_named('channel_file')
        except ValueError as e:
            body['channels_ok'] = False
            body['message'] = f"{body.get('message', '')} · {e}".strip(' ·')
            return jsonify(body), status
        if ch_path:
            try:
                by_date = _parse_collection_channels_by_date(ch_path, date)
                ch_rows = by_date.pop(date)
                ch_extra = by_date
            except Exception as e:
                logger.warning(f'GrowwithmeDB channel sync: parse failed: {e}')
                body['channels_ok'] = False
                body['message'] = f"{body.get('message', '')} · Mode of Collection FAILED: {e}".strip(' ·')
                _cleanup(ch_path, ch_temp)
                return jsonify(body), status
            finally:
                _cleanup(ch_path, ch_temp)

    if ch_rows:

        # Via _push_batch so a big channel file (thousands of officer×channel rows)
        # is split rather than rejected by the proxy with a 413.
        res_c, _st = _push_batch('/api/collection/sync-channels', {'period_date': date, 'rows': ch_rows})
        ok_c = res_c.get('success')
        if ok_c:
            unresolved = int(res_c.get('unresolved_officers') or 0)
            body['channels_ok'] = True
            body['channel_rows'] = int(res_c.get('inserted') or 0)
            body['channel_unresolved_officers'] = unresolved
            extra = f' ({unresolved} officer(s) unmapped)' if unresolved else ''
            body['message'] = f"{body.get('message', '')} · Mode of Collection: {res_c.get('inserted', 0)} rows{extra}".strip(' ·')
            logger.info(f'GrowwithmeDB channel sync: {res_c.get("inserted")} rows for {date}, {unresolved} unmapped')
        else:
            body['channels_ok'] = False
            why = res_c.get('message') or res_c
            body['message'] = f"{body.get('message', '')} · Mode of Collection push FAILED: {why}".strip(' ·')
            logger.warning(f'GrowwithmeDB channel sync failed: {why}')

        # ── Refresh the OTHER dates the channel workbook carries ──────────
        # Only after the main date succeeded (a failing API would fail them all),
        # and strictly best-effort: a bad extra date is reported, never fatal.
        if ok_c and ch_extra:
            refreshed, failed = [], []
            for d in sorted(ch_extra):
                res_e, _st_e = _push_batch('/api/collection/sync-channels',
                                           {'period_date': d, 'rows': ch_extra[d]})
                if res_e.get('success'):
                    refreshed.append(f"{d} ({int(res_e.get('inserted') or 0)})")
                else:
                    failed.append(d)
                    logger.warning(f'GrowwithmeDB channel refresh {d} failed: {res_e.get("message")}')
            if refreshed:
                body['channel_refreshed_dates'] = refreshed
                body['message'] = (f"{body.get('message', '')} · Mode of Collection also refreshed: "
                                   f"{', '.join(refreshed)}").strip(' ·')
                logger.info(f'GrowwithmeDB channel refresh: {len(refreshed)} extra date(s) pushed')
            if failed:
                body['message'] = (f"{body.get('message', '')} · channel refresh FAILED for "
                                   f"{', '.join(failed)} — re-run this sync to retry").strip(' ·')

    return jsonify(body), status


@growwithme_bp.route('/sync-hourly', methods=['POST'])
def sync_hourly():
    """Push the latest Quick Report into GrowwithmeDB (collection grain 1).

    Body (JSON or multipart): optional `date` (YYYY-MM-DD) + `period_hour` (0-23),
    plus an optional `file` (a Quick Report .xlsx). With a file, that file is
    parsed; without one, the latest generated Quick Report is used. The Quick
    Report has no hour column, so the hour defaults to the current local hour.
    """
    date = _param('date') or datetime.now().strftime('%Y-%m-%d')
    raw_hour = _param('period_hour')
    period_hour = int(raw_hour) if raw_hour else datetime.now().hour

    try:
        up, is_temp = _uploaded_file()
    except ValueError as e:
        return jsonify({'success': False, 'message': str(e)}), 400
    path = up or _quick_report_path()
    if not path:
        return jsonify({'success': False,
                        'message': 'No hourly report found. Run Hourly/Quick processing first, or upload one.'}), 404
    try:
        rows = _parse_quick_report(path)
    except Exception as e:
        logger.warning(f'GrowwithmeDB hourly sync: report parse failed: {e}')
        return jsonify({'success': False, 'message': f'Quick Report parse failed: {e}'}), 500
    finally:
        _cleanup(up, is_temp)
    if not rows:
        return jsonify({'success': False, 'message': 'No employee rows found in Quick Report.'}), 400

    batch = {'rows': [_explode(r, period_date=date, period_hour=period_hour) for r in rows]}
    body, status = _push_batch('/api/hourly/sync', batch)
    body.update(date=date, period_hour=period_hour)
    if body.get('success'):
        logger.info(f"GrowwithmeDB hourly sync: {body['inserted']} rows for {date} h{period_hour}")
    return jsonify(body), status


# ── Portfolio (POS) sync ──────────────────────────────────────────────
# Portfolio is NOT in the daily EOD report — it comes from the Month-End
# Employee Report's `POS` sheet (branch+product PrincipalOS, computed from PAR).
# That sheet's grain matches GrowwithmeDB's portfolio_period (branch+product+month)
# exactly, so we read it directly and push one row per (branch, product).
#
# The growwithme-local /api/portfolio/sync endpoint does a whole-month override
# (delete the month, then insert), so every call replaces the month.

# Product (sheet/Product Name) -> GrowwithmeDB product_type_id. VVY == IL.
_PORTFOLIO_PT_ID = {'IGL': 1, 'FIG': 2, 'IL': 3, 'VVY': 3}

# Three-letter month label -> month number, for {"month":"MAR","year":2026}.
_MONTH_NUM = {m: f'{i:02d}' for i, m in enumerate(
    ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC'], start=1)}

# POS column header -> growwithme pos bucket key (sent to /api/portfolio/sync).
_POS_HEADER_KEY = [
    ('regular_pos', 'regular'),
    ('sma0_pos', 'sma0'),
    ('sma1_pos', 'sma1'),
    ('pnpa_pos', 'pnpa'),
    ('npa_pos', 'npa'),
    ('total_pos', 'total'),
]

# Raw-PAR fallback: map the PAR's "DPD Days" bucket text -> growwithme pos key,
# mirroring the Month-End engine's POS derivation (eod_processor.build_employee_report).
# Keys are matched lower-cased/stripped. 61-90 = PNPA; everything 91+ = NPA.
_PAR_DPD_POS = {
    '0 days': 'regular', '0days': 'regular', '0': 'regular',
    '1: 1-30': 'sma0', '1-30': 'sma0',
    '2: 31-60': 'sma1', '31-60': 'sma1',
    '3: 61-90': 'pnpa', '61-90': 'pnpa',
    '4: 91-120': 'npa', '91-120': 'npa',
    '5: 121-180': 'npa', '121-180': 'npa', '121-150': 'npa', '151-180': 'npa',
    '6: 181-365': 'npa', '181-365': 'npa', '181-210': 'npa', '211-250': 'npa', '251-365': 'npa',
    '7: >365 days': 'npa', '>365 days': 'npa', '>365': 'npa', '>365days': 'npa', '>120': 'npa',
}


def _par_dpd_bucket(v):
    """Resolve a DPD-band cell to regular/sma0/sma1/pnpa/npa, for BOTH workbook
    flavours the Portfolio tab accepts:

      * PAR export ("Par as on <date>.xlsx", Sheet1)  — '1: 1-30', '7: >365 Days', …
      * Active-clients export ("Active clients details as on <date>.xlsx", Data)
        — 'Regular' / '0 Days' / '1-30' / '31-60' / '61-90' / 'NPA'

    Returns None ONLY for a blank cell. Anything non-blank that isn't a
    performing band falls through to 'npa', matching the authoritative parser in
    growwithme-local/tools/client_account_extract.py (bucket_from_dpd), which
    treats every unrecognised overdue label as NPA.

    Why this exists: the old code did a strict `_PAR_DPD_POS.get(...)` and
    `continue`d on a miss. The Active-clients export writes the literal 'NPA',
    which was not a key — so every NPA account was silently DROPPED, giving that
    month a zero NPA figure and an understated Total. Falling through to 'npa'
    instead of skipping is what makes an Active-clients upload correct.
    """
    s = str(v or '').strip().lower()
    if not s:
        return None
    hit = _PAR_DPD_POS.get(s)
    if hit:
        return hit
    if s.startswith('regular'):
        return 'regular'
    if s.startswith('0 day'):
        return 'regular'
    if '1-30' in s or '1 - 30' in s:
        return 'sma0'
    if '31-60' in s or '31 - 60' in s:
        return 'sma1'
    if '61-90' in s or '61 - 90' in s:
        return 'pnpa'
    return 'npa'


def _month_end_report_path():
    """Locate the latest Month-End Employee Report (has the POS sheet).

    Falls back to the EOD Employee Report only if it happens to carry a POS
    sheet (it normally does not).
    """
    for name in ('Quick_Month_End_Employee_Latest.xlsx', 'Employee_Report_Latest.xlsx'):
        p = config.BACKEND_DATA_DIR / name
        if p.exists():
            return p
    return None


def _parse_pos_sheet(path):
    """Parse the report's `POS` sheet into branch+product POS rows.

    Sheet columns: Region, Division, Area, BranchName, Product Name,
    Regular_POS, SMA0_POS, SMA1_POS, PNPA_POS, NPA_POS, Total_POS, and (when the
    report carries it) a Total_Account / No_of_Account column.
    Returns [{branch, product_type_id, pos:{regular,sma0,sma1,pnpa,npa,total},
              acc: <int|None>}, ...]. `acc` is None when the sheet has no
    account-count column (then the Total Account push is skipped).
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if 'POS' not in wb.sheetnames:
            raise ValueError("No 'POS' sheet in the report — generate a Month-End report first.")
        rows = list(wb['POS'].iter_rows(values_only=True))
        if not rows:
            return []
        hdr = {_norm(h): i for i, h in enumerate(rows[0]) if h is not None}
        b_i = hdr.get('branchname')
        p_i = hdr.get('product name')
        if b_i is None or p_i is None:
            raise ValueError("POS sheet missing BranchName / Product Name columns.")
        pos_idx = [(hdr.get(h), key) for h, key in _POS_HEADER_KEY]
        # Account-count column for the "Total Account" card. Optional — match the
        # first header that mentions 'account' but is NOT a *_POS amount column.
        acc_i = next((i for h, i in hdr.items() if 'account' in h and 'pos' not in h), None)
        if acc_i is None:
            logger.info("GrowwithmeDB portfolio sync: no account-count column in POS sheet — "
                        "Total Account push will be skipped.")

        out = []
        for row in rows[1:]:
            if not row or len(row) <= b_i:
                continue
            branch = str(row[b_i]).strip() if row[b_i] is not None else ''
            prod = str(row[p_i]).strip().upper() if (p_i < len(row) and row[p_i] is not None) else ''
            if not branch or not prod:
                continue
            pt = _PORTFOLIO_PT_ID.get(prod)
            if pt is None:
                logger.info(f"GrowwithmeDB portfolio sync: skipping unknown product '{prod}'")
                continue
            pos = {key: (_num(row[i]) if (i is not None and i < len(row)) else 0) for i, key in pos_idx}
            acc = _num(row[acc_i]) if (acc_i is not None and acc_i < len(row)) else None
            out.append({'branch': branch, 'product_type_id': pt, 'pos': pos, 'acc': acc})
        return out
    finally:
        wb.close()


# Demand-bucket account-count columns that sum to the "Total Account" figure —
# matching the live site's derivation: regular_demand + 1-30 + 31-60 + pnpa_demand
# + npa_cases. (db_col, normalised header) — these are COUNT columns (not _amt).
_ACC_FIELDS = [
    ('regular_demand', 'regular demand'),
    ('demand_1_30',    '1-30 demand'),
    ('demand_31_60',   '31-60 demand'),
    ('pnpa_demand',    'pnpa demand'),
    ('npa_cases',      'npa cases'),
]


def _parse_demand_accounts(path):
    """Derive per branch×product ACCOUNT COUNTS for the "Total Account" card the
    same way the live site does: SUM(regular_demand + 1-30 + 31-60 + pnpa_demand +
    npa_cases) of the account-count columns, aggregated from the report's per-product
    EOD sheets (IGL/FIG/VVY) up to branch×product.

    Returns {(BRANCH_UPPER, product_type_id): acc_int}. Empty when the report has no
    such sheets/columns (caller then falls back to a POS-sheet account column).
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    acc = {}
    try:
        for sheet_name in wb.sheetnames:
            if sheet_name.endswith('_FY') or sheet_name in ('POS', 'EMP_POS'):
                continue
            pt = _PORTFOLIO_PT_ID.get(sheet_name.strip().upper())  # IGL/FIG/IL/VVY -> id
            if pt is None:
                continue
            rows = list(wb[sheet_name].iter_rows(values_only=True))
            if not rows:
                continue
            hdr = {}
            for i, h in enumerate(_norm(x) for x in rows[0]):
                if h and h not in hdr:
                    hdr[h] = i
            b_i = next((hdr[c] for c in ('branchname', 'branch name', 'branch') if c in hdr), None)
            col_idx = [hdr.get(txt) for _, txt in _ACC_FIELDS]
            if b_i is None or all(c is None for c in col_idx):
                continue  # not a parseable per-product sheet — skip (fallback handles it)
            for row in rows[1:]:
                if not row or b_i >= len(row) or row[b_i] is None:
                    continue
                branch = str(row[b_i]).strip()
                if not branch:
                    continue
                total = sum(_num(row[ci]) if (ci is not None and ci < len(row)) else 0 for ci in col_idx)
                key = (branch.upper(), pt)
                acc[key] = acc.get(key, 0) + total
        return acc
    finally:
        wb.close()


def _resolve_period_month(data):
    """Resolve period_month ('YYYY-MM-01') from the request body.

    Accepts {"period_month":"YYYY-MM[-DD]"} directly, or {"month":"MAR","year":2026}.
    Returns the normalised string, or None if it can't be resolved.
    """
    pm = (data.get('period_month') or '').strip()
    if pm:
        import re
        mo = re.match(r'^(\d{4})-(\d{2})', pm)
        if mo:
            return f'{mo.group(1)}-{mo.group(2)}-01'
    label = (data.get('month') or '').strip().upper()[:3]
    year = data.get('year')
    if label in _MONTH_NUM and year:
        return f'{int(year)}-{_MONTH_NUM[label]}-01'
    return None


def _future_month(period_month):
    """True if 'YYYY-MM-01' is later than the month we're currently in.

    A portfolio snapshot can't exist for a month that hasn't started. The sync is
    an upsert keyed on (branch, product, month) and never questions the month it
    is handed, so a mistyped year ('2027-07') would happily create a phantom month
    that then sorts to the top of /portfolio/months and becomes the MIS default.
    """
    import datetime
    try:
        y, m = int(period_month[0:4]), int(period_month[5:7])
    except (TypeError, ValueError, IndexError):
        return False
    today = datetime.date.today()
    return (y, m) > (today.year, today.month)


def _has_pos_sheet(path):
    """True if the workbook carries a 'POS' sheet (a generated Month-End report).
    False for a raw PAR (Sheet1/Sheet2) or on any read error."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            return 'POS' in wb.sheetnames
        finally:
            wb.close()
    except Exception:
        return False


def _par_product_type(prod, pid):
    """Resolve a PAR row's product to a GrowwithmeDB product_type_id (1 IGL / 2 FIG /
    3 IL). 'IGL & FIG' is split via ProductID (starts '6' or contains FIG -> FIG)."""
    p = (prod or '').strip().upper()
    if p == 'IGL & FIG':
        s = str(pid or '').upper()
        p = 'FIG' if (s.startswith('6') or 'FIG' in s) else 'IGL'
    if p not in ('IGL', 'FIG'):
        p = 'IL'
    return _PORTFOLIO_PT_ID.get(p)


# ── Client detail (loan-account grain) ────────────────────────────────────
# The Portfolio sync stores branch x product POS aggregates (323 rows from a
# 263k-row PAR). The "Customer details" panel reads a different pair of tables —
# client_snapshot / client_account — which nothing in the UI populated, so that
# panel stayed empty no matter how often the PAR was uploaded. This parser emits
# the account-level rows so the same upload can fill it.
#
# Mirrors growwithme-local/tools/client_account_extract.py exactly: same bucket
# vocabulary, same product split, same field names.

_CA_FIG_IDS = {'60401', '60201', 'MFIG'}
_CA_DPD_LABEL = {'regular': '0 Days', 'sma0': '1-30', 'sma1': '31-60', 'sma2': '61-90', 'npa': 'NPA'}
_CA_PRODUCT_NAME = {1: 'IGL', 2: 'FIG', 3: 'VVY'}


def _ca_bucket_from_aging(v):
    """PAR AgingofOverdue -> bucket. Anything past 61-90 is NPA."""
    s = str(v or '').strip().lower()
    if s.startswith('regular'):
        return 'regular'
    if s.startswith('1 - 30') or s.startswith('1-30'):
        return 'sma0'
    if s.startswith('31 - 60') or s.startswith('31-60'):
        return 'sma1'
    if s.startswith('61 - 90') or s.startswith('61-90'):
        return 'sma2'
    return 'npa'


def _ca_bucket_from_dpd(v):
    """Active-clients 'DPD Days' -> bucket (also tolerates the PAR encoding)."""
    s = str(v or '').strip().lower()
    if s.startswith('0 day') or s == '0':
        return 'regular'
    if s == 'npa':
        return 'npa'
    if '1-30' in s or '1 - 30' in s:
        return 'sma0'
    if '31-60' in s or '31 - 60' in s:
        return 'sma1'
    if '61-90' in s or '61 - 90' in s:
        return 'sma2'
    return 'npa'


def _ca_product_type_id(pid, pname):
    """1=IGL, 2=FIG, 3=IL/VVY. Handles both workbook flavours."""
    p = str(pname or '').strip().upper()
    if p == 'VVY':
        return 3
    if p == 'FIG':
        return 2
    if p == 'IGL':
        return 1
    return 2 if str(pid or '').strip() in _CA_FIG_IDS else 1


def _ca_date(v):
    import datetime
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime('%Y-%m-%d')
    s = str(v or '').strip()
    return s[:10] if len(s) >= 10 and s[4:5] == '-' else None


def _parse_client_accounts(path, chunk=1000):
    """Stream a PAR / Active-clients workbook and YIELD lists of account rows.

    Yields in chunks rather than returning one list: a PAR is ~263k rows and
    holding them all in memory (then serialising them at once) is what makes this
    fall over. Each chunk is small enough to POST on its own.

    CHUNK SIZE IS A BODY-SIZE LIMIT, NOT A ROW LIMIT. Measured on the real
    31-07-2026 PAR, one account row serialises to ~661 bytes, so the old default
    of 4000 produced a 2.52 MB request — over nginx's 1 MB `client_max_body_size`
    default, which is exactly the 413 the Portfolio sync returned. 1000 rows is
    ~0.65 MB, leaving headroom under that default. _push_client_accounts also
    halves a batch and retries on a 413, so a proxy stricter still than 1 MB
    degrades to more requests rather than a failed sync.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        names = wb.sheetnames
        sheet = 'Sheet1' if 'Sheet1' in names else ('Data' if 'Data' in names else names[0])
        ws = wb[sheet]
        it = ws.iter_rows(values_only=True)
        header = [str(c).strip() if c is not None else '' for c in next(it)]
        H = {h: i for i, h in enumerate(header)}
        if 'PrincipalOS' not in H:
            raise ValueError('No PrincipalOS column — is this a PAR / Active-clients file?')
        is_par = 'AgingofOverdue' in H

        def col(*names_):
            for n in names_:
                if n in H:
                    return H[n]
            return None

        i = {k: col(*v) for k, v in {
            'obid': ('OurBranchID',), 'reg': ('Region',), 'div': ('Division', 'Division '),
            'area': ('Area',), 'br': ('BranchName',), 'oid': ('OfficerID',), 'onm': ('OfficerName',),
            'gid': ('GroupID',), 'gnm': ('GroupName',), 'cid': ('ClientID',), 'cnm': ('Client Name',),
            'mob': ('Mobile No',), 'acc': ('AccountID',), 'pn': ('Product Name',), 'pid': ('ProductID',),
            'dd': ('DisbursementDate',), 'la': ('LoanAmount',), 'ia': ('InstallmentAmount',),
            'arr': ('TotalArrear',), 'due': ('DueDays',), 'dpd': ('DPD Days',),
            'pos': ('PrincipalOS',), 'mat': ('LoanMaturityDate',), 'ag': ('AgingofOverdue',),
        }.items()}

        def g(row, key):
            j = i.get(key)
            return row[j] if j is not None and j < len(row) else None

        batch = []
        for row in it:
            if g(row, 'br') is None and g(row, 'acc') is None:
                continue
            bucket = _ca_bucket_from_aging(g(row, 'ag')) if is_par else _ca_bucket_from_dpd(g(row, 'dpd'))
            ptid = _ca_product_type_id(g(row, 'pid'), g(row, 'pn'))
            batch.append({
                'our_branch_id': g(row, 'obid'), 'region': g(row, 'reg'), 'division': g(row, 'div'),
                'area': g(row, 'area'), 'branch_name': g(row, 'br'),
                'officer_code': g(row, 'oid'), 'officer_name': g(row, 'onm'),
                'group_code': g(row, 'gid'), 'group_name': g(row, 'gnm'),
                'client_code': g(row, 'cid'), 'client_name': g(row, 'cnm'), 'mobile_no': g(row, 'mob'),
                'account_no': g(row, 'acc'), 'product_name': _CA_PRODUCT_NAME.get(ptid), 'product_type_id': ptid,
                'disbursement_date': _ca_date(g(row, 'dd')), 'loan_amount': _num(g(row, 'la')),
                'installment_amount': _num(g(row, 'ia')), 'total_arrear': _num(g(row, 'arr')),
                'due_days': _num(g(row, 'due')), 'dpd_days': _CA_DPD_LABEL.get(bucket),
                'bucket': bucket, 'principal_os': _num(g(row, 'pos')),
                'loan_maturity_date': _ca_date(g(row, 'mat')),
            })
            if len(batch) >= chunk:
                yield batch
                batch = []
        if batch:
            yield batch
    finally:
        wb.close()


def _push_client_accounts(path, as_on_date, period_month, source_file):
    """Push every account row from `path` to /api/clients/sync, chunk by chunk.

    Returns (ok, message). The FIRST request carries replace:true so the snapshot
    is cleared exactly once; every later one appends. Failing mid-way leaves a
    PARTIAL snapshot, so the message says so rather than implying success.

    On a 413 the batch is halved and retried, because the body limit belongs to
    whatever proxy sits in front of the API and is not knowable from here. That
    keeps the sync working on a box whose nginx still has the 1 MB default.
    """
    kind = 'par'
    state = {'sent': 0, 'total': 0}

    def push(batch, depth=0):
        """POST one batch, splitting on 413. Returns (ok, err_or_None)."""
        ok, res = _post('/api/clients/sync', {
            'as_on_date': as_on_date, 'period_month': period_month,
            'source_file': source_file, 'source_kind': kind,
            # replace ONLY on the very first request that leaves this process —
            # not per chunk, or a split batch would wipe what it just wrote.
            'replace': state['sent'] == 0, 'rows': batch,
        })
        if ok:
            state['sent'] += 1
            state['total'] += len(batch)
            if state['sent'] % 10 == 0:
                logger.info(f"GrowwithmeDB client detail: {state['total']:,} rows pushed…")
            return True, None
        # 413 = body too large for the proxy. Halve and retry; 25 rows (~16 KB)
        # is small enough that a further 413 means something other than size.
        if '(413)' in str(res) and len(batch) > 25 and depth < 8:
            half = len(batch) // 2
            logger.warning(f'GrowwithmeDB client detail: 413 on {len(batch)} rows — '
                           f'retrying as {half} + {len(batch) - half}')
            for part in (batch[:half], batch[half:]):
                pok, perr = push(part, depth + 1)
                if not pok:
                    return False, perr
            return True, None
        return False, str(res)

    try:
        for batch in _parse_client_accounts(path):
            ok, err = push(batch)
            if not ok:
                partial = (f" — {state['total']:,} row(s) already written, snapshot is INCOMPLETE"
                           if state['total'] else '')
                hint = ''
                if '(413)' in str(err):
                    hint = (' · still 413 at the smallest batch — raise '
                            'client_max_body_size in nginx on the API box')
                return False, f"client detail push failed after {state['total']:,} rows: {err}{partial}{hint}"
    except Exception as e:
        return False, f'client detail parse failed: {e}'
    logger.info(f"GrowwithmeDB client detail: {state['total']:,} account rows for {as_on_date}")
    return True, f"{state['total']:,} client account rows"


def _par_as_on_date(path, period_month):
    """As-on date for a PAR: prefer the DD-MM-YYYY in its filename, else the last
    day of period_month. A PAR is a month-end snapshot, so the month's final day
    is the right fallback — not today."""
    import re as _re, calendar, datetime
    m = _re.search(r'(\d{2})-(\d{2})-(\d{4})', os.path.basename(str(path) or ''))
    if m:
        dd, mm, yyyy = m.groups()
        if 1 <= int(mm) <= 12 and 1 <= int(dd) <= 31:
            return f'{yyyy}-{mm}-{dd}'
    y, mo = int(period_month[0:4]), int(period_month[5:7])
    return f'{y:04d}-{mo:02d}-{calendar.monthrange(y, mo)[1]:02d}'


def _parse_par_pos(path):
    """Build branch×product POS rows straight from a RAW PAR file, so the Portfolio
    tab can ingest a PAR without first generating a Month-End report.

    Mirrors the Month-End engine (eod_processor.build_employee_report): bucket each
    account's PrincipalOS by its DPD Days into regular/sma0/sma1/pnpa/npa, then sum
    per (BranchName, product). Returns the SAME shape as _parse_pos_sheet, plus
    per-bucket account counts:
      [{branch, product_type_id, pos:{regular,sma0,sma1,pnpa,npa,total},
        acc, acc_buckets:{regular,sma0,sma1,pnpa,npa}}, ...]
    where `acc` is the total account (row) count per branch×product (Total Account
    card) and `acc_buckets` are the per-DPD-bucket counts (Active Accounts card =
    non-NPA buckets). `acc` == sum(acc_buckets).

    Streams the sheet row-by-row with openpyxl read_only — a raw PAR can be 100 MB+
    with 700k rows, so we never load it into a DataFrame.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        # Prefer the account-level detail sheet ('Sheet1'); else the first sheet
        # whose header carries PrincipalOS.
        sheet = None
        for name in (['Sheet1'] + [s for s in wb.sheetnames if s != 'Sheet1']):
            ws = wb[name]
            hdr = next(ws.iter_rows(values_only=True), None)
            if hdr and any(str(c).strip() == 'PrincipalOS' for c in hdr if c is not None):
                sheet = name
                break
        if sheet is None:
            raise ValueError('PAR has no sheet with a PrincipalOS column.')

        ws = wb[sheet]
        it = ws.iter_rows(values_only=True)
        header = [str(c).strip() if c is not None else '' for c in next(it)]
        idx = {h: i for i, h in enumerate(header)}
        dpd_col = next((c for c in ('DPD Days', 'Days Group', 'Days group', 'DaysGroup') if c in idx), None)
        if dpd_col is None or 'PrincipalOS' not in idx or 'BranchName' not in idx:
            raise ValueError('PAR missing required columns (need BranchName, DPD Days, PrincipalOS).')
        i_branch, i_dpd, i_pos = idx['BranchName'], idx[dpd_col], idx['PrincipalOS']
        i_prod = idx.get('Product Name')
        i_pid = idx.get('ProductID')

        # (branch, product_type_id) -> {bucket: pos_sum, ..., '_acc': count}
        agg = {}
        # Dropped-row tallies. Silently skipping rows is how a month ends up with a
        # zero NPA figure and nobody notices, so they're logged after the sweep.
        skip_blank_dpd = skip_product = 0
        for row in it:
            if not row or i_branch >= len(row):
                continue
            branch = str(row[i_branch]).strip() if row[i_branch] is not None else ''
            if not branch:
                continue
            bucket = _par_dpd_bucket(row[i_dpd] if i_dpd < len(row) else None)
            if not bucket:
                skip_blank_dpd += 1
                continue
            prod = row[i_prod] if (i_prod is not None and i_prod < len(row)) else ''
            pid = row[i_pid] if (i_pid is not None and i_pid < len(row)) else ''
            pt = _par_product_type(prod, pid)
            if pt is None:
                skip_product += 1
                continue
            pos_val = _num(row[i_pos]) if (i_pos < len(row)) else 0
            key = (branch, pt)
            slot = agg.get(key)
            if slot is None:
                slot = agg[key] = {
                    'pos': {'regular': 0.0, 'sma0': 0.0, 'sma1': 0.0, 'pnpa': 0.0, 'npa': 0.0},
                    'acc': {'regular': 0, 'sma0': 0, 'sma1': 0, 'pnpa': 0, 'npa': 0},
                }
            slot['pos'][bucket] += pos_val
            slot['acc'][bucket] += 1

        if skip_blank_dpd or skip_product:
            logger.info(f'GrowwithmeDB portfolio sync: skipped {skip_blank_dpd} row(s) with a blank DPD band '
                        f'and {skip_product} with an unresolvable product.')

        out = []
        for (branch, pt), slot in agg.items():
            pos = slot['pos']
            pos['total'] = pos['regular'] + pos['sma0'] + pos['sma1'] + pos['pnpa'] + pos['npa']
            acc_buckets = slot['acc']
            out.append({
                'branch': branch, 'product_type_id': int(pt),
                'pos': pos,
                'acc': int(sum(acc_buckets.values())),
                'acc_buckets': acc_buckets,
            })
        return out
    finally:
        wb.close()


@growwithme_bp.route('/sync-portfolio', methods=['POST'])
def sync_portfolio():
    """Push portfolio POS into GrowwithmeDB.portfolio_* (monthly).

    Body (JSON or multipart): {"period_month":"YYYY-MM-01"} (or month+year), plus
    an optional `file` (.xlsx). The file may be EITHER a Month-End report (its POS
    sheet is read) OR a raw PAR (POS is built here from PrincipalOS × DPD, same as
    the Month-End engine). Without a file, the latest generated Month-End report is
    used. Whole-month override — re-running for the same month replaces it.
    """
    period_month = _resolve_period_month({
        'period_month': _param('period_month'), 'month': _param('month'), 'year': _param('year'),
    })
    if not period_month:
        return jsonify({'success': False,
                        'message': 'period_month required (YYYY-MM-01), or month+year (e.g. "MAR" + 2026).'}), 400
    if _future_month(period_month):
        return jsonify({'success': False,
                        'message': f'{period_month[:7]} is in the future — a portfolio snapshot cannot exist for a '
                                   f'month that has not started. Check the selected month.'}), 400

    try:
        up, is_temp = _uploaded_file()
    except ValueError as e:
        return jsonify({'success': False, 'message': str(e)}), 400
    path = up or _month_end_report_path()
    if not path:
        return jsonify({'success': False,
                        'message': 'No Month-End Employee Report found. Generate one first, or upload one.'}), 404
    try:
        if _has_pos_sheet(path):
            # Generated Month-End report — read its POS sheet.
            rows = _parse_pos_sheet(path)
            # Derive Total Account from the demand-bucket counts (matches the live site).
            # Best-effort — a derivation failure just falls back to a POS-sheet column.
            try:
                acc_map = _parse_demand_accounts(path)
            except Exception as e:
                logger.warning(f'GrowwithmeDB portfolio sync: account derivation failed: {e}')
                acc_map = {}
        else:
            # Raw PAR upload — build the POS (PrincipalOS × DPD) directly. Account
            # counts come from the PAR itself (per branch×product), so no acc_map.
            logger.info('GrowwithmeDB portfolio sync: no POS sheet — treating upload as a raw PAR.')
            rows = _parse_par_pos(path)
            acc_map = {}
        is_raw_par = not _has_pos_sheet(path)
    except Exception as e:
        logger.warning(f'GrowwithmeDB portfolio sync: POS parse failed: {e}')
        _cleanup(up, is_temp)
        return jsonify({'success': False, 'message': f'POS parse failed: {e}'}), 500
    # NOTE: the temp file is deliberately NOT cleaned up here — the client-detail
    # push below re-reads it. It is removed in the `finally` at the end instead.
    if not rows:
        _cleanup(up, is_temp)
        return jsonify({'success': False, 'message': 'No POS rows found in the report.'}), 400

    batch = {'period_month': period_month, 'rows': rows}
    body, status = _push_batch('/api/portfolio/sync', batch)
    body.update(period_month=period_month, branch_products=len(rows))
    if not body.get('success'):
        return jsonify(body), status
    logger.info(f"GrowwithmeDB portfolio sync: {body['inserted']} branch×product rows for {period_month}")

    # Total Account counts (pos_status 'total_acc'). Prefer the live-style derivation
    # (sum of demand-bucket account counts per branch×product); fall back to a POS-
    # sheet account column when the report lacks the per-product demand sheets.
    # Best-effort, runs after the POS push succeeds.
    acc_rows = []
    for r in rows:
        acc = acc_map.get((str(r['branch']).strip().upper(), r['product_type_id']))
        if acc is None:
            acc = r.get('acc')  # POS-sheet column / PAR-derived fallback
        if acc is not None:
            entry = {'branch': r['branch'], 'product_type_id': r['product_type_id'], 'acc': int(round(acc))}
            # PAR-derived rows also carry per-DPD-bucket counts — drives the
            # "Active Accounts" card (pos_status 8-12). Absent for POS-sheet syncs.
            if r.get('acc_buckets'):
                entry['acc_buckets'] = {k: int(v) for k, v in r['acc_buckets'].items()}
            acc_rows.append(entry)
    if acc_rows:
        ok_a, res_a = _post('/api/portfolio/sync-accounts', {'period_month': period_month, 'rows': acc_rows})
        if ok_a:
            matched = int((res_a or {}).get('matched') or 0)
            bucket_rows = int((res_a or {}).get('bucketRows') or 0)
            body['accounts_matched'] = matched
            body['active_account_rows'] = bucket_rows
            extra = f' (+ Active Accounts for {bucket_rows})' if bucket_rows else ''
            body['message'] = f"{body.get('message', '')} · {matched} Total Account rows{extra}".strip(' ·')
            logger.info(f"GrowwithmeDB portfolio accounts sync: {matched} matched, {bucket_rows} with buckets for {period_month}")
        else:
            body['accounts_ok'] = False
            body['message'] = f"{body.get('message', '')} · Total Account push FAILED: {res_a}".strip()
            logger.warning(f'GrowwithmeDB portfolio accounts sync failed: {res_a}')

    # ── Client detail (loan-account grain) ────────────────────────────────
    # Only a RAW PAR carries per-account rows; a generated Month-End report has
    # POS totals only, nothing to load. Runs after the POS push so a client-detail
    # problem can never fail an otherwise-good portfolio sync.
    if is_raw_par:
        as_on = _par_as_on_date(path, period_month)
        ok_c, msg_c = _push_client_accounts(path, as_on, period_month, os.path.basename(str(path)))
        body['client_detail_ok'] = bool(ok_c)
        body['client_detail_as_on'] = as_on
        if ok_c:
            body['message'] = f"{body.get('message', '')} · {msg_c}".strip(' ·')
            # ── Overview board figures (FTOD - PAR Addition / Total 1+ PAR /
            #    Disbursement) ────────────────────────────────────────────────
            # The snapshot just pushed is the same account-level DPD data the
            # offline build_overview_metrics.py reads from the raw workbooks, so
            # the API can now derive the Overview's PAR-derived rows itself,
            # against the PRIOR month's snapshot as the 0-DPD baseline. This is
            # what makes the Overview's "FTOD - PAR Addition" cell fill in from
            # this one button instead of a separate script + curl run. Best-
            # effort: a failure here never fails an otherwise-good portfolio sync.
            ok_o, res_o = _post('/api/overview/compute', {'period_month': period_month})
            if ok_o and (res_o or {}).get('ok'):
                body['overview_ok'] = True
                ftod = res_o.get('ftodParAddition')
                ftod_txt = f"FTOD - PAR Addition {ftod} Cr" if ftod is not None else \
                    "FTOD skipped (no prior-month snapshot as baseline)"
                body['message'] = (f"{body.get('message', '')} · Overview updated: {ftod_txt}, "
                                   f"Total 1+ PAR {res_o.get('total1PlusPar')} Cr").strip(' ·')
                logger.info(f"GrowwithmeDB overview compute: {res_o.get('message')}")
            else:
                body['overview_ok'] = False
                detail = (res_o or {}).get('message') or (res_o if isinstance(res_o, str) else res_o)
                body['message'] = (f"{body.get('message', '')} · Overview board figures NOT updated: "
                                   f"{detail} (an older API build without /api/overview/compute "
                                   f"answers 404 — deploy api/routes/overview.js, api/lib/overview.js "
                                   f"and api/index.js, then re-run this sync)").strip(' ·')
                logger.warning(f'GrowwithmeDB overview compute failed: {detail}')
        else:
            body['message'] = f"{body.get('message', '')} · Customer details FAILED: {msg_c}".strip(' ·')
            logger.warning(f'GrowwithmeDB client detail: {msg_c}')
    else:
        body['client_detail_ok'] = None
        body['message'] = (f"{body.get('message', '')} · Customer details skipped "
                           f"(Month-End report has no per-account rows)").strip(' ·')

    _cleanup(up, is_temp)
    return jsonify(body), status


# ── Staff (HR master) sync — refresh employee DETAIL fields (name, phone, joining
#    date, DOB, reporting manager) into GrowwithmeDB. Mirrors the Coll_Db staff
#    upload format: a "Working" sheet where row 0 = column numbers, row 1 = headers,
#    row 2+ = data. DETAILS-ONLY on the API side (never touches branch/role/hierarchy).
_STAFF_COLS = {
    'emp_id':               ['nmempid', 'emp id', 'emp_id', 'empid'],
    'full_name':            ['name(asperaadhar)', 'name', 'as per aadhaar'],
    'mobile':               ['personalmobile', 'personal mobile', 'mobile'],
    'date_of_joining':      ['date of joining', 'doj', 'joining'],
    'date_of_birth':        ['date of birth', 'dob'],
    'reporting_officer_id': ['reportingofficerempid', 'reporting officer emp'],
}


def _excel_date(v):
    """Normalise an Excel cell to 'YYYY-MM-DD' (or None). Handles datetimes, Excel
    serial numbers, and common date strings."""
    if v is None or v == '':
        return None
    if hasattr(v, 'strftime'):
        try:
            return v.strftime('%Y-%m-%d')
        except Exception:
            return None
    s = str(v).strip()
    if not s:
        return None
    try:
        n = float(s)
        if n > 1000:  # Excel serial date
            from datetime import datetime as _dt, timedelta
            return (_dt(1899, 12, 30) + timedelta(days=n)).strftime('%Y-%m-%d')
    except (TypeError, ValueError):
        pass
    from datetime import datetime as _dt
    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%d-%b-%Y'):
        try:
            return _dt.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return None


def _parse_staff_sheet(path):
    """Parse the staff master's 'Working' sheet into detail rows for /sync-staff.
    Returns [{emp_id, full_name, mobile, date_of_joining, date_of_birth,
    reporting_officer_id}, ...]. Header row is row index 1 (row 0 = column numbers)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = next((s for s in wb.sheetnames if 'working' in s.lower()), wb.sheetnames[0])
        rows = list(wb[sheet].iter_rows(values_only=True))
        if len(rows) < 3:
            return []
        headers = [_norm(h) for h in rows[1]]

        def findcol(keys):
            for k in keys:
                for i, h in enumerate(headers):
                    if h and k in h:
                        return i
            return None

        idx = {field: findcol(keys) for field, keys in _STAFF_COLS.items()}
        if idx['emp_id'] is None:
            raise ValueError("No employee-id column (NMEmpId / EMP ID) in the Working sheet.")

        def cell(row, field):
            i = idx[field]
            if i is None or i >= len(row) or row[i] is None:
                return None
            return row[i]

        out, seen = [], set()
        for row in rows[2:]:
            if not row:
                continue
            raw = cell(row, 'emp_id')
            code = str(raw).strip() if raw is not None else ''
            if not code or code in seen:
                continue
            seen.add(code)
            txt = lambda f: (str(cell(row, f)).strip() if cell(row, f) is not None else None)
            out.append({
                'emp_id': code,
                'full_name': txt('full_name'),
                'mobile': txt('mobile'),
                'date_of_joining': _excel_date(cell(row, 'date_of_joining')),
                'date_of_birth': _excel_date(cell(row, 'date_of_birth')),
                'reporting_officer_id': txt('reporting_officer_id'),
            })
        return out
    finally:
        wb.close()


# ── Onboarding / access sheet ─────────────────────────────────────────────
# The 'Employee_Onboarding_Template.xlsx' / 'Employee_Access_Current.xlsx' sheet
# ('Employees', headers on the FIRST row) carries a `scope` column and optional
# branch/role/designation. Unlike the HR 'Working' master (details-only), this
# drives the FULL onboard on the API side (create employee + login/password +
# info + assignment + SCOPE). Detected by the presence of a `scope` column.
_ONBOARD_FIELDS = ['emp_id', 'name', 'mobile', 'branch', 'role', 'designation',
                   'scope', 'reporting_officer_id', 'date_of_joining', 'date_of_birth', 'gender']


def _parse_onboard_sheet(path):
    """If `path` is an onboarding/access sheet (has emp_id + scope columns), return
    its rows as onboard dicts. Otherwise return None (caller falls back to the HR
    'Working' staff parser). Header row is auto-detected in the first 3 rows."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        names = wb.sheetnames
        sheet = next((s for s in names if s.strip().lower() == 'employees'), names[0])
        rows = list(wb[sheet].iter_rows(values_only=True))
        low = lambda x: str(x).strip().lower() if x is not None else ''
        hi, colmap = None, None
        for r in range(min(3, len(rows))):
            hdr = [low(c) for c in rows[r]]
            cm = {}
            for f in _ONBOARD_FIELDS:
                alt = f.replace('_', ' ')
                for i, h in enumerate(hdr):
                    if h == f or h == alt:
                        cm[f] = i
                        break
            if 'emp_id' in cm and 'scope' in cm:
                hi, colmap = r, cm
                break
        if colmap is None:
            return None  # not an onboarding sheet

        def cell(row, f):
            i = colmap.get(f)
            if i is None or i >= len(row):
                return None
            return row[i]

        out, seen = [], set()
        for row in rows[hi + 1:]:
            if not row:
                continue
            raw = cell(row, 'emp_id')
            code = str(raw).strip() if raw is not None else ''
            if not code or code.lower() == 'emp_id' or code in seen:
                continue
            seen.add(code)
            txt = lambda f: (str(cell(row, f)).strip() if cell(row, f) is not None else None)
            out.append({
                'emp_id': code,
                'name': txt('name'),
                'mobile': txt('mobile'),
                'branch': txt('branch'),
                'role': txt('role'),
                'designation': txt('designation'),
                'scope': txt('scope'),
                'reporting_officer_id': txt('reporting_officer_id'),
                'date_of_joining': _excel_date(cell(row, 'date_of_joining')),
                'date_of_birth': _excel_date(cell(row, 'date_of_birth')),
                'gender': txt('gender'),
            })
        return out
    finally:
        wb.close()


@growwithme_bp.route('/sync-staff', methods=['POST'])
def sync_staff():
    """Upload handler for BOTH staff formats. If the file is an onboarding/access
    sheet (an 'Employees' sheet with a `scope` column) it runs the FULL onboard
    (create employee + login/password + info + branch/role/designation assignment
    + SCOPE). Otherwise it treats the file as the HR 'Working' master and refreshes
    DETAILS ONLY (name/phone/joining/DOB/manager). Upsert; never deletes; idempotent."""
    try:
        up, is_temp = _uploaded_file()
    except ValueError as e:
        return jsonify({'success': False, 'message': str(e)}), 400
    if not up:
        return jsonify({'success': False, 'message': 'Upload a staff / access Excel file.'}), 400
    try:
        # Access/onboarding sheet? (has a scope column) → full onboard.
        onboard = _parse_onboard_sheet(up)
        if onboard is not None:
            if not onboard:
                return jsonify({'success': False, 'message': 'No rows found in the Employees sheet.'}), 400
            ok, res = _post('/api/employees/onboard', {'rows': onboard})
            if not ok:
                logger.warning(f'GrowwithmeDB onboard failed: {res}')
                return jsonify({'success': False, 'message': res}), 502
            res = res or {}
            msg = (f"{res.get('created', 0)} new · {res.get('existed', 0)} updated · "
                   f"{res.get('scoped', 0)} scope set · {res.get('assignments_set', 0)} assignments · "
                   f"{res.get('logins_created', 0)} new logins")
            if res.get('warnings'):
                msg += f" · {len(res['warnings'])} warning(s)"
            logger.info(f'GrowwithmeDB onboard: {len(onboard)} rows → {msg}')
            return jsonify({'success': True, 'onboard_rows': len(onboard), 'message': msg, **res})

        # Otherwise the HR 'Working' master → details-only sync.
        rows = _parse_staff_sheet(up)
        if not rows:
            return jsonify({'success': False, 'message': 'No staff rows found (need a "Working" sheet or an "Employees" access sheet with a scope column).'}), 400
        ok, res = _post('/api/employees/sync-staff', {'rows': rows})
        if not ok:
            logger.warning(f'GrowwithmeDB staff sync failed: {res}')
            return jsonify({'success': False, 'message': res}), 502
        res = res or {}
        msg = (f"{res.get('inserted_employees', 0)} new · {res.get('name_updates', 0)} updated · "
               f"{res.get('contacts', 0)} phones · {res.get('personals', 0)} joining/DOB · "
               f"{res.get('managers_set', 0)} managers")
        logger.info(f'GrowwithmeDB staff sync: {len(rows)} rows → {msg}')
        return jsonify({'success': True, 'staff_rows': len(rows), 'message': msg, **res})
    except Exception as e:
        logger.warning(f'GrowwithmeDB staff/onboard sync: parse failed: {e}')
        return jsonify({'success': False, 'message': f'Upload failed: {e}'}), 500
    finally:
        _cleanup(up, is_temp)


# ── Single-employee editor (fetch + save details/scope) ───────────────────
@growwithme_bp.route('/scope-options', methods=['GET'])
def scope_options():
    """Region/division/area/branch name lists for the scope-target picker."""
    ok, res = _post('/api/employees/scope-options', {})
    if not ok:
        return jsonify({'success': False, 'message': str(res)}), 502
    return jsonify({'success': True, 'options': res or {}})


@growwithme_bp.route('/employee/<code>', methods=['GET'])
def get_employee(code):
    """Fetch one employee's editable details + current scope from GrowwithmeDB.
    Returns found=False (still success) when the code isn't in the DB, so the UI
    can offer to create it — distinct from a real fetch failure (success=False)."""
    ok, res = _post('/api/employees/lookup', {'emp_id': str(code).strip()})
    if not ok:
        return jsonify({'success': False, 'message': str(res)}), 502
    res = res or {}
    if not res.get('found'):
        return jsonify({'success': True, 'found': False, 'emp_id': str(code).strip()})
    return jsonify({'success': True, 'found': True, 'employee': res})


@growwithme_bp.route('/employee', methods=['POST'])
def save_employee():
    """Save one employee's edited details + scope via the full-onboard endpoint."""
    row = request.get_json(silent=True) or {}
    if not str(row.get('emp_id', '')).strip():
        return jsonify({'success': False, 'message': 'emp_id is required'}), 400
    ok, res = _post('/api/employees/onboard', {'rows': [row]})
    if not ok:
        return jsonify({'success': False, 'message': str(res)}), 502
    res = res or {}
    res.pop('passwords', None)  # never surface passwords to the UI
    created = 'created' if res.get('created') else 'updated'
    parts = [f"{created}", f"scope={row.get('scope') or 'self'}"]
    if res.get('warnings'):
        parts.append('; '.join(res['warnings'][:3]))
    return jsonify({'success': True, 'message': ' · '.join(parts), **res})


@growwithme_bp.route('/sync-disbursement', methods=['POST'])
def sync_disbursement():
    """Push a disbursement file into GrowwithmeDB.disbursement (monthly grain).

    Form fields:
      - file:  disbursement CSV/XLSX
      - dates: comma-separated YYYY-MM-DD to keep. Empty/missing -> all dates.
    Daily rows are aggregated to the month (db_month = first-of-month) and sent
    as a single batched POST /api/disbursement/sync {rows:[...]}.
    """
    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'success': False, 'message': 'No file uploaded'}), 400
    if not f.filename.lower().endswith(('.csv', '.xlsx', '.xls')):
        return jsonify({'success': False, 'message': 'Expected .csv / .xlsx file'}), 400

    raw_dates = (request.form.get('dates') or '').strip()
    keep_dates = {d.strip() for d in raw_dates.split(',') if d.strip()} or None

    from blueprints.disbursement import _parse_file, _aggregate, _save_upload
    tmp = _save_upload(f, f.filename)
    try:
        parsed = _parse_file(tmp, f.filename)
        agg, _dates = _aggregate(parsed, keep_dates=keep_dates)
        if not agg:
            return jsonify({'success': False, 'message': (
                'No Active disbursement rows match the selected dates.' if keep_dates
                else 'No Active disbursement rows found in file.')}), 400

        # Per-day rows (grain = disb_date) for the Daily tab — `agg` already holds
        # the daily grain, so the same aggregates feed both pushes below.
        daily_rows = [
            {
                'disb_date': iso,
                'branch_name': branch,
                'emp_id': emp_id or None,
                'product_type_id': _DISB_PRODUCT_TYPE_ID.get(prod, 1),
                'officer_name': v.get('officer_name') or None,
                'disb_count': v['cnt'],
                'disb_amount': round(v['amt'], 2),
            }
            for (iso, branch, emp_id, prod), v in agg.items()
        ]

        # Roll the same aggregates up to the month for the monthly Disbursement tab.
        months = {}
        for (iso, branch, emp_id, prod), v in agg.items():
            db_month = iso[:7] + '-01'
            key = (db_month, branch, emp_id or None, prod)
            m = months.setdefault(key, {'cnt': 0, 'amt': 0.0})
            m['cnt'] += v['cnt']
            m['amt'] += v['amt']

        rows = [
            {
                'branch': branch,
                'emp_id': emp_id,
                'product_type_id': _DISB_PRODUCT_TYPE_ID.get(prod, 1),
                'db_month': db_month,
                'disb_count': m['cnt'],
                'disb_amount': round(m['amt'], 2),
            }
            for (db_month, branch, emp_id, prod), m in months.items()
        ]

        # 1) Monthly grain — drives the Disbursement tab. A failure here aborts.
        ok, res = _post('/api/disbursement/sync', {'rows': rows})
        if not ok:
            logger.warning(f'GrowwithmeDB disbursement sync failed: {res}')
            return jsonify({'success': False, 'message': res}), 502
        monthly_inserted = int((res or {}).get('count') or 0)

        # 2) Daily grain — drives the Disbursement → Daily tab (per-date override on
        #    the API side). Best-effort: a daily failure does not undo the monthly
        #    push but is surfaced in the response message.
        ok_d, res_d = _post('/api/disbursement/sync-daily', {'rows': daily_rows})
        daily_inserted = int((res_d or {}).get('inserted') or 0) if ok_d else 0
        daily_dates = (res_d or {}).get('dates_overridden') or [] if ok_d else []

        msg = f'{monthly_inserted} monthly rows synced to growwithme-local (disbursement)'
        if ok_d:
            msg += f' · {daily_inserted} daily rows across {len(daily_dates)} date(s)'
        else:
            logger.warning(f'GrowwithmeDB disbursement daily sync failed: {res_d}')
            msg += f' · daily push FAILED: {res_d}'

        logger.info(f'GrowwithmeDB disbursement sync: monthly={monthly_inserted} '
                    f'daily={daily_inserted} (daily_ok={ok_d})')
        return jsonify({
            'success': True,
            'inserted': monthly_inserted,
            'daily_inserted': daily_inserted,
            'daily_ok': ok_d,
            'message': msg,
        })
    finally:
        try:
            tmp.unlink()
        except Exception:
            pass
