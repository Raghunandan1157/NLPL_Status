"""
Report parsers (shared)
=======================
Excel parsers for the generated EOD Employee Report and Quick Hourly Report,
plus the product-type id maps. Extracted from the old ``supabase_sync`` blueprint
so the GrowwithmeDB push (``growwithme_sync``) keeps working after the Supabase
sync was removed. No network/Supabase code here — parsing only.
"""

import logging

from openpyxl import load_workbook

import config

logger = logging.getLogger(__name__)

# Sheet (product) name -> product_type_id. VVY is renamed to IL.
_PRODUCT_TYPE_ID = {'IGL': 1, 'FIG': 2, 'IL': 3, 'VVY': 3}

# Disbursement product name -> product_type_id.
_DISB_PRODUCT_TYPE_ID = {'IGL': 1, 'FIG': 2, 'IL': 3}

# Quick Report carries no product split — all rows are combined (id 0).
_QUICK_HOURLY_PT_ID = 0

# (db_column, normalised header text) — order defines positional fallback.
_METRIC_HEADERS = [
    ('regular_demand',         'regular demand'),
    ('regular_collection',     'regular collection'),
    ('demand_1_30',            '1-30 demand'),
    ('collection_1_30',        '1-30 collection'),
    ('demand_31_60',           '31-60 demand'),
    ('collection_31_60',       '31-60 collection'),
    ('pnpa_demand',            'pnpa demand'),
    ('pnpa_collection',        'pnpa collection'),
    ('npa_cases',              'npa cases'),
    ('npa_act_acc',            'npa act acc'),
    ('npa_act_amt',            'npa act amt'),
    ('npa_clo_acc',            'npa clo acc'),
    ('npa_clo_amt',            'npa clo amt'),
    ('on_date_demand',         'on-date demand'),
    ('on_date_collection',     'on-date collection'),
    ('regular_demand_amt',     'regular demand amt'),
    ('regular_collection_amt', 'regular collection amt'),
    ('demand_1_30_amt',        '1-30 demand amt'),
    ('collection_1_30_amt',    '1-30 collection amt'),
    ('demand_31_60_amt',       '31-60 demand amt'),
    ('collection_31_60_amt',   '31-60 collection amt'),
    ('pnpa_demand_amt',        'pnpa demand amt'),
    ('pnpa_collection_amt',    'pnpa collection amt'),
    ('on_date_demand_amt',     'on-date demand amt'),
    ('on_date_collection_amt', 'on-date collection amt'),
]
_METRIC_DB_COLS = [c for c, _ in _METRIC_HEADERS]


def _norm(h):
    return ' '.join(str(h or '').strip().lower().split())


def _num(v):
    if v is None or v == '':
        return 0
    try:
        n = float(v)
    except (TypeError, ValueError):
        return 0
    return n if n == n and n not in (float('inf'), float('-inf')) else 0


def _safe_num(v):
    if v is None or v == '' or v == '-':
        return 0
    try:
        n = float(v)
    except (TypeError, ValueError):
        return 0
    return n if n == n and n not in (float('inf'), float('-inf')) else 0


def _employee_report_path():
    """Locate the latest EOD Employee Report."""
    for name in ('Employee_Report_Latest.xlsx', 'EOD_Report_Latest.xlsx'):
        p = config.BACKEND_DATA_DIR / name
        if p.exists():
            return p
    return None


def _parse_report(path):
    """Return list of row dicts: {emp_id, product_type_id, <25 metrics>}.

    Sheets ending in `_FY` and the `POS` / `EMP_POS` sheets are skipped, so
    only the per-product EOD sheets (IGL / FIG / VVY) are returned.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    out = []
    try:
        for sheet_name in wb.sheetnames:
            if sheet_name.endswith('_FY') or sheet_name in ('POS', 'EMP_POS'):
                continue
            pt_id = _PRODUCT_TYPE_ID.get(sheet_name.strip().upper())
            if pt_id is None:
                logger.info(f"Report parse: skipping unknown sheet '{sheet_name}'")
                continue

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            header = rows[0]
            # emp id column
            emp_idx = 3
            norm_hdr = [_norm(h) for h in header]
            for i, h in enumerate(norm_hdr):
                if h in ('emp id', 'empid', 'emp_id'):
                    emp_idx = i
                    break
            metrics_start = emp_idx + 2

            # header-name -> column index (first match wins)
            hdr_idx = {}
            for i, h in enumerate(norm_hdr):
                if h and h not in hdr_idx:
                    hdr_idx[h] = i
            col_map = {db: hdr_idx[txt] for db, txt in _METRIC_HEADERS if txt in hdr_idx}
            use_header_map = len(col_map) >= 20

            for row in rows[1:]:
                if not row or len(row) <= emp_idx:
                    continue
                if not row[0] or not row[emp_idx]:
                    continue
                emp_id = str(row[emp_idx]).strip()
                if not emp_id:
                    continue
                rec = {'emp_id': emp_id, 'product_type_id': pt_id}
                for i, db_col in enumerate(_METRIC_DB_COLS):
                    if use_header_map and db_col in col_map:
                        ci = col_map[db_col]
                    else:
                        ci = metrics_start + i
                    rec[db_col] = _num(row[ci]) if ci < len(row) else 0
                out.append(rec)
    finally:
        wb.close()
    return out


def _quick_report_path():
    """Locate the latest hourly report to sync — the Hourly module's report or the
    Quick module's report (same format). Prefer the newest file that actually has a
    parseable per-employee table ('Employee Data' sheet, else an OverAll officer
    section); fall back to the newest existing file."""
    candidates = [
        config.BACKEND_DATA_DIR / 'Hourly_Report_Latest.xlsx',
        config.BACKEND_DATA_DIR / 'Hourly_Fast_Report_Latest.xlsx',
        config.BACKEND_DATA_DIR / 'Quick_Report_Latest.xlsx',
    ]
    existing = [p for p in candidates if p.exists()]
    if not existing:
        return None
    existing.sort(key=lambda p: p.stat().st_mtime, reverse=True)  # newest first

    def _parseable(p):
        try:
            wb = load_workbook(p, read_only=True)
            try:
                if 'Employee Data' in wb.sheetnames:
                    return True
                if 'OverAll' in wb.sheetnames:
                    for row in wb['OverAll'].iter_rows(values_only=True):
                        if row and row[0] is not None and str(row[0]).strip() == 'EMP ID':
                            return True
                return False
            finally:
                wb.close()
        except Exception:
            return False

    return next((p for p in existing if _parseable(p)), existing[0])


def _parse_employee_data_sheet(ws):
    """Parse the 'Employee Data' sheet of the newer Hourly Report into hourly row
    dicts (product_type_id = 0) — same shape as _parse_quick_report.

    This report's 'OverAll' is region-wise; the per-employee data lives here in
    grouped blocks. Column layout (0-indexed): EMP ID = 1; Regular D/C = 7/8;
    1-30 D/C = 11/12; 31-60 D/C = 15/16; PNPA D/C = 19/20; NPA cases = 27;
    NPA activation acc/amt = 28/29; closure acc/amt = 30/31. On-date is region-wise
    in this report (not per-employee), so it is 0 here.
    """
    rows = list(ws.iter_rows(values_only=True))
    hdr_idx, emp_col = -1, 1
    for r, row in enumerate(rows):
        if not row:
            continue
        for i, c in enumerate(row):
            if c is not None and str(c).strip().upper() == 'EMP ID':
                hdr_idx, emp_col = r, i
                break
        if hdr_idx >= 0:
            break
    if hdr_idx < 0:
        raise ValueError("Employee Data sheet has no 'EMP ID' header")

    def g(row, i):
        return _safe_num(row[i]) if i < len(row) else 0

    out = []
    for row in rows[hdr_idx + 1:]:
        if not row or len(row) <= emp_col or row[emp_col] is None:
            continue
        emp = str(row[emp_col]).strip()
        if not emp or emp.upper() in ('EMP ID', 'GRAND TOTAL', 'TOTAL'):
            continue
        out.append({
            'emp_id': emp,
            'product_type_id': _QUICK_HOURLY_PT_ID,
            'regular_demand': g(row, 7), 'regular_collection': g(row, 8),
            'demand_1_30': g(row, 11), 'collection_1_30': g(row, 12),
            'demand_31_60': g(row, 15), 'collection_31_60': g(row, 16),
            'pnpa_demand': g(row, 19), 'pnpa_collection': g(row, 20),
            'npa_cases': g(row, 27),
            'npa_act_acc': g(row, 28), 'npa_act_amt': g(row, 29),
            'npa_clo_acc': g(row, 30), 'npa_clo_amt': g(row, 31),
            'on_date_demand': 0, 'on_date_collection': 0,
            'regular_demand_amt': 0, 'regular_collection_amt': 0,
            'demand_1_30_amt': 0, 'collection_1_30_amt': 0,
            'demand_31_60_amt': 0, 'collection_31_60_amt': 0,
            'pnpa_demand_amt': 0, 'pnpa_collection_amt': 0,
            'on_date_demand_amt': 0, 'on_date_collection_amt': 0,
        })
    return out


def _parse_quick_report(path):
    """Parse the hourly report into row dicts (product_type_id = 0).

    Supports two layouts:
      1. Legacy Quick Report — officer section inside the 'OverAll' sheet.
      2. Newer Hourly Report — per-employee data in an 'Employee Data' sheet
         (its 'OverAll' is region-wise).

    Mirrors the Coll_Db /api/upload-hourly Quick Report parser:
      - OverAll sheet Branch+Officer section (starts 4 rows after 'EMP ID')
      - OverAll_On-Date sheet supplies on_date_demand / on_date_collection
      - column layout (0-indexed): 0=EMP ID 2=Reg Demand 3=Reg Collection
        6=1-30 Demand 7=1-30 Collection 10=31-60 Demand 11=31-60 Collection
        14=PNPA Demand 15=PNPA Collection 22=NPA Cases 23=NPA Act Account
        24=NPA Act Amount.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        # Prefer the authoritative per-employee 'Employee Data' sheet (newer Hourly
        # Report — clean grouped columns, one row per officer). The legacy Quick
        # Report has no such sheet and falls through to the OverAll officer section.
        # (The Hourly Report's OverAll officer section is a partial/region breakdown
        # with different totals, so it must NOT be used when Employee Data exists.)
        if 'Employee Data' in wb.sheetnames:
            return _parse_employee_data_sheet(wb['Employee Data'])
        if 'OverAll' not in wb.sheetnames:
            raise ValueError("Unsupported hourly file: no 'Employee Data' or 'OverAll' sheet.")
        over_rows = list(wb['OverAll'].iter_rows(values_only=True))

        officer_start = -1
        for r, row in enumerate(over_rows):
            if row and row[0] is not None and str(row[0]).strip() == 'EMP ID':
                officer_start = r
                break
        if officer_start < 0:
            # Newer Hourly Report: 'OverAll' is region-wise; per-employee data is in
            # a separate 'Employee Data' sheet. Parse that instead.
            if 'Employee Data' in wb.sheetnames:
                return _parse_employee_data_sheet(wb['Employee Data'])
            raise ValueError(
                "Unsupported hourly file: 'OverAll' has no 'EMP ID' officer section "
                "and there is no 'Employee Data' sheet.")
        data_start = officer_start + 4

        # On-Date sheet → {emp_id: {demand, collection}}
        on_date = {}
        if 'OverAll_On-Date' in wb.sheetnames:
            od_rows = list(wb['OverAll_On-Date'].iter_rows(values_only=True))
            od_start = -1
            for r, row in enumerate(od_rows):
                if row and row[0] is not None and str(row[0]).strip() == 'EMP ID':
                    od_start = r
                    break
            if od_start >= 0:
                for row in od_rows[od_start + 4:]:
                    if not row or not row[0]:
                        continue
                    emp = str(row[0]).strip()
                    if not emp or emp == 'EMP ID':
                        continue
                    on_date[emp] = {
                        'demand': _safe_num(row[2] if len(row) > 2 else 0),
                        'collection': _safe_num(row[3] if len(row) > 3 else 0),
                    }

        out = []
        for row in over_rows[data_start:]:
            if not row:
                continue
            col0 = str(row[0]).strip() if row[0] is not None else ''
            # Skip branch rows (blank col0), Grand Total, nested headers.
            if not col0 or col0 in ('Grand Total', 'EMP ID'):
                continue
            emp = col0
            od = on_date.get(emp, {'demand': 0, 'collection': 0})

            def g(i):
                return _safe_num(row[i]) if len(row) > i else 0

            out.append({
                'emp_id': emp,
                'product_type_id': _QUICK_HOURLY_PT_ID,
                'regular_demand': g(2),
                'regular_collection': g(3),
                'demand_1_30': g(6),
                'collection_1_30': g(7),
                'demand_31_60': g(10),
                'collection_31_60': g(11),
                'pnpa_demand': g(14),
                'pnpa_collection': g(15),
                'npa_cases': g(22),
                'npa_act_acc': g(23),
                'npa_act_amt': g(24),
                'npa_clo_acc': 0,
                'npa_clo_amt': 0,
                'on_date_demand': od['demand'],
                'on_date_collection': od['collection'],
                'regular_demand_amt': 0, 'regular_collection_amt': 0,
                'demand_1_30_amt': 0, 'collection_1_30_amt': 0,
                'demand_31_60_amt': 0, 'collection_31_60_amt': 0,
                'pnpa_demand_amt': 0, 'pnpa_collection_amt': 0,
                'on_date_demand_amt': 0, 'on_date_collection_amt': 0,
            })
        return out
    finally:
        wb.close()
